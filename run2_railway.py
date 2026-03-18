# app.py | CRM Firmantes - Flask + SQLite + Cartera XLSX (openpyxl) | Fusion: Firmantes + Precalificados => "Firmantes Precalificados" | - ABM firmantes (manual + CSV) desde la misma solapa | - Cálculo usado/disponible por cartera.xlsx + bloqueos diarios | - Orden por disponible (negativos arriba) y selector de "scope" (propio/3ros/peor) | - Vencimiento de límites (propio y 3ros) => al vencer mantienen límite, pasan a inactivos y se audita como VENCIMIENTO / RENOVACION | - Auditoría ABM: ALTAS, BAJAS, MODIFICACIONES, VENCIMIENTOS, RENOVACIONES | - Roles: admin / risk / sales (sales no ABM límites)

import ssl

# ===== FORZAR DESACTIVACIÓN GLOBAL DE VERIFICACIÓN SSL =====
ssl._create_default_https_context = ssl._create_unverified_context
# =============================================================

import os
import sys
import re
import io
import csv
import sqlite3
import secrets
import json
import uuid
import hashlib
import urllib.request
import urllib.error
import subprocess

import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from functools import wraps

from flask import Flask, request, session, redirect, url_for, render_template_string, abort, send_file
from werkzeug.routing import BuildError

from werkzeug.security import generate_password_hash, check_password_hash

from zoneinfo import ZoneInfo

ARG_TZ = ZoneInfo("America/Argentina/Buenos_Aires")

try:
    import openpyxl
except Exception:
    openpyxl = None

# --------------------------- | Config | ---------------------------

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, "crm.db")
CARTERA_XLSX = os.path.join(APP_DIR, "cartera.xlsx")
FACTURAS_XLSX = os.path.join(APP_DIR, "Facturas.xlsx")
SECRET_PATH = os.path.join(APP_DIR, ".secret_key")

ROLE_ADMIN = "admin"
ROLE_RISK = "risk"
ROLE_SALES = "sales"

ROLE_LABELS = {
    ROLE_ADMIN: "Admin",
    ROLE_RISK: "Analista de riesgos",
    ROLE_SALES: "Ejecutivo comercial",
}

TABLE_FIRMANTES = "firmantes"

TABLE_GROUPS = "grupos_economicos"
TABLE_IMPORT_BATCHES = "import_batches"

# --------------------------- | Alertas diarias - Cheques rechazados (SIN FONDOS) | ---------------------------

SLACK_WEBHOOK_URL = (
    os.environ.get("CRM_SLACK_WEBHOOK_URL")
    or os.environ.get("SLACK_WEBHOOK_URL")
    or ""
)

def send_slack_message(text_msg: str) -> bool:
    if not SLACK_WEBHOOK_URL:
        return False

    payload = {"text": text_msg}
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")

    req = urllib.request.Request(
        SLACK_WEBHOOK_URL,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            _ = resp.read()
        return True
    except Exception as e:
        print(f"[slack] error enviando mensaje: {e}")
        return False

BCRA_CHEQ_ENDPOINT = "https://api.bcra.gob.ar/centraldedeudores/v1.0/Deudas/ChequesRechazados"

TABLE_CHEQ_SEEN = "cheq_seen_events"
TABLE_CHEQ_RUNS = "cheq_monitor_runs"
TABLE_CHEQ_LOCK = "cheq_monitor_lock"
TABLE_CHEQ_RAW = "cheq_events_raw"

TABLE_CHEQ_PROGRESS = "cheq_monitor_progress"

def ensure_secret_key():
    if os.path.exists(SECRET_PATH):
        try:
            with open(SECRET_PATH, "r", encoding="utf-8") as f:
                key = f.read().strip()
                if key:
                    return key
        except Exception:
            pass
    key = os.environ.get("CRM_SECRET_KEY") or secrets.token_hex(32)
    try:
        with open(SECRET_PATH, "w", encoding="utf-8") as f:
            f.write(key)
    except Exception:
        pass
    return key

app = Flask(__name__)
app.secret_key = ensure_secret_key()
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
# Si lo ponés detrás de HTTPS, podés activar: | SESSION_COOKIE_SECURE=True,
)

# --------------------------- | Utils | ---------------------------

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def today_local_date():
    return datetime.now().strftime("%Y-%m-%d")

def cuit_digits(s: str) -> str:
    if not s:
        return ""
    return "".join(ch for ch in str(s) if ch.isdigit())

def cuit_digits_excel(s) -> str:
    """Normalize CUITs coming from Excel cells.
    Handles ints, floats like 30657314974.0 and string/scientific-looking values.
    Returns only digits, preferring 11-digit CUIT when possible.
    """
    if s is None:
        return ""
    try:
        if isinstance(s, bool):
            return ""
        if isinstance(s, int):
            txt = str(s)
            return txt if txt.isdigit() else cuit_digits(txt)
        if isinstance(s, float):
            if s != s:
                return ""
            if abs(s - int(s)) < 1e-9:
                txt = str(int(s))
                return txt if txt.isdigit() else cuit_digits(txt)
    except Exception:
        pass
    txt = str(s).strip()
    if not txt:
        return ""
    m = re.match(r'^([0-9]{11})[\.,]0+$', txt)
    if m:
        return m.group(1)
    digs = ''.join(ch for ch in txt if ch.isdigit())
    if len(digs) == 12 and digs.endswith('0'):
        return digs[:-1]
    return digs

def normalize_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u").replace("ü", "u").replace("ñ", "n")
    s = s.replace(".", "").replace("-", " ").replace("_", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s

def parse_bool_si(x) -> bool:
    if x is None:
        return False
    s = normalize_header(str(x))
    return s in ("si", "s", "1", "true", "x", "yes")

def parse_es_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace("$", "").replace("ARS", "").replace("ar$", "").replace(" ", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    if "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        if s.count(".") > 1:
            s = s.replace(".", "")
    try:
        v = float(s)
        return -v if neg else v
    except Exception:
        return None

def fmt_ars(v):
    try:
        x = float(v)
    except Exception:
        x = 0.0
    neg = x < 0
    x = abs(x)
    s = f"{x:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"$ {'-' if neg else ''}{s}"

def parse_date_any(s):
    """
    Acepta:
      - YYYY-MM-DD
      - DD/MM/YYYY
      - DD-MM-YYYY
    Devuelve YYYY-MM-DD o "" si vacío/invalid.
    """
    if s is None:
        return ""
    t = str(s).strip()
    if not t:
        return ""
# ya ISO
    m = re.match(r"^\d{4}-\d{2}-\d{2}$", t)
    if m:
        return t
# dd/mm/yyyy o dd-mm-yyyy
    m = re.match(r"^(\d{2})[\/\-](\d{2})[\/\-](\d{4})$", t)
    if m:
        dd, mm, yy = m.group(1), m.group(2), m.group(3)
        try:
            dt = datetime(int(yy), int(mm), int(dd))
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return ""
    return ""

def line_is_expired(expiry_value: str, today: str | None = None) -> bool:
    exp = (expiry_value or "").strip()
    if not exp:
        return False
    today = today or today_local_date()
    return exp <= today

def effective_line_active(raw_active, expiry_value: str, today: str | None = None) -> bool:
    today = today or today_local_date()
    active = int(raw_active or 0) == 1
    return bool(active and not line_is_expired(expiry_value, today=today))

def line_status_label(raw_active, expiry_value: str, today: str | None = None) -> str:
    today = today or today_local_date()
    if line_is_expired(expiry_value, today=today):
        return "Inactiva por vencimiento"
    return "Activa" if int(raw_active or 0) == 1 else "Inactiva"

# --------------------------- | DB | ---------------------------

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
# Mejor convivencia Flask + worker (evita freezes por locks)
    try:
        conn.execute('PRAGMA journal_mode=WAL;')
        conn.execute('PRAGMA synchronous=NORMAL;')
        conn.execute('PRAGMA busy_timeout=5000;')
    except Exception:
        pass

    conn.execute("PRAGMA foreign_keys=ON;")

# anti-lock (recomendado para multiusuario chico)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA busy_timeout=5000;")

    return conn

def table_columns(conn, table):
    return {r["name"] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()}

def init_db():
    conn = get_db()
    cur = conn.cursor()

# Users
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            pass_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_users_active ON users(is_active)")

# Firmantes
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_FIRMANTES} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            razon_social TEXT NOT NULL,
            cuit TEXT NOT NULL,
            cuit_digits TEXT NOT NULL UNIQUE,

            credit_limit_3ros REAL NOT NULL DEFAULT 0,
            credit_limit_propio REAL NOT NULL DEFAULT 0,
            credit_limit_fce REAL NOT NULL DEFAULT 0,

            limit_expiry_3ros TEXT NOT NULL DEFAULT '',
            limit_expiry_propio TEXT NOT NULL DEFAULT '',
            line_active_3ros INTEGER NOT NULL DEFAULT 1,
            line_active_propio INTEGER NOT NULL DEFAULT 1,

            is_active INTEGER NOT NULL DEFAULT 1,

            grupo_id INTEGER,
            primera_linea INTEGER NOT NULL DEFAULT 0,

            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)
# Grupos económicos
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_GROUPS} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            credit_limit_grupal REAL NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_grupos_nombre ON {TABLE_GROUPS}(nombre)")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_grupos_active ON {TABLE_GROUPS}(is_active)")

# Staging para importaciones (CSV firmantes con grupos no registrados)
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_IMPORT_BATCHES} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            batch_id TEXT NOT NULL UNIQUE,
            username TEXT NOT NULL,
            payload_json TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_batches_user ON {TABLE_IMPORT_BATCHES}(username)")

# MIGRACIONES (si venías de DB vieja)
    cols = table_columns(conn, TABLE_FIRMANTES)

    def add_col(coldef):
        col = coldef.split()[0]
        if col not in cols:
            try:
                conn.execute(f"ALTER TABLE {TABLE_FIRMANTES} ADD COLUMN {coldef}")
            except Exception:
                pass

    add_col("credit_limit_3ros REAL NOT NULL DEFAULT 0")
    add_col("credit_limit_propio REAL NOT NULL DEFAULT 0")
    add_col("credit_limit_fce REAL NOT NULL DEFAULT 0")
    add_col("limit_expiry_3ros TEXT NOT NULL DEFAULT ''")
    add_col("limit_expiry_propio TEXT NOT NULL DEFAULT ''")
    add_col("line_active_3ros INTEGER NOT NULL DEFAULT 1")
    add_col("line_active_propio INTEGER NOT NULL DEFAULT 1")
    add_col("is_active INTEGER NOT NULL DEFAULT 1")
    add_col("created_at TEXT NOT NULL DEFAULT ''")
    add_col("updated_at TEXT NOT NULL DEFAULT ''")
    add_col("cuit_digits TEXT NOT NULL DEFAULT ''")
    add_col("grupo_id INTEGER")
    add_col("primera_linea INTEGER NOT NULL DEFAULT 0")

    cols = table_columns(conn, TABLE_FIRMANTES)

# backfill cuit_digits si quedó vacío
    rows0 = conn.execute(f"SELECT id, cuit, cuit_digits FROM {TABLE_FIRMANTES}").fetchall()
    for rr in rows0:
        if not (rr["cuit_digits"] or "").strip():
            cd = cuit_digits(rr["cuit"] or "")
            conn.execute(f"UPDATE {TABLE_FIRMANTES} SET cuit_digits=? WHERE id=?", (cd, rr["id"]))

# backfill timestamps si quedaron vacíos
    now = now_str()
    conn.execute(f"UPDATE {TABLE_FIRMANTES} SET created_at=? WHERE created_at=''", (now,))
    conn.execute(f"UPDATE {TABLE_FIRMANTES} SET updated_at=? WHERE updated_at=''", (now,))
    today = today_local_date()
    if 'line_active_3ros' in cols and 'limit_expiry_3ros' in cols:
        conn.execute(f"UPDATE {TABLE_FIRMANTES} SET line_active_3ros=0 WHERE (limit_expiry_3ros <> '' AND limit_expiry_3ros <= ?)", (today,))
        conn.execute(f"UPDATE {TABLE_FIRMANTES} SET line_active_3ros=1 WHERE (limit_expiry_3ros = '' OR limit_expiry_3ros > ?) AND COALESCE(line_active_3ros, 1)=0", (today,))
    if 'line_active_propio' in cols and 'limit_expiry_propio' in cols:
        conn.execute(f"UPDATE {TABLE_FIRMANTES} SET line_active_propio=0 WHERE (limit_expiry_propio <> '' AND limit_expiry_propio <= ?)", (today,))
        conn.execute(f"UPDATE {TABLE_FIRMANTES} SET line_active_propio=1 WHERE (limit_expiry_propio = '' OR limit_expiry_propio > ?) AND COALESCE(line_active_propio, 1)=0", (today,))
    conn.commit()

# Índices firmantes (sólo si existen cols)
    cols = table_columns(conn, TABLE_FIRMANTES)
    if "razon_social" in cols:
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_firmantes_rs ON {TABLE_FIRMANTES}(razon_social)")
    if "is_active" in cols:
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_firmantes_active ON {TABLE_FIRMANTES}(is_active)")
    if "cuit_digits" in cols:
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_firmantes_cuitdigits ON {TABLE_FIRMANTES}(cuit_digits)")
    if "grupo_id" in cols:
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_firmantes_grupo ON {TABLE_FIRMANTES}(grupo_id)")
    if "primera_linea" in cols:
        conn.execute(f"CREATE INDEX IF NOT EXISTS idx_firmantes_primera_linea ON {TABLE_FIRMANTES}(primera_linea)")

# Auditoría ABM
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firmante_cuit_digits TEXT NOT NULL,
            action TEXT NOT NULL,         -- ALTA / BAJA / MODIFICACION / VENCIMIENTO
            username TEXT NOT NULL,
            details TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_firmante ON audit_log(firmante_cuit_digits)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_log(created_at)")

# Bloqueos diarios
    cur.execute("""
        CREATE TABLE IF NOT EXISTS limit_blocks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            firmante_cuit_digits TEXT NOT NULL,
            scope TEXT NOT NULL,            -- 'propio' o '3ros'
            username TEXT NOT NULL,
            amount REAL NOT NULL,
            block_date TEXT NOT NULL,       -- YYYY-MM-DD
            created_at TEXT NOT NULL
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_blocks_firmante_date ON limit_blocks(firmante_cuit_digits, block_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_blocks_user_date ON limit_blocks(username, block_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_blocks_scope_date ON limit_blocks(scope, block_date)")

# Alertas diarias - tablas
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_CHEQ_SEEN} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cuit_digits TEXT NOT NULL,
            event_id TEXT NOT NULL UNIQUE,
            entidad INTEGER,
            nro_cheque TEXT,
            fecha_rechazo TEXT,
            monto REAL,
            pagado INTEGER NOT NULL DEFAULT 0,
            detected_at TEXT NOT NULL
        )
    """)
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_cheq_seen_cuit ON {TABLE_CHEQ_SEEN}(cuit_digits)")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_cheq_seen_fecha ON {TABLE_CHEQ_SEEN}(fecha_rechazo)")

# Tabla raw para guardar TODOS los eventos de cheques (todas las causales) tal como vienen de BCRA. | Se usa para auditoría / historial completo. No reemplaza cheq_seen_events (subset SIN FONDOS usado por alertas).
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_CHEQ_RAW} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cuit_digits TEXT NOT NULL,
            event_id TEXT NOT NULL UNIQUE,
            causal TEXT,
            entidad INTEGER,
            payload_json TEXT NOT NULL,
            detected_at TEXT NOT NULL
        )
    """)
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_cheq_raw_cuit ON {TABLE_CHEQ_RAW}(cuit_digits)")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_cheq_raw_causal ON {TABLE_CHEQ_RAW}(causal)")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_cheq_raw_detected ON {TABLE_CHEQ_RAW}(detected_at)")

    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_CHEQ_RUNS} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_type TEXT NOT NULL,
            scope TEXT NOT NULL,
            scope_cuit TEXT,
            started_at TEXT NOT NULL,
            ended_at TEXT,
            total_cuits INTEGER NOT NULL DEFAULT 0,
            new_events INTEGER NOT NULL DEFAULT 0,
            alerts_sent INTEGER NOT NULL DEFAULT 0,
            errors INTEGER NOT NULL DEFAULT 0,
            notes TEXT
        )
    """)
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_cheq_runs_started ON {TABLE_CHEQ_RUNS}(started_at)")

    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_CHEQ_LOCK} (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            locked_at TEXT,
            locked_by TEXT,
            expires_at TEXT
        )
    """)
    cur.execute(f"INSERT OR IGNORE INTO {TABLE_CHEQ_LOCK}(id, locked_at, locked_by, expires_at) VALUES (1, NULL, NULL, NULL)")

# Progreso (sync/monitor)
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_CHEQ_PROGRESS} (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            run_id TEXT,
            mode TEXT,
            status TEXT,
            total INTEGER NOT NULL DEFAULT 0,
            done INTEGER NOT NULL DEFAULT 0,
            errors INTEGER NOT NULL DEFAULT 0,
            started_at TEXT,
            updated_at TEXT,
            message TEXT,
            result_json TEXT
        )
    """)
    cur.execute(f"INSERT OR IGNORE INTO {TABLE_CHEQ_PROGRESS}(id, status, total, done, errors) VALUES (1, 'idle', 0, 0, 0)")

    conn.commit()

    conn.close()

def audit_log(cd: str, action: str, username: str, details: str, conn=None):
    own_conn = conn is None
    if own_conn:
        conn = get_db()
    try:
        conn.execute("""
            INSERT INTO audit_log (firmante_cuit_digits, action, username, details, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (cd, action, username, details, now_str()))
        if own_conn:
            conn.commit()
    finally:
        if own_conn:
            conn.close()

def append_observation(details: str, observacion: str | None) -> str:
    obs = (observacion or "").strip()
    base = (details or "").strip()
    if not obs:
        return base or "Sin cambios"
    if not base:
        return f"Observación: {obs}"
    return f"{base} | Observación: {obs}"

# --------------------------- | Auth & Roles | ---------------------------

def current_user():
    return session.get("username")

def current_role():
    return (session.get("role") or "").lower()

def is_admin():
    return current_role() == ROLE_ADMIN

def is_risk():
    return current_role() == ROLE_RISK

def is_sales():
    return current_role() == ROLE_SALES

def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper

def roles_required(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                return redirect(url_for("login", next=request.path))
            if current_role() not in roles:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco

# --------------------------- | Flash | ---------------------------

def set_flash(msg, kind="ok"):
    session["_flash"] = {"msg": msg, "kind": kind}

def pop_flash():
    return session.pop("_flash", None)

# Helpers de compatibilidad (atajos)
def flash_ok(msg: str):
    set_flash(msg, "ok")

def flash_err(msg: str):
    set_flash(msg, "err")

def pull_flash():
# alias histórico
    return pop_flash()

# --------------------------- | Cartera cache + agregación | ---------------------------

_cartera_cache = {
    "mtime": None,
    "headers": [],
    "rows": [],
    "agg": {},          # {cuit_digits: {"3ros": sum, "propio": sum}}
    "status": "OK",
}

_facturas_cache = {
    "mtime": None,
    "headers": [],
    "rows": [],
    "agg": {},          # {cuit_digits: sum_importe}
    "status": "OK",
}

def load_cartera():
    global _cartera_cache

    if not os.path.exists(CARTERA_XLSX):
        return [], [], {}, "No se encontró cartera.xlsx en la carpeta del app.py."
    if openpyxl is None:
        return [], [], {}, "Falta openpyxl. Instalá: pip install openpyxl"

    mtime = os.path.getmtime(CARTERA_XLSX)
    if _cartera_cache["mtime"] == mtime:
        return _cartera_cache["headers"], _cartera_cache["rows"], _cartera_cache["agg"], _cartera_cache["status"]

    try:
        wb = openpyxl.load_workbook(CARTERA_XLSX, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return [], [], {}, "cartera.xlsx está vacío."

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        headers_norm = [normalize_header(h) for h in headers]

        def find_first_col(predicate):
            for i, h in enumerate(headers_norm):
                if predicate(h):
                    return i
            return None

        idx_cuit = find_first_col(lambda h: h.startswith("cuit") and ("2" not in h))
        idx_cuit2 = find_first_col(lambda h: "cuit" in h and ("2" in h))
        idx_importe = find_first_col(lambda h: ("importe" in h) or ("monto" in h))

        rows = []
        agg = {}

        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None:
                continue
            if all(v is None or str(v).strip() == "" for v in r):
                continue

            row = list(r)
            if len(row) < len(headers):
                row += [None] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            rows.append(row)

            if idx_cuit is None or idx_importe is None:
                continue

            c1 = cuit_digits(row[idx_cuit])
            if not c1:
                continue
            c2 = cuit_digits(row[idx_cuit2]) if idx_cuit2 is not None else ""
            imp = parse_es_number(row[idx_importe])
            if imp is None:
                imp = 0.0

            agg.setdefault(c1, {"3ros": 0.0, "propio": 0.0})
            if c2 and c2 != c1:
                agg[c1]["3ros"] += float(imp)
            else:
                agg[c1]["propio"] += float(imp)

        wb.close()

        status = "OK"
        if idx_cuit is None or idx_importe is None:
            status = "cartera.xlsx: faltan columnas obligatorias (CUIT y Importe/Monto)."

        _cartera_cache = {
            "mtime": mtime,
            "headers": headers,
            "rows": rows,
            "agg": agg,
            "status": status,
        }
        return headers, rows, agg, status

    except Exception as e:
        return [], [], {}, f"Error leyendo cartera.xlsx: {e}"

def load_facturas():
    global _facturas_cache

    if not os.path.exists(FACTURAS_XLSX):
        return [], [], {}, "No se encontró Facturas.xlsx en la carpeta del app.py."
    if openpyxl is None:
        return [], [], {}, "Falta openpyxl. Instalá: pip install openpyxl"

    mtime = os.path.getmtime(FACTURAS_XLSX)
    if _facturas_cache["mtime"] == mtime:
        return _facturas_cache["headers"], _facturas_cache["rows"], _facturas_cache["agg"], _facturas_cache["status"]

    try:
        wb = openpyxl.load_workbook(FACTURAS_XLSX, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return [], [], {}, "Facturas.xlsx está vacío."

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        headers_norm = [normalize_header(h) for h in headers]

        def find_first_col(*cands):
            candset = {normalize_header(c) for c in cands}
            for i, h in enumerate(headers_norm):
                if h in candset:
                    return i
            return None

        idx_firmante = find_first_col("firmante", "cuit", "cuit firmante")
        idx_importe = find_first_col("importe", "monto", "importe real")

        rows = []
        agg = {}

        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None:
                continue
            if all(v is None or str(v).strip() == "" for v in r):
                continue

            row = list(r)
            if len(row) < len(headers):
                row += [None] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            rows.append(row)

            if idx_firmante is None or idx_importe is None:
                continue

            cd = cuit_digits_excel(row[idx_firmante])
            if not cd:
                continue

            imp = parse_es_number(row[idx_importe])
            if imp is None:
                imp = 0.0

            agg[cd] = float(agg.get(cd, 0.0)) + float(imp)

        wb.close()

        status = "OK"
        if idx_firmante is None or idx_importe is None:
            status = "Facturas.xlsx: faltan columnas obligatorias (Firmante y Importe/Monto)."

        _facturas_cache.update({
            "mtime": mtime,
            "headers": headers,
            "rows": rows,
            "agg": agg,
            "status": status,
        })
        return headers, rows, agg, status

    except Exception as e:
        return [], [], {}, f"Error leyendo Facturas.xlsx: {e}"

def reorder_facturas_for_ui(headers, rows):
    """Select and reorder Facturas.xlsx columns for UI.
    Output order:
    CUIT(Firmante), Razón, Fe Pago, Liquidacion, letra, prefijo, numero, Moneda, Importe
    """
    if not headers:
        return [], []
    headers_norm = [normalize_header(h) for h in headers]

    wanted = [
        ("Firmante", "CUIT", ["firmante"]),
        ("Razon", "Razón", ["razon", "razón"]),
        ("Fe Pago", "Fe Pago", ["fe pago", "fecha pago", "fe_pago"]),
        ("Liquidacion", "Liquidacion", ["liquidacion", "liquidación"]),
        ("letra", "letra", ["letra"]),
        ("prefijo", "prefijo", ["prefijo"]),
        ("numero", "numero", ["numero", "número"]),
        ("Moneda", "Moneda", ["moneda"]),
        ("Importe", "Importe", ["importe", "monto"]),
    ]

    idxs = []
    out_headers = []
    for _, out_name, aliases in wanted:
        idx = None
        alias_set = {normalize_header(a) for a in aliases}
        for i, h in enumerate(headers_norm):
            if h in alias_set:
                idx = i
                break
        if idx is not None:
            idxs.append(idx)
            out_headers.append(out_name)

    out_rows = []
    for r in rows or []:
        rr = list(r)
        out_rows.append([rr[i] if i < len(rr) else "" for i in idxs])

    return out_headers, out_rows

def _cheq_event_id(cuit_digits_str: str, e: dict) -> str:
    cd = (cuit_digits_str or "").strip()
    entidad = str(e.get("entidad") or "").strip()
    nro = str(e.get("nro_cheque") or "").strip()
    fr = str(e.get("fecha_rechazo") or "").strip()[:10]

    try:
        monto = float(e.get("monto") or 0.0)
    except Exception:
        monto = 0.0

    payload = f"{cd}|{entidad}|{nro}|{fr}|{monto:.2f}"

    return hashlib.sha1(payload.encode("utf-8", errors="ignore")).hexdigest()

def _cheq_raw_event_id(cuit_digits_str: str, causal: str, entidad, detalle_raw: dict) -> str:
    """
    Stable unique id for ANY cheque event (all causales).
    """
    cd = (cuit_digits_str or "").strip()
    csl = (causal or "").strip().upper()
    ent = "" if entidad is None else str(entidad).strip()

    d = detalle_raw or {}
    nro = str(d.get("nroCheque") or "").strip()
    fr = str(d.get("fechaRechazo") or "").strip()[:10]
    try:
        monto = float(d.get("monto") or 0.0)
    except Exception:
        monto = 0.0
    fp = str(d.get("fechaPago") or "").strip()[:10]
    fpm = str(d.get("fechaPagoMulta") or "").strip()[:10]
    em = str(d.get("estadoMulta") or "").strip()
    flags = "|".join([
        "1" if d.get("ctaPersonal") else "0",
        "1" if d.get("enRevision") else "0",
        "1" if d.get("procesoJud") else "0",
    ])

    payload = f"{cd}|{csl}|{ent}|{nro}|{fr}|{monto:.2f}|{fp}|{fpm}|{em}|{flags}"
    return hashlib.sha1(payload.encode("utf-8", errors="ignore")).hexdigest()

def _precalificado_map_from_db():
    """
    Devuelve {cuit_digits: bool} para firmantes precalificados.
    Se considera precalificado si tiene límite propio o de 3ros > 0.
    """
    out = {}
    conn = None
    try:
        conn = get_db()
        rows = conn.execute(f"""
            SELECT cuit_digits, credit_limit_3ros, credit_limit_propio, credit_limit_fce,
                   line_active_3ros, line_active_propio,
                   limit_expiry_3ros, limit_expiry_propio, is_active
            FROM {TABLE_FIRMANTES}
        """).fetchall()
        for r in rows:
            cd = (r["cuit_digits"] or "").strip()
            if not cd:
                continue
            try:
                lim3 = float(r["credit_limit_3ros"] or 0.0)
            except Exception:
                lim3 = 0.0
            try:
                limp = float(r["credit_limit_propio"] or 0.0)
            except Exception:
                limp = 0.0
            try:
                limf = float(r["credit_limit_fce"] or 0.0)
            except Exception:
                limf = 0.0
# Por defecto monitoreamos solo firmantes activos del ABM
            active = int(r["is_active"] or 0) == 1
            line3_active = effective_line_active(r["line_active_3ros"], r["limit_expiry_3ros"] or "")
            linep_active = effective_line_active(r["line_active_propio"], r["limit_expiry_propio"] or "")
            out[cd] = bool(active and ((lim3 > 0 and line3_active) or (limp > 0 and linep_active) or limf > 0))
    except Exception as e:
        print(f"[alerts] error leyendo precalificados: {e}")
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return out

def _cartera_al_dia_map():
    """
    Devuelve {cuit_digits: monto} tomando SOLO la columna CUIT (firmante principal)
    y sumando Importe/Monto > 0 desde cartera.xlsx.
    NO usa CUIT 2.
    """
    out = {}
    headers, rows, _, status = load_cartera()
    if not headers or not rows:
        return out

    headers_norm = [normalize_header(h) for h in headers]

    idx_cuit = None
    idx_importe = None
    for i, h in enumerate(headers_norm):
        if h == "cuit":
            idx_cuit = i
            break
    for i, h in enumerate(headers_norm):
        if ("importe" in h) or ("monto" in h):
            idx_importe = i
            break

    if idx_cuit is None or idx_importe is None:
        return out

    for r in rows:
        try:
            cd = cuit_digits(r[idx_cuit])
            if not cd:
                continue
            imp = parse_es_number(r[idx_importe])
            if imp is None:
                imp = 0.0
            imp = float(imp)
            if imp <= 0:
                continue
            out[cd] = float(out.get(cd, 0.0)) + imp
        except Exception:
            continue

    return out

def _monitoreable_cuits():
    """
    Universo definitivo de monitoreo:
      1) todos los firmantes precalificados activos
      2) todos los CUIT con cartera (columna CUIT del Excel, no CUIT 2)
    """
    precal_map = _precalificado_map_from_db()
    cartera_map = _cartera_al_dia_map()

    universe = set()
    for cd, is_precal in precal_map.items():
        if cd and is_precal:
            universe.add(cd)
    for cd, monto in cartera_map.items():
        if cd and float(monto or 0.0) > 0:
            universe.add(cd)

    return sorted(universe)

def _cheq_try_lock(*, locked_by: str = "unknown", ttl_seconds: int = 20 * 60) -> bool:
    """Acquire a single-process lock for cheque monitor/sync.

    Uses table cheq_monitor_lock (single row id=1). If a lock exists and has not
    expired, returns False. Otherwise sets/refreshes the lock and returns True.
    """
    now = datetime.now(ARG_TZ)
    now_s = now.isoformat(timespec="seconds")
    exp_s = (now + timedelta(seconds=int(ttl_seconds))).isoformat(timespec="seconds")

    conn = get_db()
    try:
# IMMEDIATE: take a RESERVED lock to avoid races across processes
        conn.execute("BEGIN IMMEDIATE;")
        row = conn.execute(
            "SELECT locked_by, expires_at FROM cheq_monitor_lock WHERE id=1"
        ).fetchone()

        if row is not None:
            expires_at = (row["expires_at"] or "").strip()
            if expires_at and expires_at > now_s:
                conn.execute("ROLLBACK;")
                return False

        conn.execute(
            """
            INSERT OR REPLACE INTO cheq_monitor_lock (id, locked_at, locked_by, expires_at)
            VALUES (1, ?, ?, ?)
            """,
            (now_s, (locked_by or "").strip()[:200], exp_s),
        )
        conn.execute("COMMIT;")
        return True
    except Exception:
        try:
            conn.execute("ROLLBACK;")
        except Exception:
            pass
        return False
    finally:
        conn.close()

def _cheq_release_lock(*, locked_by: str | None = None) -> None:
    """Release the cheque monitor lock (best-effort)."""
    conn = get_db()
    try:
        if locked_by:
            conn.execute(
                "DELETE FROM cheq_monitor_lock WHERE id=1 AND locked_by=?",
                ((locked_by or "").strip()[:200],),
            )
        else:
            conn.execute("DELETE FROM cheq_monitor_lock WHERE id=1")
        conn.commit()
    except Exception:
# best-effort, never crash caller
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        conn.close()

def _cheq_lock_is_active() -> bool:
    """Return True if the cheque lock is currently held and not expired."""
    try:
        conn = get_db()
        now_s = datetime.now(ARG_TZ).isoformat(timespec="seconds")
        row = conn.execute("SELECT expires_at FROM cheq_monitor_lock WHERE id=1").fetchone()
        conn.close()
        if not row:
            return False
        exp = (row["expires_at"] or "").strip()
        return bool(exp and exp > now_s)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass
        return False

def _cheq_progress_autofix_if_stale() -> None:
    """
    Evita UI colgada en 'running' si la corrida murió:
    - Si progress dice running pero el lock NO está activo -> resetea a idle.
    - Si running y updated_at está viejo (>= 180s) -> marca error.
    Best-effort.
    """
    try:
        d = _cheq_progress_get()
        if (d.get("status") or "") != "running":
            return

# Si no hay lock, no debería estar corriendo.
        if not _cheq_lock_is_active():
            _cheq_progress_reset()
            return

# Watchdog por timestamp
        upd = (d.get("updated_at") or "").strip()
        if upd:
            try:
                dt = datetime.fromisoformat(upd)
                if (datetime.now(ARG_TZ) - dt).total_seconds() >= 180:
                    _cheq_progress_set(
                        mode=(d.get("mode") or ""),
                        status="error",
                        total=int(d.get("total") or 0),
                        done=int(d.get("done") or 0),
                        errors=max(1, int(d.get("errors") or 0)),
                        message="Se interrumpió la corrida (watchdog). Reintentar.",
                        run_id=(d.get("run_id") or None),
                    )
            except Exception:
                pass
    except Exception:
        pass

def _cheq_progress_set(*, mode: str, status: str, total: int, done: int, errors: int, message: str = "", run_id: str | None = None, result: dict | None = None) -> None:
    """Persist progress for UI polling (single row). Best-effort."""
    try:
        conn = get_db()
        now_s = datetime.now(ARG_TZ).isoformat(timespec="seconds")
        result_json = json.dumps(result, ensure_ascii=False) if isinstance(result, dict) else None
        conn.execute(
            f"""
            UPDATE {TABLE_CHEQ_PROGRESS}
               SET run_id=?,
                   mode=?,
                   status=?,
                   total=?,
                   done=?,
                   errors=?,
                   started_at=COALESCE(started_at, ?),
                   updated_at=?,
                   message=?,
                   result_json=COALESCE(?, result_json)
             WHERE id=1
            """,
            (
                (run_id or None),
                (mode or "").strip()[:20],
                (status or "").strip()[:20],
                int(total or 0),
                int(done or 0),
                int(errors or 0),
                now_s,
                now_s,
                (message or "").strip()[:400],
                result_json,
            ),
        )
        conn.commit()
        conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass

def _cheq_progress_reset() -> None:
    try:
        conn = get_db()
        conn.execute(
            f"UPDATE {TABLE_CHEQ_PROGRESS} SET run_id=NULL, mode=NULL, status='idle', total=0, done=0, errors=0, started_at=NULL, updated_at=NULL, message='', result_json=NULL WHERE id=1"
        )
        conn.commit()
        conn.close()
    except Exception:
        try:
            conn.close()
        except Exception:
            pass

def _cheq_progress_get() -> dict:
    """Return current progress row as dict."""
    try:
        conn = get_db()
        row = conn.execute(f"SELECT * FROM {TABLE_CHEQ_PROGRESS} WHERE id=1").fetchone()
        conn.close()
        if not row:
            return {"status": "idle", "total": 0, "done": 0, "errors": 0, "message": ""}
        d = dict(row)
# Parse result_json if present
        rj = (d.get("result_json") or "").strip()
        if rj:
            try:
                d["result"] = json.loads(rj)
            except Exception:
                d["result"] = None
        else:
            d["result"] = None
        d.pop("result_json", None)
        return d
    except Exception:
        return {"status": "idle", "total": 0, "done": 0, "errors": 0, "message": ""}

def bcra_fetch_cheques_sin_fondos(cuit_digits_str: str) -> dict:
    out = {"denominacion": "", "events": [], "error": ""}
    url = f"{BCRA_CHEQ_ENDPOINT}/{cuit_digits_str}"

    def _do_request() -> dict:
        req = urllib.request.Request(
            url,
            headers={
                "Accept": "application/json",
                "User-Agent": "MeridianoCRM/1.0",  # ayuda a evitar algunos bloqueos
            },
        )
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = resp.read().decode("utf-8", errors="ignore")
        return json.loads(data)

# Reintento simple (errores transitorios)
    last_err = ""
    for attempt in range(3):  # 0,1,2 (2 retries)
        try:
            j = _do_request()
            last_err = ""
            break
        except urllib.error.HTTPError as e:
            code = getattr(e, "code", "")
            last_err = f"HTTP {code}"
# Reintentar si es 429 o 5xx
            if code in (429, 500, 502, 503, 504) and attempt < 2:
                time.sleep(1.2 * (attempt + 1))
                continue
            out["error"] = last_err
            return out
        except Exception as e:
            last_err = str(e)
# timeout / network: reintentar
            if attempt < 2:
                time.sleep(1.2 * (attempt + 1))
                continue
            out["error"] = last_err
            return out

def bcra_fetch_cheques_all(cuit_digits_str: str) -> dict:
    """
    Fetch ALL causales/events from BCRA cheques rechazados endpoint.
    Returns:
        {"denominacion": str, "events": [ {causal, entidad, detalle_raw(dict)} ], "error": str}
    """
    out = {"denominacion": "", "events": [], "error": ""}
    url = f"{BCRA_CHEQ_ENDPOINT}/{cuit_digits_str}"

    def _do_request() -> dict:
        req = urllib.request.Request(
            url,
            headers={
                "Accept": "application/json",
                "User-Agent": "MeridianoCRM/1.0",
            },
        )
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = resp.read().decode("utf-8", errors="ignore")
        return json.loads(data)

    last_err = ""
    for attempt in range(3):
        try:
            j = _do_request()
            last_err = ""
            break
        except urllib.error.HTTPError as e:
            code = getattr(e, "code", "")
            last_err = f"HTTP {code}"
            if code in (429, 500, 502, 503, 504) and attempt < 2:
                time.sleep(1.2 * (attempt + 1))
                continue
            out["error"] = last_err
            return out
        except Exception as e:
            last_err = str(e)
            if attempt < 2:
                time.sleep(1.2 * (attempt + 1))
                continue
            out["error"] = last_err
            return out

    if last_err:
        out["error"] = last_err
        return out

    results = (j or {}).get("results") or {}
    if isinstance(results, dict):
        out["denominacion"] = str(results.get("denominacion") or "").strip()

    causales = results.get("causales") or []
    if not isinstance(causales, list):
        causales = []

    for c in causales:
        if not isinstance(c, dict):
            continue
        causal = str(c.get("causal") or "").strip()
        entidades = c.get("entidades") or []
        if not isinstance(entidades, list):
            entidades = []
        for ent in entidades:
            if not isinstance(ent, dict):
                continue
            entidad = ent.get("entidad")
            detalle = ent.get("detalle") or []
            if not isinstance(detalle, list):
                detalle = []
            for d in detalle:
                if not isinstance(d, dict):
                    continue
                out["events"].append({"causal": causal, "entidad": entidad, "detalle_raw": d})

    def _k(ev: dict):
        d = (ev or {}).get("detalle_raw") or {}
        fr = str(d.get("fechaRechazo") or "")[:10]
        nro = str(d.get("nroCheque") or "")
        return (fr, nro)

    out["events"].sort(key=_k, reverse=True)
    return out

    if last_err:
        out["error"] = last_err
        return out

    results = (j or {}).get("results") or {}
    if isinstance(results, dict):
        out["denominacion"] = str(results.get("denominacion") or "").strip()

    causales = results.get("causales") or []
    if not isinstance(causales, list):
        causales = []

    for c in causales:
        if not isinstance(c, dict):
            continue
        causal = str(c.get("causal") or "").strip().upper()
# robustez: por si viene "SIN FONDOS - ..."
        if not causal.startswith("SIN FONDOS"):
            continue

        entidades = c.get("entidades") or []
        if not isinstance(entidades, list):
            entidades = []

        for ent in entidades:
            if not isinstance(ent, dict):
                continue
            entidad = ent.get("entidad")
            detalle = ent.get("detalle") or []
            if not isinstance(detalle, list):
                detalle = []

            for d in detalle:
                if not isinstance(d, dict):
                    continue
                nro = d.get("nroCheque")
                fr = d.get("fechaRechazo")
                monto = d.get("monto")
                fpago = d.get("fechaPago")

                fr_s = str(fr)[:10] if fr else ""
                try:
                    monto_f = float(monto) if monto is not None and str(monto).strip() != "" else 0.0
                except Exception:
                    monto_f = 0.0

                out["events"].append({
                    "entidad": entidad,
                    "nro_cheque": str(nro) if nro is not None else "",
                    "fecha_rechazo": fr_s,
                    "monto": monto_f,
                    "pagado": bool(fpago),
                })

    out["events"].sort(
        key=lambda e: (e.get("fecha_rechazo") or "", e.get("nro_cheque") or ""),
        reverse=True
    )
    return out

def run_cheq_monitor(run_type: str, scope: str = "all", scope_cuit: str = "", alert_enabled: bool = True) -> dict:
    """
    Monitorea cheques rechazados (SIN FONDOS) y opcionalmente envía alertas.
    En corridas tipo sync (o con alert_enabled=False) además persiste TODOS los eventos (todas las causales)
    en cheq_events_raw.

    Optimizaciones:
    - 1 request por CUIT: se pide ALL y de ahí se deriva SIN FONDOS.
    - Concurrencia controlada para fetch (ThreadPoolExecutor); escritura SQLite en el hilo principal.
    - Commits en lote.
    """
    stats = {
        "ok": True,
        "run_type": run_type,
        "scope": scope,
        "scope_cuit": scope_cuit,
        "started_at": now_str(),
        "ended_at": "",
        "total_cuits": 0,
        "new_events": 0,
        "alerts_sent": 0,
        "errors": 0,
    }

    mode = "sync" if (not alert_enabled) or (run_type == "sync") else "alert"
    run_uuid = str(uuid.uuid4())
    _cheq_progress_set(mode=mode, status="running", total=0, done=0, errors=0, message="Preparando…", run_id=run_uuid)

# log run start (best-effort)
    run_id = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            f"INSERT INTO {TABLE_CHEQ_RUNS}(run_type, scope, scope_cuit, started_at) VALUES (?,?,?,?)",
            (run_type, scope, scope_cuit or None, stats["started_at"]),
        )
        run_id = cur.lastrowid
        conn.commit()
        conn.close()
    except Exception:
        pass

# scope: soporta "single" y "one"
    if scope in ("single", "one"):
        cd1 = cuit_digits(scope_cuit)
        cuits = [cd1] if cd1 else []
    else:
        cuits = _monitoreable_cuits()

    cuits = [c for c in cuits if c]
    stats["total_cuits"] = len(cuits)

    _cheq_progress_set(mode=mode, status="running", total=stats["total_cuits"], done=0, errors=0, message=f"Iniciando consultas… (0/{stats['total_cuits']})", run_id=run_uuid)

    def fmt_amt(x: float) -> str:
        try:
            return f"$ {x:,.0f}".replace(",", ".")
        except Exception:
            return f"$ {x}"

    err_samples = []

# Prefetch RS for speed
    rs_map = {}
    try:
        if cuits:
            conn0 = get_db()
            qmarks = ",".join(["?"] * len(cuits))
            rows0 = conn0.execute(
                f"SELECT cuit_digits, razon_social FROM {TABLE_FIRMANTES} WHERE cuit_digits IN ({qmarks})",
                tuple(cuits),
            ).fetchall()
            conn0.close()
            for r in rows0:
                rs_map[(r["cuit_digits"] or "").strip()] = (r["razon_social"] or "").strip()
    except Exception:
        try:
            conn0.close()
        except Exception:
            pass

# Single writer connection (SQLite)
    conn_ins = None
    cur_ins = None
    try:
        conn_ins = get_db()
# PRAGMAs para acelerar inserts (balance razonable)
        try:
            conn_ins.execute("PRAGMA journal_mode=WAL;")
            conn_ins.execute("PRAGMA synchronous=NORMAL;")
            conn_ins.execute("PRAGMA temp_store=MEMORY;")
        except Exception:
            pass
        cur_ins = conn_ins.cursor()
    except Exception:
        conn_ins = None
        cur_ins = None

    done_count = 0
    commit_counter = 0

# Control de concurrencia (fetch en paralelo, write en main thread)
    try:
        WORKERS = int(os.getenv("CHEQ_WORKERS", "6") or "6")
    except Exception:
        WORKERS = 20
    WORKERS = max(1, min(16, WORKERS))  # hard cap para evitar 429 masivo

    def _to_sin_fondos_events(bcra_all: dict) -> list:
        """Deriva eventos SIN FONDOS (formato normalizado) desde respuesta ALL."""
        out = []
        for ev in (bcra_all or {}).get("events") or []:
            try:
                causal = str((ev or {}).get("causal") or "").strip().upper()
                if causal != "SIN FONDOS":
                    continue
                entidad = (ev or {}).get("entidad")
                d = (ev or {}).get("detalle_raw") or {}
                nro = d.get("nroCheque")
                fr = d.get("fechaRechazo")
                monto = d.get("monto")
                fpago = d.get("fechaPago")
                try:
                    monto_f = float(monto) if monto is not None and str(monto).strip() != "" else 0.0
                except Exception:
                    monto_f = 0.0
                out.append(
                    {
                        "entidad": entidad,
                        "nro_cheque": str(nro) if nro is not None else "",
                        "fecha_rechazo": str(fr or "").strip()[:10],
                        "monto": monto_f,
                        "pagado": bool(fpago),
                    }
                )
            except Exception:
                continue
        out.sort(key=lambda e: (e.get("fecha_rechazo") or "", e.get("nro_cheque") or ""), reverse=True)
        return out

    def _fetch_one(cd: str) -> tuple:
# 1 sola request por cuit
        return cd, bcra_fetch_cheques_all(cd)

# Fetch en paralelo
    futures = []
    try:
        with ThreadPoolExecutor(max_workers=WORKERS) as ex:
            fut_map = {ex.submit(_fetch_one, cd): cd for cd in cuits}
            for fut in as_completed(fut_map):
                cd = fut_map.get(fut) or ""
                bcra_all = None
                try:
                    cd, bcra_all = fut.result()
                except Exception as e:
                    stats["errors"] += 1
                    if len(err_samples) < 10:
                        err_samples.append(f"{cd}: {e}")
                    bcra_all = {"denominacion": "", "events": [], "error": str(e)}

# --- Procesamiento / persistencia (main thread) ---
                try:
# Persist raw ALL events ONLY for sync runs (o alert_enabled=False)
                    if ((not alert_enabled) or (run_type == "sync")) and bcra_all and not bcra_all.get("error"):
                        try:
                            cur = cur_ins
                            if cur is not None:
                                for ev in (bcra_all.get("events") or []):
                                    causal = (ev or {}).get("causal") or ""
                                    entidad_raw = (ev or {}).get("entidad")
                                    detalle_raw = (ev or {}).get("detalle_raw") or {}
                                    eid_raw = _cheq_raw_event_id(cd, str(causal), entidad_raw, detalle_raw)
                                    cur.execute(
                                        f"INSERT OR IGNORE INTO {TABLE_CHEQ_RAW}(cuit_digits, event_id, causal, entidad, payload_json, detected_at) VALUES (?,?,?,?,?,?)",
                                        (cd, eid_raw, str(causal), entidad_raw, json.dumps(detalle_raw, ensure_ascii=False), now_str()),
                                    )
                        except Exception:
# raw es best-effort
                            pass

                    if bcra_all.get("error"):
                        stats["errors"] += 1
                        if len(err_samples) < 10:
                            err_samples.append(f"{cd}: {bcra_all.get('error')}")
# progreso y continuar
                        done_count += 1
                        if done_count <= 10 or stats["total_cuits"] <= 20 or (done_count % 5 == 0) or done_count == stats["total_cuits"]:
                            _cheq_progress_set(
                                mode=mode,
                                status="running",
                                total=stats["total_cuits"],
                                done=done_count,
                                errors=stats["errors"],
                                message=f"Error para {cd}: {bcra_all.get('error')}",
                                run_id=run_uuid,
                            )
                        continue

                    events = _to_sin_fondos_events(bcra_all)
                    if not events:
                        done_count += 1
                        if done_count <= 10 or stats["total_cuits"] <= 20 or (done_count % 5 == 0) or done_count == stats["total_cuits"]:
                            _cheq_progress_set(
                                mode=mode,
                                status="running",
                                total=stats["total_cuits"],
                                done=done_count,
                                errors=stats["errors"],
                                message=f"OK {cd} (sin eventos)",
                                run_id=run_uuid,
                            )
                        continue

                    denom_api = (bcra_all.get("denominacion") or "").strip()
                    rs_db = (rs_map.get(cd) or _get_firmante_rs_from_db(cd))
                    title_rs = rs_db or denom_api or "VER"

                    new_items = []
                    cur = cur_ins
                    if cur is None:
                        raise RuntimeError("DB not available")

                    for e in events:
                        eid = _cheq_event_id(cd, e)
                        cur.execute(
                            f"""
                            INSERT OR IGNORE INTO {TABLE_CHEQ_SEEN}
                            (cuit_digits, event_id, entidad, nro_cheque, fecha_rechazo, monto, pagado, detected_at)
                            VALUES (?,?,?,?,?,?,?,?)
                            """,
                            (
                                cd,
                                eid,
                                e.get("entidad"),
                                e.get("nro_cheque"),
                                e.get("fecha_rechazo"),
                                float(e.get("monto") or 0.0),
                                1 if e.get("pagado") else 0,
                                now_str(),
                            ),
                        )
                        if cur.rowcount == 1:
                            new_items.append(e)

# commit in batches for speed
                    commit_counter += 1
                    if conn_ins is not None and (commit_counter % 25 == 0):
                        try:
                            conn_ins.commit()
                        except Exception:
                            pass

                    if new_items:
                        stats["new_events"] += len(new_items)

                        if alert_enabled:
# Armamos mensaje para Slack (y log en DB)
                            total_monto = sum(float(e.get("monto") or 0.0) for e in new_items)
                            total_pagado = sum(float(e.get("monto") or 0.0) for e in new_items if e.get("pagado"))

                            lines = []
                            lines.append(f"*🚨 Cheques rechazados SIN FONDOS* — *{title_rs}* ({cd})")
                            lines.append(f"Nuevos eventos: *{len(new_items)}* | Monto: *{fmt_amt(total_monto)}* | Pagados: *{fmt_amt(total_pagado)}*")
                            lines.append("")
                            lines.append("Fecha Rechazo | Monto | Pagado")
                            lines.append("---|---:|:---:")
                            rows = []
                            for e in sorted(new_items, key=lambda x: (x.get("fecha_rechazo") or "", float(x.get("monto") or 0.0)), reverse=True)[:10]:
                                fr = e.get("fecha_rechazo") or ""
                                mo = fmt_amt(float(e.get("monto") or 0.0))
                                pg = "SI" if e.get("pagado") else "NO"
                                rows.append((fr, mo, pg))
                            for fr, mo, pg in rows:
                                lines.append(f"{fr} | {mo} | {pg}")

                            extra = max(0, len(new_items) - len(rows))
                            if extra > 0:
                                lines.append(f"(+ {extra} más)")

                            msg = "\n".join(lines)

                            sent_ok = False
                            try:
                                sent_ok = send_slack_message(msg)
                            except Exception:
                                sent_ok = False

                            if sent_ok:
                                stats["alerts_sent"] += 1

# log alert (best-effort)
                            try:
                                if run_id is not None:
                                    cur.execute(
                                        f"INSERT INTO {TABLE_CHEQ_ALERTS}(run_id, cuit_digits, razon_social, message, sent_ok, sent_at) VALUES (?,?,?,?,?,?)",
                                        (run_id, cd, title_rs, msg, 1 if sent_ok else 0, now_str()),
                                    )
                            except Exception:
                                pass

# progreso
                    done_count += 1
                    if done_count <= 10 or stats["total_cuits"] <= 20 or (done_count % 5 == 0) or done_count == stats["total_cuits"]:
                        _cheq_progress_set(
                            mode=mode,
                            status="running",
                            total=stats["total_cuits"],
                            done=done_count,
                            errors=stats["errors"],
                            message=f"[{done_count}/{stats['total_cuits']}] {cd}…",
                            run_id=run_uuid,
                        )

                except Exception as e:
                    stats["errors"] += 1
                    if len(err_samples) < 10:
                        err_samples.append(f"{cd}: {e}")
                    done_count += 1
                    if done_count <= 10 or stats["total_cuits"] <= 20 or (done_count % 5 == 0) or done_count == stats["total_cuits"]:
                        _cheq_progress_set(
                            mode=mode,
                            status="running",
                            total=stats["total_cuits"],
                            done=done_count,
                            errors=stats["errors"],
                            message=f"Error interno {cd}",
                            run_id=run_uuid,
                        )
                    continue

    except Exception as e:
        stats["ok"] = False
        stats["errors"] += 1
        if len(err_samples) < 10:
            err_samples.append(str(e))

# finalize
    stats["ended_at"] = now_str()

# log run end (best-effort)
    try:
        if run_id is not None:
            conn = get_db()
            conn.execute(
                f"UPDATE {TABLE_CHEQ_RUNS} SET ended_at=?, total_cuits=?, new_events=?, alerts_sent=?, errors=? WHERE id=?",
                (stats["ended_at"], stats["total_cuits"], stats["new_events"], stats["alerts_sent"], stats["errors"], run_id),
            )
            conn.commit()
            conn.close()
    except Exception:
        pass

# push progress done
    final_status = "done" if stats["ok"] else "error"
    _cheq_progress_set(
        mode=mode,
        status=final_status,
        total=stats["total_cuits"],
        done=stats["total_cuits"],
        errors=stats["errors"],
        message="Listo." if stats["ok"] else "Finalizó con errores.",
        run_id=run_uuid,
    )

# close insert conn
    try:
        if conn_ins is not None:
            conn_ins.commit()
            conn_ins.close()
    except Exception:
        try:
            conn_ins.close()
        except Exception:
            pass

    return stats

# --------------------------- | Bloqueos diarios | ---------------------------

_last_blocks_cleanup = None

def cleanup_daily_blocks_if_needed():
    global _last_blocks_cleanup
    today = today_local_date()
    if _last_blocks_cleanup == today:
        return
    conn = get_db()
    conn.execute("DELETE FROM limit_blocks WHERE block_date <> ?", (today,))
    conn.commit()
    conn.close()
    _last_blocks_cleanup = today

def blocks_agg_today_and_list():
    today = today_local_date()
    conn = get_db()
    rows = conn.execute("""
        SELECT id, firmante_cuit_digits, scope, username, amount, created_at
        FROM limit_blocks
        WHERE block_date=?
        ORDER BY id DESC
    """, (today,)).fetchall()

    conn.close()

    agg = {}
    by_firmante = {}
    for r in rows:
        cd = r["firmante_cuit_digits"]
        by_firmante.setdefault(cd, []).append(r)

        agg.setdefault(cd, {"3ros": 0.0, "propio": 0.0, "fce": 0.0})
        agg[cd][r["scope"]] += float(r["amount"] or 0.0)

    return agg, by_firmante

# --------------------------- | Vencimientos de límites (daily) | ---------------------------

_last_expiry_run = None

def apply_expired_limits_if_needed():
    global _last_expiry_run
    today = today_local_date()
    if _last_expiry_run == today:
        return
    _last_expiry_run = today

    conn = get_db()

# buscamos firmantes activos con líneas vencidas todavía marcadas como activas
    rows = conn.execute(f"""
        SELECT cuit_digits, razon_social, credit_limit_3ros, credit_limit_propio,
               limit_expiry_3ros, limit_expiry_propio,
               line_active_3ros, line_active_propio
        FROM {TABLE_FIRMANTES}
        WHERE is_active=1
          AND (
                (limit_expiry_3ros <> '' AND limit_expiry_3ros <= ? AND COALESCE(line_active_3ros, 1)=1)
             OR (limit_expiry_propio <> '' AND limit_expiry_propio <= ? AND COALESCE(line_active_propio, 1)=1)
          )
    """, (today, today)).fetchall()

    for r in rows:
        cd = r["cuit_digits"]
        old3 = float(r["credit_limit_3ros"] or 0.0)
        oldp = float(r["credit_limit_propio"] or 0.0)
        exp3 = (r["limit_expiry_3ros"] or "").strip()
        expp = (r["limit_expiry_propio"] or "").strip()

        line3_active = int(r["line_active_3ros"] or 0)
        linep_active = int(r["line_active_propio"] or 0)
        parts = []

        if exp3 and exp3 <= today and line3_active == 1:
            line3_active = 0
            parts.append(f"3ros vencido ({exp3}): límite se mantiene en {fmt_ars(old3)} y pasa a inactiva")

        if expp and expp <= today and linep_active == 1:
            linep_active = 0
            parts.append(f"propio vencido ({expp}): límite se mantiene en {fmt_ars(oldp)} y pasa a inactiva")

        if parts:
            conn.execute(f"""
                UPDATE {TABLE_FIRMANTES}
                SET line_active_3ros=?,
                    line_active_propio=?,
                    updated_at=?
                WHERE cuit_digits=?
            """, (line3_active, linep_active, now_str(), cd))
            audit_log(cd, "VENCIMIENTO", "system", " | ".join(parts), conn=conn)

    conn.commit()
    conn.close()

@app.before_request
def _daily_hooks():
    cleanup_daily_blocks_if_needed()
    apply_expired_limits_if_needed()

# --------------------------- | Firmantes helpers | ---------------------------

def get_firmante(cd: str):
    conn = get_db()
    r = conn.execute(f"SELECT * FROM {TABLE_FIRMANTES} WHERE cuit_digits=?", (cd,)).fetchone()
    conn.close()
    return r

def _get_firmante_rs_from_db(cuit_digits: str) -> str | None:
    """Devuelve razón social/denominación desde DB para un CUIT (solo dígitos).
    Helper defensivo para el módulo de alertas.
    """
    try:
        r = get_firmante((cuit_digits or "").strip())
    except Exception:
        return None
    if not r:
        return None
# sqlite3.Row expone keys()
    try:
        keys = set(r.keys())
    except Exception:
        keys = set()
    for k in ("razon_social", "denominacion", "razon", "nombre", "name"):
        if (k in keys) or (not keys):
            try:
                v = r[k]
            except Exception:
                continue
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
    return None

# --------------------------- | Grupos económicos helpers | ---------------------------

def get_group(group_id: int):
    conn = get_db()
    r = conn.execute(f"SELECT * FROM {TABLE_GROUPS} WHERE id=?", (int(group_id),)).fetchone()
    conn.close()
    return r

def get_group_by_name(nombre: str):
    if nombre is None:
        return None
    nombre = str(nombre).strip()
    if not nombre:
        return None
    conn = get_db()
    r = conn.execute(f"SELECT * FROM {TABLE_GROUPS} WHERE nombre=?", (nombre,)).fetchone()
    conn.close()
    return r

def upsert_group(nombre: str, limite_grupal, actor: str, allow_update=True):
    nombre = (nombre or "").strip()
    if not nombre:
        return False, "El nombre del grupo es obligatorio.", None

# limite_grupal puede ser None => no pisa si existe
    lim = None if limite_grupal is None else float(limite_grupal or 0.0)

    conn = get_db()
    existing = conn.execute(f"SELECT * FROM {TABLE_GROUPS} WHERE nombre=?", (nombre,)).fetchone()

    if existing is None:
        lim0 = float(lim or 0.0)
        conn.execute(f"""
            INSERT INTO {TABLE_GROUPS} (nombre, credit_limit_grupal, is_active, created_at, updated_at)
            VALUES (?, ?, 1, ?, ?)
        """, (nombre, lim0, now_str(), now_str()))
        conn.commit()
        gid = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
        conn.close()
        return True, "Grupo creado.", int(gid)

    if not allow_update:
        gid = int(existing["id"])
        conn.close()
        return False, "El grupo ya existe.", gid

    old_lim = float(existing["credit_limit_grupal"] or 0.0)
    new_lim = old_lim if lim is None else float(lim)

    conn.execute(f"""
        UPDATE {TABLE_GROUPS}
        SET credit_limit_grupal=?,
            updated_at=?
        WHERE id=?
    """, (new_lim, now_str(), int(existing["id"])))
    conn.commit()
    conn.close()

# No audit log específico para grupos por requerimiento.
    return True, "Grupo actualizado.", int(existing["id"])

def list_groups():
    conn = get_db()
    rows = conn.execute(f"SELECT * FROM {TABLE_GROUPS} WHERE is_active=1 ORDER BY nombre COLLATE NOCASE").fetchall()
    conn.close()
    return rows

def compute_group_available_map(agg=None, facturas_agg=None, blocks_agg=None):
    if agg is None:
        _, _, agg, _ = load_cartera()
    if facturas_agg is None:
        _, _, facturas_agg, _ = load_facturas()
    if blocks_agg is None:
        blocks_agg, _ = blocks_agg_today_and_list()

    agg = agg or {}
    facturas_agg = facturas_agg or {}
    blocks_agg = blocks_agg or {}

    conn = get_db()
    groups_rows = conn.execute(
        f"SELECT id, credit_limit_grupal FROM {TABLE_GROUPS} WHERE is_active=1"
    ).fetchall()
    firm_rows = conn.execute(
        f"SELECT cuit_digits, grupo_id FROM {TABLE_FIRMANTES} WHERE grupo_id IS NOT NULL"
    ).fetchall()
    conn.close()

    g_to_cuits = {}
    for fr in firm_rows:
        try:
            gid = int(fr["grupo_id"])
        except Exception:
            continue
        cd = (fr["cuit_digits"] or "").strip()
        if not cd:
            continue
        g_to_cuits.setdefault(gid, set()).add(cd)

    out = {}
    for g in groups_rows:
        gid = int(g["id"])
        cuits = g_to_cuits.get(gid, set())
        used = 0.0
        blocked = 0.0
        for cd in cuits:
            a = agg.get(cd) or {"3ros": 0.0, "propio": 0.0}
            used += float(a.get("3ros", 0.0)) + float(a.get("propio", 0.0)) + float(facturas_agg.get(cd, 0.0))
            b = blocks_agg.get(cd) or {"3ros": 0.0, "propio": 0.0, "fce": 0.0}
            blocked += float(b.get("3ros", 0.0)) + float(b.get("propio", 0.0)) + float(b.get("fce", 0.0))
        lim = float(g["credit_limit_grupal"] or 0.0)
        out[gid] = {
            "limit": lim,
            "used": used,
            "blocked": blocked,
            "avail": lim - used - blocked,
        }
    return out

_KEEP = object()

def upsert_firmante(razon_social, cuit, lim3, limp, limf, exp3, expp, actor,
                    allow_update=True, reactivate=False, group_id=_KEEP, primera_linea=_KEEP,
                    observacion: str = ""):
    cd = cuit_digits(cuit)
    if not cd or not razon_social:
        return False, "Razón social y CUIT son obligatorios."

    lim3_val = None if lim3 is None else float(lim3 or 0.0)
    limp_val = None if limp is None else float(limp or 0.0)
    limf_val = None if limf is None else float(limf or 0.0)

    exp3_val = None if exp3 is None else parse_date_any(exp3)
    expp_val = None if expp is None else parse_date_any(expp)

    conn = get_db()
    existing = conn.execute(f"SELECT * FROM {TABLE_FIRMANTES} WHERE cuit_digits=?", (cd,)).fetchone()

    if existing is None:
        conn.execute(f"""
            INSERT INTO {TABLE_FIRMANTES}
            (razon_social, cuit, cuit_digits,
             credit_limit_3ros, credit_limit_propio, credit_limit_fce,
             limit_expiry_3ros, limit_expiry_propio,
             line_active_3ros, line_active_propio,
             grupo_id, primera_linea,
             is_active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
        """, (
            razon_social.strip(), str(cuit).strip(), cd,
            float(lim3_val or 0.0), float(limp_val or 0.0), float(limf_val or 0.0),
            (exp3_val or "") if exp3_val is not None else "",
            (expp_val or "") if expp_val is not None else "",
            0 if line_is_expired((exp3_val or "") if exp3_val is not None else "") else 1,
            0 if line_is_expired((expp_val or "") if expp_val is not None else "") else 1,
            None if group_id is _KEEP else (group_id if group_id is not None else None),
            1 if (False if primera_linea is _KEEP else bool(primera_linea)) else 0,
            now_str(), now_str()
        ))
        conn.commit()
        conn.close()
        audit_log(
            cd, "ALTA", actor,
            append_observation(
                f"Límites: 3ros={fmt_ars(float(lim3_val or 0.0))} (vto {(exp3_val or '—') if exp3_val is not None else '—'}); "
                f"propio={fmt_ars(float(limp_val or 0.0))} (vto {(expp_val or '—') if expp_val is not None else '—'}); "
                f"FCE={fmt_ars(float(limf_val or 0.0))}",
                observacion
            )
        )
        return True, "Firmante creado."

    if not allow_update:
        conn.close()
        return False, "El firmante ya existe."

    old3 = float(existing["credit_limit_3ros"] or 0.0)
    oldp = float(existing["credit_limit_propio"] or 0.0)
    oldf = float(existing["credit_limit_fce"] or 0.0)
    olde3 = (existing["limit_expiry_3ros"] or "").strip()
    oldep = (existing["limit_expiry_propio"] or "").strip()
    old_active = int(existing["is_active"] or 0)
    old_line3_active = int(existing["line_active_3ros"] or 0)
    old_linep_active = int(existing["line_active_propio"] or 0)
    old_gid = existing["grupo_id"]
    old_primera = int(existing["primera_linea"] or 0)

    new_active = old_active
    if reactivate and old_active == 0:
        new_active = 1

    new3 = old3 if lim3_val is None else float(lim3_val)
    newp = oldp if limp_val is None else float(limp_val)
    newf = oldf if limf_val is None else float(limf_val)
    newe3 = olde3 if exp3_val is None else (exp3_val or "")
    newep = oldep if expp_val is None else (expp_val or "")

    new_line3_active = old_line3_active
    if exp3_val is not None:
        new_line3_active = 0 if line_is_expired(newe3) else 1
    elif lim3_val is not None and line_is_expired(newe3):
        new_line3_active = 0

    new_linep_active = old_linep_active
    if expp_val is not None:
        new_linep_active = 0 if line_is_expired(newep) else 1
    elif limp_val is not None and line_is_expired(newep):
        new_linep_active = 0

    if group_id is _KEEP:
        new_gid = old_gid
    else:
        new_gid = group_id

    if primera_linea is _KEEP:
        new_primera = old_primera
    else:
        new_primera = 1 if bool(primera_linea) else 0

    conn.execute(f"""
        UPDATE {TABLE_FIRMANTES}
        SET razon_social=?,
            cuit=?,
            credit_limit_3ros=?,
            credit_limit_propio=?,
            credit_limit_fce=?,
            limit_expiry_3ros=?,
            limit_expiry_propio=?,
            line_active_3ros=?,
            line_active_propio=?,
            grupo_id=?,
            primera_linea=?,
            is_active=?,
            updated_at=?
        WHERE cuit_digits=?
    """, (
        razon_social.strip(), str(cuit).strip(),
        new3, newp, newf, newe3, newep, new_line3_active, new_linep_active,
        new_gid, new_primera, new_active, now_str(), cd
    ))
    conn.commit()
    conn.close()

    changes = []
    if old_active == 0 and new_active == 1:
        changes.append("Reactivado")
    if old3 != new3:
        changes.append(f"3ros: {fmt_ars(old3)} → {fmt_ars(new3)}")
    if oldp != newp:
        changes.append(f"propio: {fmt_ars(oldp)} → {fmt_ars(newp)}")
    if oldf != newf:
        changes.append(f"FCE: {fmt_ars(oldf)} → {fmt_ars(newf)}")
    if olde3 != newe3:
        changes.append(f"vto 3ros: {olde3 or '—'} → {newe3 or '—'}")
    if oldep != newep:
        changes.append(f"vto propio: {oldep or '—'} → {newep or '—'}")
    if old_line3_active != new_line3_active:
        changes.append(f"estado 3ros: {line_status_label(old_line3_active, olde3)} → {line_status_label(new_line3_active, newe3)}")
    if old_linep_active != new_linep_active:
        changes.append(f"estado propio: {line_status_label(old_linep_active, oldep)} → {line_status_label(new_linep_active, newep)}")

    if old_gid != new_gid:
        old_name = "—"
        new_name = "—"
        if old_gid:
            og = get_group(int(old_gid))
            if og:
                old_name = og["nombre"]
        if new_gid:
            ng = get_group(int(new_gid))
            if ng:
                new_name = ng["nombre"]
        changes.append(f"grupo: {old_name} → {new_name}")
    if old_primera != new_primera:
        changes.append(f"primera línea: {'Sí' if old_primera else 'No'} → {'Sí' if new_primera else 'No'}")

    renewal_parts = []
    if old_line3_active == 0 and new_line3_active == 1 and line_is_expired(olde3):
        renewal_parts.append(f"3ros renovado: vto {olde3 or '—'} → {newe3 or '—'} | límite {fmt_ars(new3)}")
    if old_linep_active == 0 and new_linep_active == 1 and line_is_expired(oldep):
        renewal_parts.append(f"propio renovado: vto {oldep or '—'} → {newep or '—'} | límite {fmt_ars(newp)}")

    action = "ALTA" if (old_active == 0 and new_active == 1) else "MODIFICACION"
    audit_log(
        cd,
        action,
        actor,
        append_observation(" | ".join(changes) if changes else "Sin cambios", observacion)
    )
    if renewal_parts:
        audit_log(
            cd,
            "RENOVACION",
            actor,
            append_observation(" | ".join(renewal_parts), observacion)
        )
    return True, "Firmante actualizado."

def set_firmante_active(cd: str, active: bool, actor: str, observacion: str = ""):
    conn = get_db()
    r = conn.execute(f"SELECT * FROM {TABLE_FIRMANTES} WHERE cuit_digits=?", (cd,)).fetchone()
    if not r:
        conn.close()
        return False, "Firmante no encontrado."

    conn.execute(f"UPDATE {TABLE_FIRMANTES} SET is_active=?, updated_at=? WHERE cuit_digits=?",
                 (1 if active else 0, now_str(), cd))
    conn.commit()
    conn.close()

    audit_log(
        cd,
        "BAJA" if not active else "ALTA",
        actor,
        append_observation("Reactivación" if active else "Desactivación", observacion)
    )
    return True, "OK"


def delete_firmante_hard(cd: str, actor: str, observacion: str = ""):
    cd = cuit_digits(cd)
    if not cd:
        return False, "Firmante inválido."

    conn = get_db()
    r = conn.execute(f"SELECT * FROM {TABLE_FIRMANTES} WHERE cuit_digits=?", (cd,)).fetchone()
    if not r:
        conn.close()
        return False, "Firmante no encontrado."

    rs = (r["razon_social"] or "").strip()
    cuit = (r["cuit"] or "").strip()

    try:
        conn.execute("DELETE FROM limit_blocks WHERE firmante_cuit_digits=?", (cd,))
    except Exception:
        pass

    conn.execute(f"DELETE FROM {TABLE_FIRMANTES} WHERE cuit_digits=?", (cd,))
    audit_log(
        cd,
        "ELIMINACION",
        actor,
        append_observation(f"Eliminación definitiva de firmante: {rs} ({cuit or cd})", observacion),
        conn=conn,
    )
    conn.commit()
    conn.close()
    return True, "Firmante eliminado definitivamente."

# --------------------------- | CSV robusto | ---------------------------

def decode_bytes_best_effort(raw: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(enc)
        except Exception:
            continue
    return raw.decode("latin-1", errors="replace")

def read_csv_dictreader(text: str):
    sample = text[:5000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
    except Exception:
        dialect = csv.excel
        dialect.delimiter = ';'
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if reader.fieldnames and len(reader.fieldnames) == 1 and (";" in reader.fieldnames[0]):
        reader = csv.DictReader(io.StringIO(text), delimiter=';')
    return reader

# --------------------------- | UI (Toolbar / CSS / JS) | ---------------------------

TOOLBAR_CSS = '\n.toolbar{\n  position:fixed; top:0; left:0; right:0;\n  height:64px; display:flex; align-items:center; justify-content:space-between;\n  padding: 0 18px; z-index: 999;\n  background: rgba(8, 12, 22, 0.92);\n  border-bottom: 1px solid rgba(255,255,255,0.10);\n  backdrop-filter: blur(10px);\n}\n.tb-left,.tb-right{ display:flex; gap:10px; align-items:center; flex-wrap:wrap; }\n.tb-btn{\n  display:inline-block; padding:10px 12px; border-radius:12px;\n  text-decoration:none; border:1px solid rgba(255,255,255,0.14);\n  background: rgba(255,255,255,0.06); color:#fff;\n  font-size:14px;\n}\n.tb-btn:hover{ background: rgba(255,255,255,0.10); }\n'

BASE_FLASH_CSS = '\n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n'

LIVE_SEARCH_JS = '\n<script>\n(function(){\n  function debounce(fn, ms){\n    let t=null;\n    return function(){\n      const args = arguments;\n      clearTimeout(t);\n      t=setTimeout(()=>fn.apply(null,args), ms);\n    }\n  }\n  document.querySelectorAll(\'form[data-live-search="1"]\').forEach(form=>{\n    const delay = parseInt(form.getAttribute(\'data-live-delay\') || \'350\', 10);\n    const trigger = debounce(()=>{ form.requestSubmit(); }, delay);\n\n    form.querySelectorAll(\'.live-search\').forEach(el=>{\n      const tag = (el.tagName || \'\').toLowerCase();\n      if(tag === \'select\'){\n        el.addEventListener(\'change\', ()=>form.requestSubmit());\n      }else{\n        el.addEventListener(\'input\', trigger);\n      }\n    });\n  });\n})();\n</script>\n'

def safe_url_for(endpoint: str, **values):
    try:
        return url_for(endpoint, **values)
    except BuildError:
        return None

def render_toolbar_html():
    links = [
        (url_for("home"), "🏠 Home"),
        (url_for("fp"), "👥 Firmantes Precalificados"),
        (url_for("groups"), "🏢 Grupos Económicos"),
        (url_for("cartera"), "📁 Cartera"),
        (url_for("audit"), "🧾 Auditoría ABM"),
    ]
    if not is_sales():
        links.append((safe_url_for("alerts"), "🚨 Alertas diarias"))
    if is_admin():
        links.append((url_for("users"), "👤 Usuarios"))
    links.append((url_for("logout"), "Salir"))

    right = "".join([f'<a class="tb-btn" href="{href}">{label}</a>' for href, label in links if href])

    return f"""
<div class="toolbar">
  <div class="tb-left">
    <button class="tb-btn" type="button" onclick="history.back()">◀ Anterior</button>
    <button class="tb-btn" type="button" onclick="history.forward()">Siguiente ▶</button>
  </div>
  <div class="tb-right">{right}</div>
</div>
"""

# --------------------------- | Templates | ---------------------------

LOGIN_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Login</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ min-height:100vh; display:flex; align-items:center; justify-content:center; padding:24px; }\n  .card{\n    width: 420px; max-width: 100%;\n    background: rgba(255,255,255,0.06);\n    border:1px solid rgba(255,255,255,0.12);\n    border-radius:16px; padding:20px;\n    box-shadow: 0 10px 30px rgba(0,0,0,0.35);\n  }\n  h2{ margin:0 0 12px 0; }\n  label{ display:block; margin-top:10px; color:#cbd5e1; font-size:13px; }\n  input{\n    width:100%; padding:12px; margin-top:6px;\n    border-radius:12px; border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  button{\n    width:100%; margin-top:14px; padding:12px;\n    border:none; border-radius:12px; font-weight:800;\n    background:#fff; color:#0b1220; cursor:pointer;\n  }\n  .muted{ margin-top:10px; color:#94a3b8; font-size:12px; }\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n\n  .pwrap{ display:none; margin:12px 0 6px; padding:10px; border-radius:14px; background: rgba(0,0,0,0.18); border:1px solid rgba(255,255,255,0.10); }\n  .pbar{ height:10px; border-radius:999px; background: rgba(255,255,255,0.10); overflow:hidden; }\n  .pbar > div{ height:100%; width:0%; background:#22c55e; transition: width .2s ease; }\n  .pmeta{ display:flex; justify-content:space-between; gap:10px; margin-top:8px; color:#cbd5e1; font-size:12px; }\n  .pmeta b{ color:#fff; }\n\n</style></head><body>\n<div class="page"><div class="card">\n  <h2>Ingreso</h2><div class="muted">Sistema de gestión</div>\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n  <form method="post" action="{{ url_for(\'login\') }}">\n    <input type="hidden" name="next" value="{{ next or \'\' }}" />\n    <label>Usuario</label><input name="username" autocomplete="username" required />\n    <label>Clave</label><input name="password" type="password" autocomplete="current-password" required />\n    <button type="submit">Entrar</button>\n  </form>\n</div></div>\n<script>\n(function(){\n  const pwrap = document.getElementById(\'pwrap\');\n  const pfill = document.getElementById(\'pfill\');\n  const pmode = document.getElementById(\'pmode\');\n  const pmsg  = document.getElementById(\'pmsg\');\n  const pcount= document.getElementById(\'pcount\');\n  const perr  = document.getElementById(\'perr\');\n\n  let timer = null;\n  let jobStarted = false;\n\n  function show(){ if(pwrap) pwrap.style.display=\'block\'; }\n  function hide(){ if(pwrap) pwrap.style.display=\'none\'; }\n  function setPct(done,total){\n    const pct = total>0 ? Math.max(0, Math.min(100, Math.round((done/total)*100))) : 0;\n    if(pfill) pfill.style.width = pct + \'%\';\n  }\n\n  async function poll(){\n    try{\n      const r = await fetch(\'{{ url_for("alerts_progress") }}\', {cache:\'no-store\'});\n      if(!r.ok) return;\n      const j = await r.json();\n      const status = (j.status||\'\').toLowerCase();\n      const total = Number(j.total||0);\n      const done  = Number(j.done||0);\n      const errors= Number(j.errors||0);\n      const mode  = (j.mode||\'\').toLowerCase();\n      const msg   = j.message || \'\';\n      show();\n      pmode.textContent = (mode===\'sync\') ? \'Sincronización\' : (mode===\'alert\') ? \'Monitoreo con alertas\' : \'Proceso\';\n      pmsg.textContent = msg || (status===\'running\' ? \'Procesando…\' : \'Listo.\');\n      pcount.textContent = done + \'/\' + total;\n      perr.textContent = errors + \' errores\';\n      setPct(done,total);\n\n      if(status && status !== \'running\'){\n        if(timer){ clearInterval(timer); timer=null; }\n        // mantener visible 3s, luego ocultar si quedó idle\n        setTimeout(()=>{ \n          // si no hay otra corrida\n          fetch(\'{{ url_for("alerts_progress") }}\', {cache:\'no-store\'}).then(rr=>rr.json()).then(j2=>{\n            const st = (j2.status||\'\').toLowerCase();\n            if(st===\'idle\') hide();\n          }).catch(()=>{});\n        }, 3000);\n      }\n    }catch(e){}\n  }\n\n  async function startAsync(url, form){\n    show();\n    pmode.textContent = \'Proceso\';\n    pmsg.textContent = \'Iniciando…\';\n    pcount.textContent = \'0/0\';\n    perr.textContent = \'0 errores\';\n    setPct(0,0);\n\n    const fd = new FormData(form);\n    const r = await fetch(url, {method:\'POST\', body: fd});\n    if(r.status === 409){\n      pmsg.textContent = \'Ya hay una corrida en ejecución.\';\n      if(timer){ clearInterval(timer); timer=null; }\n      timer = setInterval(poll, 1000);\n      return;\n    }\n    if(!r.ok){\n      pmsg.textContent = \'No se pudo iniciar.\';\n      return;\n    }\n    if(timer){ clearInterval(timer); timer=null; }\n    timer = setInterval(poll, 800);\n    poll();\n  }\n\n  const frmRun = document.getElementById(\'frmRun\');\n  if(frmRun){\n    frmRun.addEventListener(\'submit\', function(ev){\n      ev.preventDefault();\n      startAsync(\'{{ url_for("alerts_run_async") }}\', frmRun);\n    });\n  }\n  const frmSync = document.getElementById(\'frmSync\');\n  if(frmSync){\n    frmSync.addEventListener(\'submit\', function(ev){\n      ev.preventDefault();\n      startAsync(\'{{ url_for("alerts_sync_async") }}\', frmSync);\n    });\n  }\n\n  // If a job is already running when opening the page, show it.\n  poll();\n  if(timer){ clearInterval(timer); timer=null; }\n  timer = setInterval(poll, 2000);\n})();\n</script>\n\n</body></html>\n'

HOME_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Home</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1100px; margin:0 auto; }\n  .card{\n    background: rgba(255,255,255,0.06);\n    border:1px solid rgba(255,255,255,0.12);\n    border-radius:16px; padding:18px;\n  }\n  .row{ display:flex; gap:10px; flex-wrap:wrap; margin-top:14px; }\n  a.btn{\n    display:inline-block; padding:12px 14px; border-radius:12px;\n    text-decoration:none; color:#0b1220; background:#fff; font-weight:800;\n  }\n  a.ghost{ background:transparent; color:#fff; border:1px solid rgba(255,255,255,0.14); font-weight:400; }\n  .muted{ color:#cbd5e1; }\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Bienvenido al sistema de gestión(BETA)</h2>\n  <div class="muted" style="margin-top:6px;">Usuario: <b>{{ username }}</b> · Rol: <b>{{ role_label }}</b></div>\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n  <div class="row">\n    <a class="btn" href="{{ url_for(\'fp\') }}">Firmantes Precalificados</a>\n    <a class="btn ghost" href="{{ url_for(\'cartera\') }}">Cartera</a>\n    <a class="btn ghost" href="{{ url_for(\'audit\') }}">Auditoría ABM</a>\n    {% if is_admin %}<a class="btn ghost" href="{{ url_for(\'users\') }}">Usuarios</a>{% endif %}\n    <a class="btn ghost" href="{{ url_for(\'logout\') }}">Salir</a>\n  </div>\n</div></div></div></body></html>\n'

FP_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Firmantes Precalificados</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1600px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n\n  .top{ display:flex; justify-content:space-between; align-items:flex-end; flex-wrap:wrap; gap:12px; }\n  .muted{ color:#cbd5e1; }\n  .pill{\n    display:inline-block; padding:6px 10px; border-radius:999px; font-size:12px;\n    border:1px solid rgba(255,255,255,0.15); color:#cbd5e1;\n  }\n  .pill-green{\n    border:1px solid rgba(34,197,94,0.35);\n    background: rgba(34,197,94,0.12);\n    color:#bbf7d0;\n  }\n\n  .searchbar{ margin-top: 12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center; }\n  .searchbar input, .searchbar select{\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  .searchbar input{ flex: 1; min-width: 260px; }\n  .searchbar button, .searchbar a.btn{\n    padding:10px 14px; border-radius:12px; cursor:pointer; text-decoration:none;\n    border:none; background:#fff; color:#0b1220; font-weight:800;\n  }\n  .searchbar a.btn{ background: transparent; border:1px solid rgba(255,255,255,0.15); color:#fff; font-weight:400; }\n\n  .grid2{ display:grid; grid-template-columns: 1fr 1fr; gap:12px; margin-top:12px; }\n  @media(max-width: 1100px){ .grid2{ grid-template-columns: 1fr; } }\n\n  .panel{\n    background: rgba(255,255,255,0.05);\n    border:1px solid rgba(255,255,255,0.10);\n    border-radius:16px; padding:14px;\n  }\n  .panel h3{ margin:0 0 8px 0; }\n  .row{ display:flex; gap:10px; flex-wrap:wrap; }\n  .row input{\n    flex:1; min-width: 200px;\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  .row input[type="date"]{ min-width: 190px; }\n  .row .small{ min-width: 170px; flex:0 0 auto; }\n\n  table{ width:100%; border-collapse: collapse; margin-top: 14px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }\n  th{ color:#dbeafe; font-size:13px; }\n\n  /* Evitar salto de línea en importes */\n  .num, .num * { white-space: nowrap; }\n  .w-num{ min-width: 170px; }\n  .w-date{ min-width: 120px; }\n\n  a.link{ color:#93c5fd; text-decoration:none; }\n  a.link:hover{ text-decoration: underline; }\n\n  .neg{ color:#fecdd3; font-weight:800; }\n  .rowneg td{ background: rgba(244,63,94,0.06); }\n\n  .rowblock td{ background: rgba(245,158,11,0.06); }\n  .lock-pill{\n    display:inline-block; margin-left:8px; padding:3px 10px; border-radius:999px; font-size:12px;\n    border:1px solid rgba(245,158,11,0.35); background: rgba(245,158,11,0.12); color:#fde68a;\n  }\n\n  .btn-small{\n    padding:8px 10px; border-radius:12px; text-decoration:none; cursor:pointer;\n    border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff; font-size:13px;\n  }\n  .btn-danger{\n    border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12);\n  }\n\n  .details{ background: rgba(255,255,255,0.04); border:1px solid rgba(255,255,255,0.10); border-radius:14px; padding:12px; }\n  .details h4{ margin:0 0 8px 0; }\n  .details table{ margin-top:10px; }\n  .sub{ font-size:12px; color:#94a3b8; margin-top:4px; }\n</style>\n</head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n\n  <div class="top">\n    <div>\n      <h2 style="margin:0;">Firmantes Precalificados</h2>\n      <div class="muted">Los <b>bloqueos</b> se reinician diariamente. Las líneas vencidas conservan el límite pero quedan inactivas hasta renovar o borrar el vencimiento.</div>\n    </div>\n\n    <div id="progWrap" style="display:none; margin-top:12px; padding:12px; border-radius:14px; background:rgba(0,0,0,0.22); border:1px solid rgba(255,255,255,0.12);">\n      <div style="font-weight:700; margin-bottom:8px;">Progreso</div>\n      <div style="width:100%; height:12px; background:rgba(255,255,255,0.12); border-radius:999px; overflow:hidden;">\n        <div id="progBar" style="width:0%; height:100%; background:rgba(0,255,170,0.65);"></div>\n      </div>\n      <div id="progTxt" style="margin-top:8px; font-size:12px; color:rgba(255,255,255,0.85);">—</div>\n    </div>\n\n    <div class="muted">Usuario: <b>{{ username }}</b> · Rol: <b>{{ role_label }}</b></div>\n  </div>\n\n  {% if cartera_status and cartera_status != "OK" %}\n    <div class="pill" style="margin-top:10px;">{{ cartera_status }}</div>\n  {% endif %}\n\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n\n  <form method="get" action="{{ url_for(\'fp\') }}" class="searchbar" data-live-search="1" data-live-delay="350">\n    <input class="live-search" type="text" name="q" value="{{ q or \'\' }}" placeholder="Buscar por CUIT o Razón Social..." />\n    <select name="view_scope" class="live-search">\n      <option value="3ros" {{ \'selected\' if view_scope==\'3ros\' else \'\' }}>Ver: 3ros</option>\n      <option value="propio" {{ \'selected\' if view_scope==\'propio\' else \'\' }}>Ver: propios</option>\n      <option value="fce" {{ \'selected\' if view_scope==\'fce\' else \'\' }}>Ver: FCE</option>\n    </select>\n    <select name="first_line" class="live-search">\n      <option value="all" {{ \'selected\' if first_line==\'all\' else \'\' }}>Primera línea: todos</option>\n      <option value="yes" {{ \'selected\' if first_line==\'yes\' else \'\' }}>Primera línea: sí</option>\n      <option value="no" {{ \'selected\' if first_line==\'no\' else \'\' }}>Primera línea: no</option>\n    </select>\n\n    <input type="hidden" name="show_inactive" value="{{ \'1\' if show_inactive else \'0\' }}" />\n    <input type="hidden" name="group_id" value="{{ group_id }}" />\n    <button type="submit">Filtrar</button>\n    <a class="btn" href="{{ url_for(\'fp\') }}">Limpiar</a>\n    <a class="btn" href="{{ url_for(\'firmantes_all\') }}">Ver todos los firmantes</a>\n    <a class="btn" href="{{ url_for(\'fp\') }}?show_inactive={{ \'0\' if show_inactive else \'1\' }}&view_scope={{ view_scope }}&first_line={{ first_line }}"> {{ \'Ocultar inactivos\' if show_inactive else \'Mostrar inactivos\' }} </a>\n    <button type="button" onclick="exportFP()">Exportar XLSX</button>\n  </form>\n\n  {% if can_edit %}\n  <div class="grid2">\n    <div class="panel">\n      <h3>Alta manual</h3>\n      <form method="post" action="{{ url_for(\'fp_create\') }}">\n        <div class="row">\n          <input name="razon_social" placeholder="Razón social" required />\n          <input name="cuit" placeholder="CUIT" required />\n        </div>\n        <div class="row" style="margin-top:10px;">\n          <input class="small" name="lim3" placeholder="Límite 3ros (ARS)" />\n          <input type="date" class="small" name="exp3" placeholder="Vto 3ros" />\n          <input class="small" name="limp" placeholder="Límite propio (ARS)" />\n          <input type="date" class="small" name="expp" placeholder="Vto propio" />\n          <input class="small" name="limf" placeholder="Límite FCE (ARS)" />\n        </div>\n        <div class="sub">Formato número: 1.234.567,89 · Vencimientos: YYYY-MM-DD</div>\n        <input name="observacion" placeholder="Observación (opcional)" style="margin-top:10px; width:100%; padding:10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff;" />\n        <button style="margin-top:10px;" type="submit">Crear firmante</button>\n      </form>\n    </div>\n\n    <div class="panel">\n      <h3>Alta masiva (CSV)</h3>\n      <div class="sub">Columnas: razon social, cuit, lim3, venc3, limp, vencp, limf, primera linea. En "primera linea": usar "si" o dejar vacío.</div>\n      <form method="post" action="{{ url_for(\'fp_upload\') }}" enctype="multipart/form-data">\n        <input type="file" name="file" accept=".csv,text/csv" required style="margin-top:8px;" />\n        <button style="margin-top:10px;" type="submit">Importar CSV</button>\n      </form>\n    </div>\n  </div>\n  {% endif %}\n\n  <div class="muted" style="margin-top:12px;">\n    Mostrando: <b>{{ firmantes|length }}</b>\n    {% if group_id %} <span class="pill" style="margin-left:8px;">Filtrado por grupo</span>{% endif %}\n    {% if first_line == \'yes\' %} <span class="pill pill-green" style="margin-left:8px;">Solo primera línea</span>{% elif first_line == \'no\' %} <span class="pill" style="margin-left:8px;">Excluye primera línea</span>{% endif %}\n  </div>\n\n  <table id="fp-main-table" data-exportable="1">\n    <thead>\n      <tr>\n        <th>Razón Social</th>\n        <th>CUIT</th>\n\n      {% if view_scope == \'3ros\' %}\n        <th class="w-num">Límite 3ros</th>\n        <th class="w-date">Vto 3ros</th>\n        <th class="w-num">Usado 3ros</th>\n        <th class="w-num">Bloq 3ros</th>\n        <th class="w-num">Disp 3ros</th>\n      {% elif view_scope == \'propio\' %}\n        <th class="w-num">Límite propio</th>\n        <th class="w-date">Vto propio</th>\n        <th class="w-num">Usado propio</th>\n        <th class="w-num">Bloq propio</th>\n        <th class="w-num">Disp propio</th>\n      {% else %}\n        <th class="w-num">Límite FCE</th>\n        <th class="w-date">Vto FCE</th>\n        <th class="w-num">Usado FCE</th>\n        <th class="w-num">Bloq FCE</th>\n        <th class="w-num">Disp FCE</th>\n      {% endif %}\n\n\n        <th>Estado</th>\n        <th>Acciones</th>\n      </tr>\n    </thead>\n    <tbody>\n    {% for f in firmantes %}\n      <tr class="{% if f._avail_any_neg %}rowneg{% endif %}{% if f._blocked_any %} rowblock{% endif %}">\n        <td>\n          <a class="link" href="{{ url_for(\'firmante_cartera\', cd=f.cuit_digits) }}">{{ f.razon_social }}</a>\n          {% if f._group_name %}<span class="pill" style="margin-left:8px;">{{ f._group_name }}</span>{% endif %}\n          {% if f.primera_linea %}<span class="pill pill-green" style="margin-left:8px;">PRIMERA LÍNEA</span>{% endif %}\n          {% if f._blocked_any %}<span class="lock-pill">🔒 con bloqueos</span>{% endif %}\n        </td>\n        <td>{{ f.cuit }}</td>\n\n      {% if view_scope == \'3ros\' %}\n        <td class="num w-num">{{ fmt_ars(f._lim3) }}</td>\n        <td class="num w-date">{{ f.limit_expiry_3ros or \'—\' }}</td>\n        <td class="num w-num">{{ fmt_ars(f._used3) }}</td>\n        <td class="num w-num">{{ fmt_ars(f._blocked3) }}</td>\n        <td class="num w-num {{ \'neg\' if f._avail3 < 0 else \'\' }}">{{ fmt_ars(f._avail3) }}</td>\n      {% elif view_scope == \'propio\' %}\n        <td class="num w-num">{{ fmt_ars(f._limp) }}</td>\n        <td class="num w-date">{{ f.limit_expiry_propio or \'—\' }}</td>\n        <td class="num w-num">{{ fmt_ars(f._usedp) }}</td>\n        <td class="num w-num">{{ fmt_ars(f._blockedp) }}</td>\n        <td class="num w-num {{ \'neg\' if f._availp < 0 else \'\' }}">{{ fmt_ars(f._availp) }}</td>\n      {% else %}\n        <td class="num w-num">{{ fmt_ars(f._limf) }}</td>\n        <td class="num w-date">—</td>\n        <td class="num w-num">{{ fmt_ars(f._usedf) }}</td>\n        <td class="num w-num">{{ fmt_ars(f._blockedf) }}</td>\n        <td class="num w-num {{ \'neg\' if f._availf < 0 else \'\' }}">{{ fmt_ars(f._availf) }}</td>\n      {% endif %}\n\n\n        <td>{% if f.is_active %}<span class="pill">Activo</span>{% else %}<span class="pill">Inactivo</span>{% endif %}</td>\n        <td>\n          <button class="btn-small" type="button" onclick="toggleDetails(\'{{ f.cuit_digits }}\')">Detalles</button>\n          {% if can_edit %}\n            {% if f.is_active %}\n              <form method="post" action="{{ url_for(\'fp_deactivate\', cd=f.cuit_digits) }}" style="display:inline;" onsubmit="return confirm(\'¿Desactivar firmante?\');">\n                <input name="observacion" placeholder="Observación (opcional)" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:220px;" />\n                <button class="btn-small btn-danger" type="submit">Desactivar</button>\n              </form>\n            {% else %}\n              <form method="post" action="{{ url_for(\'fp_reactivate\', cd=f.cuit_digits) }}" style="display:inline;" onsubmit="return confirm(\'¿Reactivar firmante?\');">\n                <input name="observacion" placeholder="Observación (opcional)" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:220px;" />\n                <button class="btn-small" type="submit">Reactivar</button>\n              </form>\n            {% endif %}\n          {% endif %}\n          {% if is_admin %}\n            <form method="post" action="{{ url_for(\'fp_delete\', cd=f.cuit_digits) }}" style="display:inline; margin-left:6px;" onsubmit="return confirm(\'¿ELIMINAR definitivamente este firmante de la DB? Esta acción no se puede deshacer.\');">\n              <input name="observacion" placeholder="Motivo eliminación (opcional)" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:220px;" />\n              <button class="btn-small btn-danger" type="submit">Eliminar DB</button>\n            </form>\n          {% endif %}\n        </td>\n      </tr>\n\n      <tr id="det-{{ f.cuit_digits }}" style="display:none;">\n        <td colspan="14">\n          <div class="details">\n            <h4>Acciones rápidas</h4>\n\n            <div class="row">\n              <div style="flex:1; min-width:300px;">\n                <div class="sub"><b>Bloqueos (hoy)</b></div>\n                <form method="post" action="{{ url_for(\'block_add\', cd=f.cuit_digits) }}" style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap;">\n                  <select name="scope" style="padding:10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff;">\n                    <option value="propio">Propio</option>\n                    <option value="3ros">3ros</option>\n                    <option value="fce">FCE</option>\n                  </select>\n                  <input name="amount" placeholder="Monto a bloquear (ej 250.000,00)" style="min-width:240px;" />\n                  <button class="btn-small" type="submit">Bloquear</button>\n                </form>\n\n                {% if f._blocks_list and (f._blocks_list|length > 0) %}\n                  <table>\n                    <thead><tr><th>Usuario</th><th>Tipo</th><th>Monto</th><th>Acción</th></tr></thead>\n                    <tbody>\n                      {% for b in f._blocks_list %}\n                        <tr>\n                          <td>{{ b.username }}</td>\n                          <td>{{ b.scope }}</td>\n                          <td class="num">{{ fmt_ars(b.amount) }}</td>\n                          <td>\n                            {% if b.can_delete %}\n                              <form method="post" action="{{ url_for(\'block_delete\', block_id=b.id) }}" style="display:inline;" onsubmit="return confirm(\'¿Eliminar bloqueo?\');">\n                                <button class="btn-small btn-danger" type="submit">Eliminar</button>\n                              </form>\n                            {% else %}\n                              <span class="sub">No autorizado</span>\n                            {% endif %}\n                          </td>\n                        </tr>\n                      {% endfor %}\n                    </tbody>\n                  </table>\n                {% else %}\n                  <div class="sub" style="margin-top:8px;">Sin bloqueos hoy.</div>\n                {% endif %}\n              </div>\n\n              {% if can_edit %}\n              <div style="flex:1; min-width:300px;">\n                <div class="sub"><b>Grupo económico</b></div>\n                <form method="post" action="{{ url_for(\'fp_set_group\', cd=f.cuit_digits) }}" style="margin-top:8px; display:flex; gap:8px; flex-wrap:wrap; align-items:center;">\n                  <input name="group_name" placeholder="Nombre del grupo" value="{{ f._group_name or \'\' }}" style="min-width:240px;" />\n                  <button class="btn-small" type="submit">Guardar</button>\n                  <button class="btn-small btn-danger" type="submit" name="clear_group" value="1" onclick="return confirm(\'¿Quitar grupo económico?\');">Quitar</button>\n                </form>\n                <div class="sub">Si el grupo no existe, el sistema te va a preguntar si querés registrarlo.</div>\n\n                <div class="sub" style="margin-top:12px;"><b>Clasificación</b></div>\n                <form method="post" action="{{ url_for(\'fp_set_first_line\', cd=f.cuit_digits, q=q, show_inactive=\'1\' if show_inactive else \'0\', group_id=group_id, first_line=first_line, view_scope=view_scope) }}" style="margin-top:8px; display:flex; gap:10px; flex-wrap:wrap; align-items:center;">\n                  <label style="display:flex; align-items:center; gap:8px; color:#e2e8f0; font-size:14px;">\n                    <input type="checkbox" name="primera_linea" value="1" {% if f.primera_linea %}checked{% endif %} />\n                    Es primera línea\n                  </label>\n                  <button class="btn-small" type="submit">Guardar clasificación</button>\n                </form>\n                <div class="sub">Tildado = el firmante queda clasificado como <b>PRIMERA LÍNEA</b>.</div>\n\n                <div class="sub" style="margin-top:12px;"><b>Modificar límites y vencimientos</b></div>\n                <form method="post" action="{{ url_for(\'fp_update_limits\', cd=f.cuit_digits) }}" style="margin-top:8px;">\n                  <div class="row">\n                {% if view_scope == \'3ros\' %}\n                  <div class="row">\n                    <input class="small" name="lim3" value="{{ f._lim3_raw }}" placeholder="Límite 3ros" />\n                    <input type="date" class="small" name="exp3" value="{{ f.limit_expiry_3ros or \'\' }}" />\n\n                    <input type="hidden" name="limp" value="{{ f._limp_raw }}" />\n                    <input type="hidden" name="expp" value="{{ f.limit_expiry_propio or \'\' }}" />\n                    <input type="hidden" name="limf" value="{{ f._limf_raw }}" />\n                  </div>\n                {% elif view_scope == \'propio\' %}\n                  <div class="row">\n                    <input class="small" name="limp" value="{{ f._limp_raw }}" placeholder="Límite propio" />\n                    <input type="date" class="small" name="expp" value="{{ f.limit_expiry_propio or \'\' }}" />\n\n                    <input type="hidden" name="lim3" value="{{ f._lim3_raw }}" />\n                    <input type="hidden" name="exp3" value="{{ f.limit_expiry_3ros or \'\' }}" />\n                    <input type="hidden" name="limf" value="{{ f._limf_raw }}" />\n                  </div>\n                {% else %}\n                  <div class="row">\n                    <input class="small" name="limf" value="{{ f._limf_raw }}" placeholder="Límite FCE" />\n\n                    <input type="hidden" name="lim3" value="{{ f._lim3_raw }}" />\n                    <input type="hidden" name="exp3" value="{{ f.limit_expiry_3ros or \'\' }}" />\n                    <input type="hidden" name="limp" value="{{ f._limp_raw }}" />\n                    <input type="hidden" name="expp" value="{{ f.limit_expiry_propio or \'\' }}" />\n                  </div>\n                {% endif %}\n\n                  </div>\n                  <input name="observacion" placeholder="Observación (opcional)" style="margin-top:10px; width:100%; padding:10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff;" />\n                <button style="margin-top:10px;" class="btn-small" type="submit">Guardar cambios</button>\n                </form>\n                <div class="sub">Se audita como <b>MODIFICACION</b>. Al vencer, se audita como <b>VENCIMIENTO</b>, el límite se conserva y la línea queda inactiva. Si luego cargás una nueva fecha futura o borrás el vencimiento, se audita como <b>RENOVACION</b>. FCE no vence.</div>\n              </div>\n              {% endif %}\n            </div>\n\n            <div style="margin-top:10px;">\n              <a class="btn-small" href="{{ url_for(\'firmante_cartera\', cd=f.cuit_digits) }}">Ver cartera (detalle)</a>\n            </div>\n\n          </div>\n        </td>\n      </tr>\n\n    {% endfor %}\n    </tbody>\n  </table>\n\n  <h3 style="margin-top:22px;">Cartera de facturas</h3>\n  <div class="footer">\n    <div>Filas FCE: <b>{{ fact_rows|length }}</b></div>\n  </div>\n\n  <table>\n    <thead>\n      <tr>\n        {% for h in fact_headers %}\n          <th>{{ h }}</th>\n        {% endfor %}\n      </tr>\n    </thead>\n    <tbody>\n      {% for r in fact_rows %}\n        <tr>\n          {% for v in r %}\n            <td>{{ v if v is not none else \'\' }}</td>\n          {% endfor %}\n        </tr>\n      {% endfor %}\n    </tbody>\n  </table>\n\n</div></div></div>\n\n<script>\nfunction toggleDetails(cd){\n  const tr = document.getElementById(\'det-\' + cd);\n  if(!tr) return;\n  tr.style.display = (tr.style.display === \'none\' || tr.style.display === \'\') ? \'table-row\' : \'none\';\n}\nfunction exportFP(){\n  const table = document.getElementById("fp-main-table");\n  if(!table){\n    alert("No encontré la tabla de firmantes para exportar.");\n    return;\n  }\n\n  const ths = Array.from(table.querySelectorAll("thead th"));\n  const visibleCols = [];\n  const headers = [];\n\n  ths.forEach((th, idx) => {\n    const txt = (th.innerText || "").trim().toLowerCase();\n    const visible = th.offsetParent !== null;\n    if(visible && txt !== "acciones"){\n      visibleCols.push(idx);\n      headers.push((th.innerText || "").trim());\n    }\n  });\n\n  const rows = [];\n  table.querySelectorAll("tbody tr").forEach(tr => {\n    if(tr.classList.contains("export-skip")) return;\n    if((tr.id || "").startsWith("det-")) return;\n\n    const tds = Array.from(tr.children);\n    if(!tds.length) return;\n\n    const row = visibleCols.map(i => {\n      const cell = tds[i];\n      return cell ? (cell.innerText || "").replace(/\\s+/g, " ").trim() : "";\n    });\n    rows.push(row);\n  });\n\n  const params = new URLSearchParams(window.location.search);\n  const meta = {\n    exported_at: new Date().toISOString(),\n    q: params.get("q") || "",\n    view_scope: params.get("view_scope") || "",\n    first_line: params.get("first_line") || "",\n    show_inactive: params.get("show_inactive") || ""\n  };\n\n  fetch("/fp/export_xlsx", {\n    method: "POST",\n    headers: {"Content-Type": "application/json"},\n    body: JSON.stringify({headers, rows, meta})\n  })\n  .then(async resp => {\n    if(!resp.ok){\n      const txt = await resp.text();\n      throw new Error(txt || "Error exportando XLSX");\n    }\n    return resp.blob();\n  })\n  .then(blob => {\n    const url = window.URL.createObjectURL(blob);\n    const a = document.createElement("a");\n    a.href = url;\n    a.download = "firmantes_precalificados.xlsx";\n    document.body.appendChild(a);\n    a.click();\n    a.remove();\n    window.URL.revokeObjectURL(url);\n  })\n  .catch(err => alert(err.message || "No se pudo exportar"));\n}\n</script>\n\n{{ live_search_js|safe }}\n</body></html>\n'
ALL_FIRMANTES_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Todos los firmantes</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1750px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n\n  .searchbar{ margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center; }\n  .searchbar input{\n    flex:1; min-width:280px; padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  .searchbar button, .searchbar a.btn{\n    padding:10px 14px; border-radius:12px; cursor:pointer; text-decoration:none;\n    border:none; background:#fff; color:#0b1220; font-weight:800;\n  }\n  .searchbar a.btn{\n    background: transparent; border:1px solid rgba(255,255,255,0.15); color:#fff; font-weight:400;\n  }\n\n  table{ width:100%; border-collapse: collapse; margin-top:14px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }\n  th{ color:#dbeafe; font-size:13px; }\n  .num{ white-space:nowrap; }\n\n  .pill{\n    display:inline-block; padding:6px 10px; border-radius:999px; font-size:12px;\n    border:1px solid rgba(255,255,255,0.15); color:#cbd5e1;\n  }\n\n  .btn-small{\n    padding:8px 10px; border-radius:12px; text-decoration:none; cursor:pointer;\n    border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff; font-size:13px;\n  }\n  .btn-danger{\n    border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12);\n  }\n\n  .subform{\n    display:flex; gap:6px; flex-wrap:wrap; align-items:center;\n  }\n  .subform input{\n    width:130px; padding:8px 10px; border-radius:10px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n\n  .actions{\n    display:flex; gap:8px; flex-wrap:wrap; align-items:center;\n  }\n    .subform{\n    display:flex;\n    flex-direction:column;\n    gap:6px;\n    }\n\n    .subform input{\n    width:120px;\n    }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Todos los firmantes</h2>\n  <div class="muted" style="margin-top:6px;">\n    Se muestran activos e inactivos. Desde acá podés modificar límites, ver detalle y activar/desactivar.\n  </div>\n\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n\n  <form method="get" action="{{ url_for(\'firmantes_all\') }}" class="searchbar">\n    <input name="q" value="{{ q or \'\' }}" placeholder="Buscar por CUIT o Razón Social..." />\n    <button type="submit">Buscar</button>\n    <a class="btn" href="{{ url_for(\'firmantes_all\') }}">Limpiar</a>\n    <a class="btn" href="{{ url_for(\'fp\') }}">Volver a FP</a>\n  </form>\n\n  <div class="muted" style="margin-top:10px;">Mostrando: <b>{{ rows|length }}</b></div>\n\n  <table>\n    <thead>\n      <tr>\n        <th>Razón Social</th>\n        <th>CUIT</th>\n        <th>Límite 3ros</th>\n        <th>Vto 3ros</th>\n        <th>Límite propio</th>\n        <th>Vto propio</th>\n        <th>Límite FCE</th>\n        <th>Estado</th>\n        <th>Editar límites</th>\n        <th>Acciones</th>\n      </tr>\n    </thead>\n    <tbody>\n    {% for r in rows %}\n      <tr>\n        <td>{{ r.razon_social }}</td>\n        <td>{{ r.cuit }}</td>\n        <td class="num">{{ fmt_ars(r.credit_limit_3ros or 0) }}</td>\n        <td>{{ r.limit_expiry_3ros or \'—\' }}</td>\n        <td class="num">{{ fmt_ars(r.credit_limit_propio or 0) }}</td>\n        <td>{{ r.limit_expiry_propio or \'—\' }}</td>\n        <td class="num">{{ fmt_ars(r.credit_limit_fce or 0) }}</td>\n        <td>{% if r.is_active %}<span class="pill">Activo</span>{% else %}<span class="pill">Inactivo</span>{% endif %}</td>\n        <td>\n            <form method="post" action="{{ url_for(\'firmantes_all_update\', cd=r.cuit_digits) }}" class="subform">\n\n            <div style="display:flex;align-items:center;gap:6px;">\n                <span style="font-size:12px;color:#cbd5e1;width:46px;">3ros</span>\n                <input name="lim3" value="{{ r.credit_limit_3ros or 0 }}" />\n                <input type="date" name="exp3" value="{{ r.limit_expiry_3ros or \'\' }}" />\n            </div>\n\n            <div style="display:flex;align-items:center;gap:6px;">\n                <span style="font-size:12px;color:#cbd5e1;width:46px;">Prop</span>\n                <input name="limp" value="{{ r.credit_limit_propio or 0 }}" />\n                <input type="date" name="expp" value="{{ r.limit_expiry_propio or \'\' }}" />\n            </div>\n\n            <div style="display:flex;align-items:center;gap:6px;">\n                <span style="font-size:12px;color:#cbd5e1;width:46px;">FCE</span>\n                <input name="limf" value="{{ r.credit_limit_fce or 0 }}" />\n            </div>\n\n            <input name="observacion" placeholder="Observación (opcional)" />\n            <button class="btn-small" type="submit">Guardar</button>\n\n            </form>\n        </td>\n        <td>\n          <div class="actions">\n            <a class="btn-small" href="{{ url_for(\'firmante_cartera\', cd=r.cuit_digits) }}">Ver detalle</a>\n            {% if r.is_active %}\n              <form method="post" action="{{ url_for(\'fp_deactivate\', cd=r.cuit_digits) }}" style="display:inline-block; margin:0;" onsubmit="return confirm(\'¿Desactivar firmante?\');">\n                <input name="observacion" placeholder="Observación (opcional)" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:220px;" />\n                <button class="btn-small btn-danger" type="submit">Desactivar</button>\n              </form>\n            {% else %}\n              <form method="post" action="{{ url_for(\'fp_reactivate\', cd=r.cuit_digits) }}" style="display:inline-block; margin:0;" onsubmit="return confirm(\'¿Reactivar firmante?\');">\n                <input name="observacion" placeholder="Observación (opcional)" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:220px;" />\n                <button class="btn-small" type="submit">Reactivar</button>\n              </form>\n            {% endif %}\n            {% if is_admin %}\n              <form method="post" action="{{ url_for(\'fp_delete\', cd=r.cuit_digits) }}" style="display:inline-block; margin:0;" onsubmit="return confirm(\'¿ELIMINAR definitivamente este firmante de la DB? Esta acción no se puede deshacer.\');">\n                <input name="observacion" placeholder="Motivo eliminación (opcional)" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:220px;" />\n                <button class="btn-small btn-danger" type="submit">Eliminar DB</button>\n              </form>\n            {% endif %}\n          </div>\n        </td>\n      </tr>\n    {% endfor %}\n    </tbody>\n  </table>\n\n</div></div></div>\n</body></html>\n'

FIRMANTE_CARTERA_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Cartera del firmante</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1650px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n  .grid{ display:flex; gap:12px; flex-wrap:wrap; margin-top:12px; }\n  .kpi{\n    flex:1; min-width:320px;\n    padding:14px; border-radius:16px;\n    background: rgba(255,255,255,0.05);\n    border:1px solid rgba(255,255,255,0.10);\n  }\n  .kpi .v{ font-size:20px; font-weight:900; margin-top:6px; white-space:nowrap; }\n  .neg{ color:#fecdd3; font-weight:900; }\n\n  .kpi-form{\n    margin-top:12px;\n    padding-top:12px;\n    border-top:1px solid rgba(255,255,255,0.10);\n  }\n  .kpi-form label{\n    display:block;\n    font-size:12px;\n    color:#cbd5e1;\n    margin:8px 0 4px 0;\n  }\n  .kpi-form input{\n    width:100%;\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  .kpi-form button{\n    margin-top:12px;\n    width:100%;\n    padding:10px 12px;\n    border:none; border-radius:12px;\n    background:#fff; color:#0b1220; font-weight:800; cursor:pointer;\n  }\n\n  .panel{\n    margin-top:14px;\n    background: rgba(255,255,255,0.05);\n    border:1px solid rgba(255,255,255,0.10);\n    border-radius:16px; padding:14px;\n  }\n  .panel h3{ margin:0 0 8px 0; }\n\n  table{ width:100%; border-collapse: collapse; margin-top:12px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); }\n  th{ color:#dbeafe; font-size:13px; }\n\n  input.colf{\n    margin-top:6px; width:100%; max-width:220px;\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n\n  .footer{\n    margin-top:10px; padding:12px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.12); background: rgba(255,255,255,0.04);\n    display:flex; gap:12px; flex-wrap:wrap; align-items:center; justify-content:space-between;\n  }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">{{ f.razon_social }}</h2>\n  <div class="muted" style="margin-top:6px;">CUIT: <b>{{ f.cuit }}</b></div>\n\n  {% if cartera_status and cartera_status != "OK" %}\n    <div class="flash-err">{{ cartera_status }}</div>\n  {% endif %}\n  {% if facturas_status and facturas_status != "OK" %}\n    <div class="flash-err">{{ facturas_status }}</div>\n  {% endif %}\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n\n  <div class="grid">\n    <div class="kpi">\n      <div class="muted">3ros · Límite / Vto</div>\n      <div class="v">{{ fmt_ars(lim3) }} · {{ exp3 or \'—\' }}</div>\n      <div class="muted">Usado: {{ fmt_ars(used3) }} · Bloq: {{ fmt_ars(blocked3) }}</div>\n      <div class="v {{ \'neg\' if avail3 < 0 else \'\' }}">Disp: {{ fmt_ars(avail3) }}</div>\n\n      <form method="post" action="{{ url_for(\'firmante_update_limits_from_cartera\', cd=f.cuit_digits) }}" class="kpi-form">\n        <label>Límite 3ros</label>\n        <input name="lim3" value="{{ f.credit_limit_3ros or 0 }}" placeholder="Límite 3ros" />\n\n        <label>Vto 3ros</label>\n        <input type="date" name="exp3" value="{{ f.limit_expiry_3ros or \'\' }}" />\n\n        <input type="hidden" name="limp" value="{{ f.credit_limit_propio or 0 }}" />\n        <input type="hidden" name="expp" value="{{ f.limit_expiry_propio or \'\' }}" />\n        <input type="hidden" name="limf" value="{{ f.credit_limit_fce or 0 }}" />\n\n        <label>Observación</label>\n        <input name="observacion" placeholder="Observación (opcional)" />\n\n        <button type="submit">Guardar 3ros</button>\n      </form>\n    </div>\n\n    <div class="kpi">\n      <div class="muted">Propio · Límite / Vto</div>\n      <div class="v">{{ fmt_ars(limp) }} · {{ expp or \'—\' }}</div>\n      <div class="muted">Usado: {{ fmt_ars(usedp) }} · Bloq: {{ fmt_ars(blockedp) }}</div>\n      <div class="v {{ \'neg\' if availp < 0 else \'\' }}">Disp: {{ fmt_ars(availp) }}</div>\n\n      <form method="post" action="{{ url_for(\'firmante_update_limits_from_cartera\', cd=f.cuit_digits) }}" class="kpi-form">\n        <label>Límite propio</label>\n        <input name="limp" value="{{ f.credit_limit_propio or 0 }}" placeholder="Límite propio" />\n\n        <label>Vto propio</label>\n        <input type="date" name="expp" value="{{ f.limit_expiry_propio or \'\' }}" />\n\n        <input type="hidden" name="lim3" value="{{ f.credit_limit_3ros or 0 }}" />\n        <input type="hidden" name="exp3" value="{{ f.limit_expiry_3ros or \'\' }}" />\n        <input type="hidden" name="limf" value="{{ f.credit_limit_fce or 0 }}" />\n\n        <label>Observación</label>\n        <input name="observacion" placeholder="Observación (opcional)" />\n\n        <button type="submit">Guardar propio</button>\n      </form>\n    </div>\n\n    <div class="kpi">\n      <div class="muted">FCE · Límite</div>\n      <div class="v">{{ fmt_ars(limf) }}</div>\n      <div class="muted">Usado: {{ fmt_ars(usedf) }} · Bloq: {{ fmt_ars(blockedf) }}</div>\n      <div class="v {{ \'neg\' if availf < 0 else \'\' }}">Disp: {{ fmt_ars(availf) }}</div>\n\n      <form method="post" action="{{ url_for(\'firmante_update_limits_from_cartera\', cd=f.cuit_digits) }}" class="kpi-form">\n        <label>Límite FCE</label>\n        <input name="limf" value="{{ f.credit_limit_fce or 0 }}" placeholder="Límite FCE" />\n\n        <input type="hidden" name="lim3" value="{{ f.credit_limit_3ros or 0 }}" />\n        <input type="hidden" name="exp3" value="{{ f.limit_expiry_3ros or \'\' }}" />\n        <input type="hidden" name="limp" value="{{ f.credit_limit_propio or 0 }}" />\n        <input type="hidden" name="expp" value="{{ f.limit_expiry_propio or \'\' }}" />\n\n                <label>Observación</label>\n        <input name="observacion" placeholder="Observación (opcional)" />\n\n<button type="submit">Guardar FCE</button>\n      </form>\n    </div>\n  </div>\n\n  <div class="panel">\n    <h3>Cartera de cheques</h3>\n    <div class="footer">\n      <div>Filas visibles: <b id="rowCountCheq">{{ rows|length }}</b></div>\n      <div>Suma Importe visible: <b id="sumImporteCheq">$ 0,00</b></div>\n    </div>\n\n    <table id="tCheq">\n      <thead>\n        <tr>\n          {% for h in headers %}\n            <th>{{ h }}<div><input class="colf cheqf" data-col="{{ loop.index0 }}" placeholder="filtrar..." /></div></th>\n          {% endfor %}\n        </tr>\n      </thead>\n      <tbody>\n        {% for r in rows %}\n          <tr>{% for v in r %}<td>{{ v if v is not none else \'\' }}</td>{% endfor %}</tr>\n        {% endfor %}\n      </tbody>\n    </table>\n  </div>\n\n  <div class="panel">\n    <h3>Cartera de facturas</h3>\n    <div class="footer">\n      <div>Filas visibles: <b id="rowCountFce">{{ fact_rows|length }}</b></div>\n      <div>Suma Importe visible: <b id="sumImporteFce">$ 0,00</b></div>\n    </div>\n\n    <table id="tFce">\n      <thead>\n        <tr>\n          {% for h in fact_headers %}\n            <th>{{ h }}<div><input class="colf fcef" data-col="{{ loop.index0 }}" placeholder="filtrar..." /></div></th>\n          {% endfor %}\n        </tr>\n      </thead>\n      <tbody>\n        {% for r in fact_rows %}\n          <tr>{% for v in r %}<td>{{ v if v is not none else \'\' }}</td>{% endfor %}</tr>\n        {% endfor %}\n      </tbody>\n    </table>\n  </div>\n\n</div></div></div>\n\n<script>\n(function(){\n  function parseEsNumber(s){\n    if(s == null) return 0;\n    s = String(s).trim();\n    if(!s) return 0;\n    s = s.replace(/\\$/g,\'\').replace(/\\s+/g,\'\');\n    let neg = false;\n    if(s.startsWith(\'(\') && s.endsWith(\')\')){ neg=true; s=s.slice(1,-1); }\n    if(s.includes(\',\')){\n      s = s.replace(/\\./g,\'\');\n      s = s.replace(\',\', \'.\');\n    }else{\n      const dots = (s.match(/\\./g)||[]).length;\n      if(dots > 1) s = s.replace(/\\./g,\'\');\n    }\n    let v = parseFloat(s);\n    if(isNaN(v)) v = 0;\n    return neg ? -v : v;\n  }\n\n  function formatARS(x){\n    const neg = x < 0;\n    x = Math.abs(x);\n    let s = x.toLocaleString(\'en-US\', {minimumFractionDigits:2, maximumFractionDigits:2});\n    s = s.replace(/,/g,\'X\').replace(/\\./g,\',\').replace(/X/g,\'.\');\n    return \'$ \' + (neg ? \'-\' : \'\') + s;\n  }\n\n  function bindTable(tableId, inputSelector, rowCountId, sumId){\n    const table = document.getElementById(tableId);\n    if(!table) return;\n    const inputs = table.querySelectorAll(inputSelector);\n    const tbody = table.querySelector(\'tbody\');\n    const rows = Array.from(tbody.querySelectorAll(\'tr\'));\n    const sumEl = document.getElementById(sumId);\n    const countEl = document.getElementById(rowCountId);\n\n    const headers = Array.from(table.querySelectorAll(\'thead th\')).map(th => (th.childNodes[0]?.textContent || \'\').trim().toLowerCase());\n    let idxImporte = headers.findIndex(h => h.includes(\'importe\') || h.includes(\'monto\'));\n    if(idxImporte < 0) idxImporte = null;\n\n    function apply(){\n      const filters = Array.from(inputs).map(i => (i.value || \'\').toLowerCase());\n      let visible = 0;\n      let sum = 0;\n\n      rows.forEach(tr => {\n        const tds = Array.from(tr.querySelectorAll(\'td\'));\n        let ok = true;\n        for(let i=0;i<filters.length;i++){\n          const f = filters[i];\n          if(!f) continue;\n          const val = (tds[i]?.textContent || \'\').toLowerCase();\n          if(!val.includes(f)){ ok=false; break; }\n        }\n        tr.style.display = ok ? \'\' : \'none\';\n        if(ok){\n          visible++;\n          if(idxImporte !== null){\n            const raw = tds[idxImporte]?.textContent || \'\';\n            sum += parseEsNumber(raw);\n          }\n        }\n      });\n\n      if(countEl) countEl.textContent = String(visible);\n      if(sumEl) sumEl.textContent = formatARS(sum);\n    }\n\n    inputs.forEach(i => i.addEventListener(\'input\', apply));\n    apply();\n  }\n\n  bindTable(\'tCheq\', \'.cheqf\', \'rowCountCheq\', \'sumImporteCheq\');\n  bindTable(\'tFce\', \'.fcef\', \'rowCountFce\', \'sumImporteFce\');\n})();\n</script>\n\n</body></html>\n'

AUDIT_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Auditoría ABM</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1500px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n  .searchbar{ margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center; }\n  .searchbar input{\n    flex: 1; min-width: 260px;\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  .searchbar button, .searchbar a.btn{\n    padding:10px 14px; border-radius:12px; cursor:pointer; text-decoration:none;\n    border:none; background:#fff; color:#0b1220; font-weight:800;\n  }\n  .searchbar a.btn{ background: transparent; border:1px solid rgba(255,255,255,0.15); color:#fff; font-weight:400; }\n  table{ width:100%; border-collapse: collapse; margin-top: 14px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }\n  th{ color:#dbeafe; font-size:13px; }\n  .pill{\n    display:inline-block; padding:6px 10px; border-radius:999px; font-size:12px;\n    border:1px solid rgba(255,255,255,0.15); color:#cbd5e1;\n  }\n  .a{ color:#93c5fd; text-decoration:none; }\n  .a:hover{ text-decoration: underline; }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Auditoría ABM</h2>\n  <div class="muted" style="margin-top:6px;">Conceptos: <b>ALTA</b>, <b>BAJA</b>, <b>MODIFICACION</b>, <b>VENCIMIENTO</b>.</div>\n\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n\n  {% if is_admin %}\n  <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center;">\n    <form method="post" action="{{ url_for(\'audit_delete_all\') }}" onsubmit="return confirm(\'¿Borrar TODOS los registros ABM? Esta acción no se puede deshacer.\');">\n      <button type="submit" class="btn-small btn-danger">🗑️ Borrar TODOS los registros ABM</button>\n    </form>\n  </div>\n  {% endif %}\n\n  <form method="get" action="{{ url_for(\'audit\') }}" class="searchbar" data-live-search="1" data-live-delay="350">\n    <input class="live-search" name="q" value="{{ q or \'\' }}" placeholder="Buscar por CUIT o Razón Social..." />\n    <button type="submit">Buscar</button>\n    <a class="btn" href="{{ url_for(\'audit\') }}">Limpiar</a>\n  </form>\n\n  <div class="muted" style="margin-top:10px;">Mostrando: <b>{{ rows|length }}</b></div>\n\n  <table>\n    <thead><tr>\n      <th>Fecha</th><th>Firmante</th><th>Acción</th><th>Usuario</th><th>Detalle</th>\n      {% if is_admin %}<th>Admin</th>{% endif %}\n    </tr></thead>\n    <tbody>\n    {% for r in rows %}\n      <tr>\n        <td>{{ r.created_at }}</td>\n        <td>\n          {% if r.razon_social %}\n            <a class="a" href="{{ url_for(\'firmante_cartera\', cd=r.firmante_cuit_digits) }}">{{ r.razon_social }}</a>\n            <div class="muted" style="font-size:12px;">{{ r.cuit }}</div>\n          {% else %}\n            {{ r.firmante_cuit_digits }}\n          {% endif %}\n        </td>\n\n        <td>{{ r.action }}</td>\n        <td>{{ r.username }}</td>\n        <td>{{ r.details }}</td>\n\n        {% if is_admin %}\n          <td>\n            <form method="post" action="{{ url_for(\'audit_delete\', audit_id=r.id) }}"\n                  style="display:inline;"\n                  onsubmit="return confirm(\'¿Eliminar este registro ABM?\');">\n              <button type="submit" class="btn-small btn-danger">Eliminar</button>\n            </form>\n          </td>\n        {% endif %}\n      </tr>\n    {% endfor %}\n    </tbody>\n\n  </table>\n\n</div></div></div>\n{{ live_search_js|safe }}\n</body></html>\n'

CARTERA_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Cartera</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1600px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n  .searchbar{ margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center; }\n  .searchbar select, .searchbar button{ padding:10px 14px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; }\n  .searchbar button{ cursor:pointer; background:#fff; color:#0b1220; border:none; font-weight:800; }\n  table{ width:100%; border-collapse: collapse; margin-top:12px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }\n  th{ color:#dbeafe; font-size:13px; }\n  input.colf{ margin-top:6px; width:100%; max-width:220px; padding:10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; }\n  .footer{ margin-top:10px; padding:12px; border-radius:12px; border:1px solid rgba(255,255,255,0.12); background: rgba(255,255,255,0.04); display:flex; gap:12px; flex-wrap:wrap; align-items:center; justify-content:space-between; }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Cartera</h2>\n  <div class="muted" style="margin-top:6px;">Elegí entre cheques y facturas.</div>\n  {% if status and status != "OK" %}<div class="flash-err">{{ status }}</div>{% endif %}\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n  <form method="get" action="{{ url_for(\'cartera\') }}" class="searchbar">\n    <select name="kind" onchange="this.form.submit()">\n      <option value="cheques" {{ \'selected\' if view_kind==\'cheques\' else \'\' }}>Cheques</option>\n      <option value="facturas" {{ \'selected\' if view_kind==\'facturas\' else \'\' }}>facturas</option>\n    </select>\n    <button type="submit">Ver</button>\n  </form>\n  <div class="footer">\n    <div>Filas visibles: <b id="rowCount">{{ (rows|length) if view_kind==\'cheques\' else (fact_rows|length) }}</b></div>\n    <div>Suma Importe visible: <b id="sumImporte">$ 0,00</b></div>\n  </div>\n  <table id="t">\n    <thead><tr>\n      {% for h in (headers if view_kind==\'cheques\' else fact_headers) %}<th>{{ h }}<div><input class="colf" placeholder="filtrar..." /></div></th>{% endfor %}\n    </tr></thead>\n    <tbody>\n      {% if view_kind==\'cheques\' %}{% for r in rows %}<tr>{% for v in r %}<td>{{ v if v is not none else \'\' }}</td>{% endfor %}</tr>{% endfor %}\n      {% else %}{% for r in fact_rows %}<tr>{% for v in r %}<td>{{ v if v is not none else \'\' }}</td>{% endfor %}</tr>{% endfor %}{% endif %}\n    </tbody>\n  </table>\n</div></div></div>\n<script>\n(function(){\n  const table=document.getElementById(\'t\'); if(!table) return;\n  const inputs=table.querySelectorAll(\'input.colf\'); const tbody=table.querySelector(\'tbody\'); const rows=Array.from(tbody.querySelectorAll(\'tr\'));\n  const sumEl=document.getElementById(\'sumImporte\'); const countEl=document.getElementById(\'rowCount\');\n  const headers=Array.from(table.querySelectorAll(\'thead th\')).map(th => (th.childNodes[0]?.textContent || \'\').trim().toLowerCase());\n  let idxImporte=headers.findIndex(h => h.includes(\'importe\') || h.includes(\'monto\')); if(idxImporte < 0) idxImporte = null;\n  function parseEsNumber(s){ if(s == null) return 0; s=String(s).trim(); if(!s) return 0; s=s.replace(/\\$/g,\'\').replace(/\\s+/g,\'\'); let neg=false; if(s.startsWith(\'(\') && s.endsWith(\')\')){ neg=true; s=s.slice(1,-1);} if(s.includes(\',\')){ s=s.replace(/\\./g,\'\'); s=s.replace(\',\', \'.\'); } else { const dots=(s.match(/\\./g)||[]).length; if(dots>1) s=s.replace(/\\./g,\'\'); } let v=parseFloat(s); if(isNaN(v)) v=0; return neg ? -v : v; }\n  function formatARS(x){ const neg=x<0; x=Math.abs(x); let s=x.toLocaleString(\'en-US\',{minimumFractionDigits:2,maximumFractionDigits:2}); s=s.replace(/,/g,\'X\').replace(/\\./g,\',\').replace(/X/g,\'.\'); return \'$ \' + (neg ? \'-\' : \'\') + s; }\n  function apply(){ const filters=Array.from(inputs).map(i => (i.value || \'\').toLowerCase()); let visible=0,sum=0; rows.forEach(tr => { const tds=Array.from(tr.querySelectorAll(\'td\')); let ok=true; for(let i=0;i<filters.length;i++){ const f=filters[i]; if(!f) continue; const val=(tds[i]?.textContent || \'\').toLowerCase(); if(!val.includes(f)){ ok=false; break;} } tr.style.display = ok ? \'\' : \'none\'; if(ok){ visible++; if(idxImporte !== null){ sum += parseEsNumber(tds[idxImporte]?.textContent || \'\'); } } }); if(countEl) countEl.textContent=String(visible); if(sumEl) sumEl.textContent=formatARS(sum); }\n  inputs.forEach(i => i.addEventListener(\'input\', apply)); apply();\n})();\n</script>\n</body></html>\n'

USERS_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Usuarios</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1100px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n  label{ display:block; margin-top:12px; color:#cbd5e1; font-size:13px; }\n  input, select{\n    width:100%; padding:12px; margin-top:6px;\n    border-radius:12px; border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  button{\n    margin-top:14px; padding:12px 14px; border-radius:12px; cursor:pointer;\n    border:none; background:#fff; color:#0b1220; font-weight:800;\n  }\n  table{ width:100%; border-collapse: collapse; margin-top: 14px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }\n  th{ color:#dbeafe; font-size:13px; }\n  .pill{\n    display:inline-block; padding:6px 10px; border-radius:999px; font-size:12px;\n    border:1px solid rgba(255,255,255,0.15); color:#cbd5e1;\n  }\n  .btn-small{\n    padding:8px 10px; border-radius:12px; cursor:pointer; font-size:13px;\n    border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff;\n  }\n  .btn-danger{\n    border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12);\n  }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Usuarios</h2>\n  <div class="muted" style="margin-top:6px;">Admin puede crear usuarios y desactivarlos (no se borran).</div>\n\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n\n  <h3 style="margin-top:16px;">Crear usuario</h3>\n  <form method="post" action="{{ url_for(\'user_create\') }}">\n    <label>Username</label><input name="username" required />\n    <label>Password</label><input name="password" type="password" required />\n    <label>Rol</label>\n    <select name="role">\n      <option value="admin">Admin</option>\n      <option value="risk">Analista de riesgos</option>\n      <option value="sales">Ejecutivo comercial</option>\n    </select>\n    <button type="submit">Crear</button>\n  </form>\n\n  <h3 style="margin-top:18px;">Listado</h3>\n  <table>\n    <thead><tr><th>Usuario</th><th>Rol</th><th>Estado</th><th>Acción</th></tr></thead>\n    <tbody>\n    {% for u in users %}\n      <tr>\n        <td>{{ u.username }}</td>\n        <td><span class="pill">{{ labels.get(u.role,u.role) }}</span></td>\n        <td>{% if u.is_active %}<span class="pill">Activo</span>{% else %}<span class="pill">Inactivo</span>{% endif %}</td>\n        <td>\n          {% if u.username != current_username %}\n            {% if u.is_active %}\n              <form method="post" action="{{ url_for(\'user_deactivate\', username=u.username) }}" style="display:inline;" onsubmit="return confirm(\'¿Desactivar usuario?\');">\n                <button class="btn-small btn-danger" type="submit">Desactivar</button>\n              </form>\n            {% else %}\n              <form method="post" action="{{ url_for(\'user_reactivate\', username=u.username) }}" style="display:inline;" onsubmit="return confirm(\'¿Reactivar usuario?\');">\n                <button class="btn-small" type="submit">Reactivar</button>\n              </form>\n            {% endif %}\n          {% else %}\n            <span class="muted" style="font-size:12px;">(vos)</span>\n          {% endif %}\n        </td>\n      </tr>\n    {% endfor %}\n    </tbody>\n  </table>\n</div></div></div>\n</body></html>\n'

GROUP_CONFIRM_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Registrar grupo económico</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 900px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n  .pillbtn{\n    padding:10px 14px; border-radius:999px; cursor:pointer;\n    border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff; font-weight:700;\n  }\n  .pillbtn.yes{ border:1px solid rgba(34,197,94,0.35); background: rgba(34,197,94,0.12); }\n  .pillbtn.no{ border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12); }\n  input{\n    padding:12px; border-radius:12px; border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff; min-width:260px;\n  }\n  button{\n    padding:12px 14px; border-radius:12px; cursor:pointer; border:none; background:#fff; color:#0b1220; font-weight:900;\n  }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Grupo económico no registrado</h2>\n  <div class="muted" style="margin-top:6px;">\n    El grupo económico <b>{{ group_name }}</b> no está registrado. ¿Querés registrarlo?\n  </div>\n\n  <div style="margin-top:14px; display:flex; gap:10px; flex-wrap:wrap; align-items:center;">\n    <form method="post" action="{{ url_for(\'group_confirm_no\') }}">\n      <input type="hidden" name="cd" value="{{ cd }}" />\n      <input type="hidden" name="group_name" value="{{ group_name }}" />\n      <button class="pillbtn no" type="submit">No</button>\n    </form>\n\n    <form method="post" action="{{ url_for(\'group_confirm_yes\') }}" style="display:flex; gap:10px; flex-wrap:wrap; align-items:center;">\n      <input type="hidden" name="cd" value="{{ cd }}" />\n      <input type="hidden" name="group_name" value="{{ group_name }}" />\n      <input name="limit" placeholder="Límite grupal (ARS)" required />\n      <button type="submit">Crear grupo</button>\n    </form>\n  </div>\n\n  <div class="muted" style="margin-top:12px; font-size:12px;">\n    Si elegís “No”, el firmante queda sin grupo.\n  </div>\n</div></div></div>\n</body></html>\n'

FP_UPLOAD_CONFIRM_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Confirmar grupos (CSV)</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1100px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  table{ width:100%; border-collapse:collapse; margin-top:12px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align:top; }\n  th{ color:#dbeafe; font-size:13px; }\n  input{\n    padding:10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff; width: 220px;\n  }\n  .pillbtn{\n    padding:8px 12px; border-radius:999px; cursor:pointer;\n    border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff; font-weight:800;\n  }\n  .yes{ border:1px solid rgba(34,197,94,0.35); background: rgba(34,197,94,0.12); }\n  .no{ border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12); }\n  button{\n    margin-top:14px; padding:12px 14px; border-radius:12px; cursor:pointer; border:none; background:#fff; color:#0b1220; font-weight:900;\n  }\n  .muted{ color:#cbd5e1; }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Confirmación de grupos no registrados</h2>\n  <div class="muted" style="margin-top:6px;">\n    Se detectaron grupos en el CSV que no están registrados. Elegí si querés crearlos (con límite) o ignorarlos (se deja el campo grupo vacío).\n  </div>\n\n  {% if flash %}<div class="{{ \'flash-ok\' if flash.kind==\'ok\' else \'flash-err\' }}">{{ flash.msg }}</div>{% endif %}\n\n  <form method="post" action="{{ url_for(\'fp_upload_confirm_apply\') }}">\n    <input type="hidden" name="batch_id" value="{{ batch_id }}" />\n\n    <table>\n      <thead>\n        <tr>\n          <th>Grupo</th>\n          <th>Acción</th>\n          <th>Límite grupal (si crea)</th>\n        </tr>\n      </thead>\n      <tbody>\n        {% for g in unknown_groups %}\n          <tr>\n            <td><b>{{ g }}</b></td>\n            <td>\n              <label><input type="radio" name="action__{{ loop.index0 }}" value="yes" checked /> <span class="pillbtn yes">Crear</span></label>\n              <label style="margin-left:10px;"><input type="radio" name="action__{{ loop.index0 }}" value="no" /> <span class="pillbtn no">No</span></label>\n              <input type="hidden" name="name__{{ loop.index0 }}" value="{{ g }}" />\n            </td>\n            <td><input name="limit__{{ loop.index0 }}" placeholder="1.000.000,00" /></td>\n          </tr>\n        {% endfor %}\n      </tbody>\n    </table>\n\n    <div class="muted" style="margin-top:10px; font-size:12px;">\n      Nota: si elegís “Crear”, el límite es obligatorio (podés poner 0).\n    </div>\n\n    <button type="submit">Aplicar importación</button>\n  </form>\n\n</div></div></div>\n</body></html>\n'

GROUPS_HTML = """
<!doctype html><html><head><meta charset="utf-8" />
<title>Grupos Económicos</title>
<style>
  *,*::before,*::after{ box-sizing:border-box; }
  {{ toolbar_css|safe }}
  
.flash-ok{
  margin-top:12px; padding:12px 14px; border-radius:12px;
  border:1px solid rgba(34,197,94,0.35);
  background: rgba(34,197,94,0.10);
  color:#dcfce7;
}
.flash-err{
  margin-top:12px; padding:12px 14px; border-radius:12px;
  border:1px solid rgba(244,63,94,0.35);
  background: rgba(244,63,94,0.10);
  color:#ffe4e6;
}

  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }
  .page{ padding:92px 28px 28px 28px; }
  .wrap{ max-width: 1600px; margin:0 auto; }
  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }
  .muted{ color:#cbd5e1; }
  .grid2{ display:grid; grid-template-columns: 1fr 1fr; gap:12px; margin-top:12px; }
  @media(max-width: 1100px){ .grid2{ grid-template-columns: 1fr; } }
  .panel{
    background: rgba(255,255,255,0.05);
    border:1px solid rgba(255,255,255,0.10);
    border-radius:16px; padding:14px;
  }
  .panel h3{ margin:0 0 8px 0; }
  .row{ display:flex; gap:10px; flex-wrap:wrap; }
  .row input{
    flex:1; min-width: 220px;
    padding:10px; border-radius:12px;
    border:1px solid rgba(255,255,255,0.15);
    background: rgba(0,0,0,0.2); color:#fff;
  }
  .searchbar{ margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center; }
  .searchbar input{
    flex:1; min-width: 280px;
    padding:10px; border-radius:12px;
    border:1px solid rgba(255,255,255,0.15);
    background: rgba(0,0,0,0.2); color:#fff;
  }
  table{ width:100%; border-collapse: collapse; margin-top: 14px; }
  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }
  th{ color:#dbeafe; font-size:13px; }
  .num, .num * { white-space: nowrap; }
  .w-num{ min-width: 170px; }
  .neg{ color:#fecdd3; font-weight:800; }
  .rowneg td{ background: rgba(244,63,94,0.06); }
  .rowblock td{ background: rgba(245,158,11,0.06); }
  .lock-pill{
    display:inline-block; margin-left:8px; padding:3px 10px; border-radius:999px; font-size:12px;
    border:1px solid rgba(245,158,11,0.35); background: rgba(245,158,11,0.12); color:#fde68a;
  }
  a.link{ color:#93c5fd; text-decoration:none; }
  a.link:hover{ text-decoration: underline; }
  button.btn-small{
    padding:8px 10px; border-radius:12px; cursor:pointer; font-size:13px;
    border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff;
  }
  .btn-danger{
    border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12);
  }
</style></head><body>
{{ toolbar_html|safe }}
<div class="page"><div class="wrap"><div class="card">
  <div style="display:flex; justify-content:space-between; align-items:flex-end; flex-wrap:wrap; gap:12px;">
    <div>
      <h2 style="margin:0;">Grupos Económicos</h2>
      <div class="muted"></div>
    </div>
    <div class="muted">Usuario: <b>{{ username }}</b> · Rol: <b>{{ role_label }}</b></div>
  </div>

  {% if flash %}<div class="{{ 'flash-ok' if flash.kind=='ok' else 'flash-err' }}">{{ flash.msg }}</div>{% endif %}

  {% if can_edit %}
  <div class="grid2">
    <div class="panel">
      <h3>Alta manual</h3>
      <form method="post" action="{{ url_for('groups_create') }}">
        <div class="row">
          <input name="nombre" placeholder="Nombre del grupo" required />
          <input name="limite" placeholder="Límite grupal (ARS)" />
        </div>
        <div class="muted" style="margin-top:6px; font-size:12px;">Formato número: 1.234.567,89</div>
        <button style="margin-top:10px;" type="submit" class="btn-small">Crear grupo</button>
      </form>
    </div>

    <div class="panel">
      <h3>Alta/Actualización masiva (CSV)</h3>
      <div class="muted" style="font-size:12px;">Columnas: grupo(nombre), limite</div>
      <form method="post" action="{{ url_for('groups_upload') }}" enctype="multipart/form-data">
        <input type="file" name="file" accept=".csv,text/csv" required style="margin-top:8px;" />
        <button style="margin-top:10px;" type="submit" class="btn-small">Importar CSV</button>
      </form>
    </div>
  </div>
  {% endif %}

  <div class="searchbar">
    <input id="groupsLiveSearch" type="text" placeholder="Buscar grupo o importes..." autocomplete="off" />
    <div class="muted">Mostrando: <b id="groupsRowCount">{{ rows|length }}</b></div>
  </div>

  <table id="groupsTable">
    <thead>
      <tr>
        <th>Grupo</th>
        <th class="w-num">Límite grupal</th>
        <th class="w-num">Usado grupal</th>
        <th class="w-num">Bloqueos</th>
        <th class="w-num">Disponible</th>
        <th>Acciones</th>
      </tr>
    </thead>
    <tbody>
      {% for g in rows %}
        <tr class="{% if g._avail < 0 %}rowneg{% endif %}{% if g._blocked_total > 0 %} rowblock{% endif %}">
          <td>
            <a class="link" href="{{ url_for('group_cartera', group_id=g.id) }}">{{ g.nombre }}</a>
            {% if g._blocked_total > 0 %}<span class="lock-pill">🔒 con bloqueos</span>{% endif %}
          </td>
          <td class="num">{{ fmt_ars(g._lim) }}</td>
          <td class="num">{{ fmt_ars(g._used) }}</td>
          <td class="num">{{ fmt_ars(g._blocked_total) }}</td>
          <td class="num {{ 'neg' if g._avail < 0 else '' }}">{{ fmt_ars(g._avail) }}</td>
          <td>
            <a class="btn-small" href="{{ url_for('fp') }}?group_id={{ g.id }}">Ver grupo</a>
            {% if can_edit %}
              <form method="post" action="{{ url_for('groups_update_limit', group_id=g.id) }}" style="display:inline; margin-left:6px;">
                <input name="limite" placeholder="Nuevo límite" style="padding:8px 10px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; width:170px;" />
                <button class="btn-small" type="submit">Guardar</button>
              </form>
            {% endif %}
          </td>
        </tr>
      {% endfor %}
    </tbody>
  </table>

</div></div></div>
<script>
(function(){
  const input = document.getElementById('groupsLiveSearch');
  const table = document.getElementById('groupsTable');
  if(!input || !table) return;
  const rows = Array.from(table.querySelectorAll('tbody tr'));
  const count = document.getElementById('groupsRowCount');

  function apply(){
    const q = (input.value || '').trim().toLowerCase();
    let visible = 0;
    rows.forEach(tr => {
      const txt = (tr.innerText || tr.textContent || '').toLowerCase();
      const ok = !q || txt.includes(q);
      tr.style.display = ok ? '' : 'none';
      if(ok) visible += 1;
    });
    if(count) count.textContent = String(visible);
  }

  input.addEventListener('input', apply);
  apply();
})();
</script>
</body></html>
"""

GROUP_CARTERA_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Cartera del grupo</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 28px 28px 28px; }\n  .wrap{ max-width: 1600px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .muted{ color:#cbd5e1; }\n  table{ width:100%; border-collapse: collapse; margin-top:12px; }\n  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }\n  th{ color:#dbeafe; font-size:13px; }\n  input.colf{\n    margin-top:6px;\n    width: 100%;\n    max-width: 220px;\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n  }\n  .footer{\n    margin-top:10px; padding:12px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.12); background: rgba(255,255,255,0.04);\n    display:flex; gap:12px; flex-wrap:wrap; align-items:center; justify-content:space-between;\n  }\n</style></head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap"><div class="card">\n  <h2 style="margin:0;">Cartera del grupo: {{ group.nombre }}</h2>\n  <div class="muted" style="margin-top:6px;">Empresas: <b>{{ firmantes_count }}</b></div>\n\n  {% if status and status != "OK" %}\n    <div class="flash-err">{{ status }}</div>\n  {% endif %}\n\n  <div class="footer">\n    <div>Filas visibles: <b id="rowCount">{{ rows|length }}</b></div>\n    <div>Suma Importe visible: <b id="sumImporte">$ 0,00</b></div>\n  </div>\n\n  <table id="t">\n    <thead><tr>\n      {% for h in headers %}\n        <th>{{ h }}<div><input class="colf" placeholder="filtrar..." /></div></th>\n      {% endfor %}\n    </tr></thead>\n    <tbody>\n      {% for r in rows %}\n        <tr>{% for v in r %}<td>{{ v if v is not none else \'\' }}</td>{% endfor %}</tr>\n      {% endfor %}\n    </tbody>\n  </table>\n\n</div></div></div>\n\n<script>\n(function(){\n  const table = document.getElementById(\'t\');\n  if(!table) return;\n  const inputs = table.querySelectorAll(\'input.colf\');\n  const tbody = table.querySelector(\'tbody\');\n  const rows = Array.from(tbody.querySelectorAll(\'tr\'));\n  const sumEl = document.getElementById(\'sumImporte\');\n  const countEl = document.getElementById(\'rowCount\');\n\n  const headers = Array.from(table.querySelectorAll(\'thead th\')).map(th => (th.childNodes[0]?.textContent || \'\').trim().toLowerCase());\n  let idxImporte = headers.findIndex(h => h.includes(\'importe\') || h.includes(\'monto\'));\n  if(idxImporte < 0) idxImporte = null;\n\n  function parseEsNumber(s){\n    if(s == null) return 0;\n    s = String(s).trim();\n    if(!s) return 0;\n    s = s.replace(/\\$/g,\'\').replace(/\\s+/g,\'\');\n    let neg = false;\n    if(s.startsWith(\'(\') && s.endsWith(\')\')){ neg=true; s=s.slice(1,-1); }\n    if(s.includes(\',\')){\n      s = s.replace(/\\./g,\'\');\n      s = s.replace(\',\', \'.\');\n    }else{\n      const dots = (s.match(/\\./g)||[]).length;\n      if(dots > 1) s = s.replace(/\\./g,\'\');\n    }\n    let v = parseFloat(s);\n    if(isNaN(v)) v = 0;\n    return neg ? -v : v;\n  }\n\n  function formatARS(x){\n    const neg = x < 0;\n    x = Math.abs(x);\n    let s = x.toLocaleString(\'en-US\', {minimumFractionDigits:2, maximumFractionDigits:2});\n    s = s.replace(/,/g,\'X\').replace(/\\./g,\',\').replace(/X/g,\'.\');\n    return \'$ \' + (neg ? \'-\' : \'\') + s;\n  }\n\n  function apply(){\n    const filters = Array.from(inputs).map(i => (i.value || \'\').toLowerCase());\n    let visible = 0;\n    let sum = 0;\n\n    rows.forEach(tr => {\n      const tds = Array.from(tr.querySelectorAll(\'td\'));\n      let ok = true;\n      for(let i=0;i<filters.length;i++){\n        const f = filters[i];\n        if(!f) continue;\n        const val = (tds[i]?.textContent || \'\').toLowerCase();\n        if(!val.includes(f)){ ok=false; break; }\n      }\n      tr.style.display = ok ? \'\' : \'none\';\n      if(ok){\n        visible++;\n        if(idxImporte !== null){\n          const raw = tds[idxImporte]?.textContent || \'\';\n          sum += parseEsNumber(raw);\n        }\n      }\n    });\n\n    if(countEl) countEl.textContent = String(visible);\n    if(sumEl) sumEl.textContent = formatARS(sum);\n  }\n\n  inputs.forEach(i => i.addEventListener(\'input\', apply));\n  apply();\n})();\n</script>\n\n</body></html>\n'

# --------------------------- | Routes: auth | ---------------------------

@app.route("/", endpoint="root")
def root():
    if current_user():
        return redirect(url_for("home"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"], endpoint="login")
def login():
    flash = pop_flash()
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = (request.form.get("password") or "").strip()
        nxt = (request.form.get("next") or "").strip()

        conn = get_db()
        row = conn.execute("SELECT * FROM users WHERE username=?", (u,)).fetchone()
        conn.close()

        if not row or not row["is_active"]:
            set_flash("Usuario inválido o inactivo.", "err")
            return redirect(url_for("login"))

        if not check_password_hash(row["pass_hash"], p):
            set_flash("Clave incorrecta.", "err")
            return redirect(url_for("login"))

        session["username"] = row["username"]
        session["role"] = (row["role"] or "").lower()
        return redirect(nxt or url_for("home"))

    nxt = request.args.get("next") or ""
    return render_template_string(LOGIN_HTML, flash=flash, next=nxt)

@app.route("/logout", endpoint="logout")
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))

# --------------------------- | Routes: home | ---------------------------

@app.route("/home", endpoint="home")
@login_required
def home():
    flash = pop_flash()
    role = current_role() or ROLE_SALES
    return render_template_string(
        HOME_HTML,
        username=current_user(),
        role_label=ROLE_LABELS.get(role, role),
        is_admin=is_admin(),
        flash=flash,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

# --------------------------- | Routes: Firmantes Precalificados (all) | ---------------------------

@app.route("/firmantes_all", endpoint="firmantes_all")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def firmantes_all():
    flash = pop_flash()
    q = (request.args.get("q") or "").strip()
    qd = cuit_digits(q)

    conn = get_db()
    where = []
    params = []

    if q:
        like = f"%{q}%"
        where.append("(razon_social LIKE ? COLLATE NOCASE OR cuit LIKE ? COLLATE NOCASE OR cuit_digits LIKE ?)")
        params.extend([like, like, f"%{qd}%" if qd else like])

    sql = f"SELECT * FROM {TABLE_FIRMANTES}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY razon_social COLLATE NOCASE"

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    return render_template_string(
        ALL_FIRMANTES_HTML,
        rows=rows,
        q=q,
        fmt_ars=fmt_ars,
        line_status_label=line_status_label,
        flash=flash,
        is_admin=is_admin(),
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

@app.route("/firmantes_all/<string:cd>/update", methods=["POST"], endpoint="firmantes_all_update")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def firmantes_all_update(cd: str):
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    rs = (f["razon_social"] or "").strip()
    cuit = (f["cuit"] or "").strip()

    lim3 = parse_es_number(request.form.get("lim3"))
    limp = parse_es_number(request.form.get("limp"))
    limf = parse_es_number(request.form.get("limf"))
    exp3 = request.form.get("exp3") or ""
    expp = request.form.get("expp") or ""
    observacion = (request.form.get("observacion") or "").strip()

    ok, msg = upsert_firmante(
        rs, cuit,
        lim3, limp, limf,
        exp3, expp,
        current_user(),
        allow_update=True,
        observacion=observacion
    )
    set_flash(msg, "ok" if ok else "err")
    return redirect(url_for("firmantes_all"))

# --------------------------- | Routes: Firmantes Precalificados (fusion) | ---------------------------

@app.route("/fp", endpoint="fp")
@login_required
def fp():
    flash = pop_flash()

    q = (request.args.get("q") or "").strip()
    qd = cuit_digits(q)
    show_inactive = (request.args.get("show_inactive") or "0").strip() == "1"
    view_scope = (request.args.get("view_scope") or "3ros").strip().lower()
    if view_scope not in ("3ros", "propio", "fce"):
        view_scope = "3ros"

    group_id = (request.args.get("group_id") or "").strip()
    group_id_int = int(group_id) if group_id.isdigit() else None
    first_line = (request.args.get("first_line") or "all").strip().lower()
    if first_line not in ("all", "yes", "no"):
        first_line = "all"

    _, _, agg, cartera_status = load_cartera()
    _, _, facturas_agg, facturas_status = load_facturas()
    blocks_agg, blocks_by_firmante = blocks_agg_today_and_list()

# grupos (para etiquetas)
    conn_g = get_db()
    g_rows = conn_g.execute(f"SELECT id, nombre FROM {TABLE_GROUPS} WHERE is_active=1").fetchall()
    conn_g.close()
    groups_map = {int(r["id"]): (r["nombre"] or "").strip() for r in g_rows}
    group_avail_map = compute_group_available_map(agg=agg, facturas_agg=facturas_agg, blocks_agg=blocks_agg)

    conn = get_db()
    where = []
    params = []

    if not show_inactive:
        where.append("is_active=1")

    if q:
        like = f"%{q}%"
        where.append("(razon_social LIKE ? COLLATE NOCASE OR cuit LIKE ? COLLATE NOCASE OR cuit_digits LIKE ?)")
        params.extend([like, like, f"%{qd}%" if qd else like])

    if group_id_int is not None:
        where.append("grupo_id=?")
        params.append(group_id_int)

    if first_line == "yes":
        where.append("COALESCE(primera_linea, 0)=1")
    elif first_line == "no":
        where.append("COALESCE(primera_linea, 0)=0")

    sql = f"SELECT * FROM {TABLE_FIRMANTES}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY razon_social COLLATE NOCASE"
    rs = conn.execute(sql, params).fetchall()
    conn.close()

    out = []
    role = current_role() or ROLE_SALES
    today = today_local_date()

    for r in rs:
        cd = (r["cuit_digits"] or "").strip() or cuit_digits(r["cuit"] or "")
        used3 = float((agg or {}).get(cd, {}).get("3ros", 0.0))
        usedp = float((agg or {}).get(cd, {}).get("propio", 0.0))
        usedf = float((facturas_agg or {}).get(cd, 0.0))

        lim3 = float(r["credit_limit_3ros"] or 0.0)
        limp = float(r["credit_limit_propio"] or 0.0)
        limf = float(r["credit_limit_fce"] or 0.0)
        exp3 = (r["limit_expiry_3ros"] or "").strip()
        expp = (r["limit_expiry_propio"] or "").strip()
        line3_active = effective_line_active(r["line_active_3ros"], exp3, today=today)
        linep_active = effective_line_active(r["line_active_propio"], expp, today=today)

        blocked3 = float((blocks_agg or {}).get(cd, {}).get("3ros", 0.0))
        blockedp = float((blocks_agg or {}).get(cd, {}).get("propio", 0.0))
        blockedf = float((blocks_agg or {}).get(cd, {}).get("fce", 0.0))

        avail3 = (lim3 - used3 - blocked3) if line3_active else 0.0
        availp = (limp - usedp - blockedp) if linep_active else 0.0
        availf = limf - usedf - blockedf

        gid_raw = r.get("grupo_id") if isinstance(r, dict) else r["grupo_id"]
        try:
            gid_int = int(gid_raw) if gid_raw is not None and str(gid_raw).strip() != "" else None
        except Exception:
            gid_int = None

        group_avail = None
        if gid_int is not None:
            group_info = group_avail_map.get(gid_int) or {}
            if "avail" in group_info:
                group_avail = float(group_info.get("avail", 0.0))
                avail3 = min(avail3, group_avail)
                availp = min(availp, group_avail)
                availf = min(availf, group_avail)

        if view_scope == "3ros":
            if lim3 <= 0:
                continue
            if not show_inactive and not line3_active:
                continue
        if view_scope == "propio":
            if limp <= 0:
                continue
            if not show_inactive and not linep_active:
                continue
        if view_scope == "fce" and limf <= 0:
            continue

# sort metric
        metric = avail3 if view_scope == "3ros" else (availp if view_scope == "propio" else availf)

        d = dict(r)
        d["_group_name"] = groups_map.get(gid_int) if gid_int is not None else ""
        d["_group_id"] = gid_int
        d["_used3"] = used3
        d["_usedp"] = usedp
        d["_blocked3"] = blocked3
        d["_blockedp"] = blockedp
        d["_avail3"] = avail3
        d["_availp"] = availp
        d["_lim3"] = lim3
        d["_limp"] = limp
        d["_limf"] = limf
        d["_line3_active"] = line3_active
        d["_linep_active"] = linep_active
        d["_line3_status"] = line_status_label(r["line_active_3ros"], exp3, today=today)
        d["_linep_status"] = line_status_label(r["line_active_propio"], expp, today=today)
        d["_usedf"] = usedf
        d["_blockedf"] = blockedf
        d["_availf"] = availf
        d["_group_avail"] = group_avail
        d["_lim3_raw"] = (r["credit_limit_3ros"] if r["credit_limit_3ros"] is not None else 0)
        d["_limp_raw"] = (r["credit_limit_propio"] if r["credit_limit_propio"] is not None else 0)
        d["_limf_raw"] = (r["credit_limit_fce"] if r["credit_limit_fce"] is not None else 0)
        d["_blocked_any"] = (blocked3 > 0 or blockedp > 0 or blockedf > 0)
        d["_avail_any_neg"] = (metric < 0)  # negativo según el criterio de orden
        d["_metric"] = metric

# bloqueos list para el detalle
        bl = []
        for b in (blocks_by_firmante.get(cd, []) if blocks_by_firmante else []):
            can_delete = True
            if role == ROLE_SALES and b["username"] != current_user():
                can_delete = False
            bl.append({
                "id": b["id"],
                "username": b["username"],
                "scope": b["scope"],
                "amount": float(b["amount"] or 0.0),
                "can_delete": can_delete,
            })
        d["_blocks_list"] = bl

        out.append(d)

# Orden: negativos arriba, luego menor a mayor
    out.sort(key=lambda x: (x["_metric"] >= 0, x["_metric"]))

    can_edit = role in (ROLE_ADMIN, ROLE_RISK)

    return render_template_string(
        FP_HTML,
        username=current_user(),
        role_label=ROLE_LABELS.get(role, role),
        flash=flash,
        q=q,
        group_id=(str(group_id_int) if group_id_int is not None else ""),
        show_inactive=show_inactive,
        view_scope=view_scope,
        first_line=first_line,
        firmantes=out,
        can_edit=can_edit,
        fmt_ars=fmt_ars,
        cartera_status=(facturas_status if view_scope == "fce" else cartera_status),
        live_search_js=LIVE_SEARCH_JS,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

@app.route("/fp/export_xlsx", methods=["POST"], endpoint="fp_export_xlsx")
@login_required
def fp_export_xlsx():
    if openpyxl is None:
        return ("openpyxl no está disponible en el servidor.", 500)
    try:
        payload = request.get_json(silent=True) or {}
        headers = payload.get("headers") or []
        rows = payload.get("rows") or []
        meta = payload.get("meta") or {}
        if not isinstance(headers, list) or not headers:
            return ("No hay columnas visibles para exportar.", 400)
        if not isinstance(rows, list):
            return ("Payload inválido para exportación.", 400)
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "FP"
        ws.sheet_view.showGridLines = False
        exported_at = meta.get("exported_at") or now_str()
        try:
            dt = datetime.fromisoformat(str(exported_at).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=ARG_TZ)
            exported_local = dt.astimezone(ARG_TZ).strftime("%d/%m/%Y %H:%M:%S")
        except Exception:
            exported_local = now_str()
        filters_txt = []
        q = (meta.get("q") or "").strip()
        if q:
            filters_txt.append(f"Búsqueda: {q}")
        filters_txt.append(f"Vista: {(meta.get('view_scope') or '3ros').upper()}")
        fl = (meta.get("first_line") or "all").strip().lower()
        fl_map = {"all": "todos", "yes": "sí", "no": "no"}
        filters_txt.append(f"Primera línea: {fl_map.get(fl, fl)}")
        filters_txt.append("Incluye inactivos" if str(meta.get("show_inactive") or "0") == "1" else "Solo activos")
        ws["A1"] = "Firmantes Precalificados"
        ws["A2"] = f"Exportado: {exported_local}"
        ws["A3"] = " | ".join(filters_txt)
        ws["A1"].font = Font(size=14, bold=True)
        ws["A2"].font = Font(italic=True, color="666666")
        ws["A3"].font = Font(color="666666")
        header_row = 5
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=str(header))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        money_keywords = ("límite", "limite", "usado", "bloq", "disp", "dispon", "monto", "saldo", "importe")
        date_keywords = ("vto", "fecha")
        thin_gray = Side(style="thin", color="D9E2F3")
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        def parse_export_number(value):
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return float(value)
            s = str(value).strip()
            if not s:
                return None
            s = s.replace("$", "").replace("ARS", "").replace("ar$", "").replace(" ", "")
            neg = False
            if s.startswith("(") and s.endswith(")"):
                neg = True
                s = s[1:-1]
            if s.startswith("-"):
                neg = True
                s = s[1:]
            if "," in s:
                s = s.replace(".", "").replace(",", ".")
            elif s.count(".") > 1:
                s = s.replace(".", "")
            try:
                num = float(s)
                return -num if neg else num
            except Exception:
                return None
        def parse_export_date(value):
            if value is None:
                return None
            s = str(value).strip()
            if not s or s == "—":
                return None
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    return datetime.strptime(s, fmt)
                except Exception:
                    pass
            return None
        widths = [len(str(h)) for h in headers]
        data_start = header_row + 1
        for row_offset, row in enumerate(rows, start=data_start):
            if not isinstance(row, list):
                continue
            for col_idx, raw_value in enumerate(row[:len(headers)], start=1):
                header = str(headers[col_idx - 1])
                header_norm = normalize_header(header)
                cell = ws.cell(row=row_offset, column=col_idx)
                value = raw_value if raw_value is not None else ""
                parsed_number = parse_export_number(value) if any(k in header_norm for k in money_keywords) else None
                parsed_date = parse_export_date(value) if parsed_number is None and any(k in header_norm for k in date_keywords) else None
                if parsed_number is not None:
                    cell.value = parsed_number
                    cell.number_format = '$ #.##0,00;[Red]-$ #.##0,00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    width_candidate = len(f"{parsed_number:,.2f}") + 4
                elif parsed_date is not None:
                    cell.value = parsed_date
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    width_candidate = 12
                else:
                    cell.value = str(value)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    width_candidate = min(max(len(str(value)) + 2, len(header) + 2), 60)
                cell.border = border
                if row_offset % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F7FBFF")
                widths[col_idx - 1] = max(widths[col_idx - 1], width_candidate)
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 40)
        ws.freeze_panes = f"A{data_start}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(ws.max_row, header_row)}"
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="firmantes_precalificados.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return (f"Error generando XLSX: {e}", 500)

@app.route("/fp/create", methods=["POST"], endpoint="fp_create")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_create():
    rs = (request.form.get("razon_social") or "").strip()
    cuit = (request.form.get("cuit") or "").strip()
    lim3 = parse_es_number(request.form.get("lim3")) or 0.0
    limp = parse_es_number(request.form.get("limp")) or 0.0
    exp3 = request.form.get("exp3") or ""
    expp = request.form.get("expp") or ""
    limf = parse_es_number(request.form.get("limf")) or 0.0
    primera_linea = parse_bool_si(request.form.get("primera_linea"))
    observacion = (request.form.get("observacion") or "").strip()

    ok, msg = upsert_firmante(
        rs, cuit, lim3, limp, limf, exp3, expp,
        current_user(),
        allow_update=False,
        primera_linea=primera_linea,
        observacion=observacion
    )
    set_flash(msg, "ok" if ok else "err")
    return redirect(url_for("fp"))

@app.route("/fp/upload", methods=["POST"], endpoint="fp_upload")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_upload():
    file = request.files.get("file")
    if not file:
        set_flash("Falta archivo CSV.", "err")
        return redirect(url_for("fp"))

    try:
        raw = file.read()
        txt = decode_bytes_best_effort(raw)
        reader = read_csv_dictreader(txt)
        if not reader.fieldnames:
            set_flash("CSV inválido: no se detectaron encabezados.", "err")
            return redirect(url_for("fp"))

        rows_in = []
        unknown = set()

        for i, row in enumerate(reader, start=2):
            row_norm = {normalize_header(k): v for k, v in (row or {}).items()}

            def pick(*names):
                for n in names:
                    key = normalize_header(n)
                    if key in row_norm:
                        return row_norm.get(key)
                return None

            rs = (pick("razon social", "razon_social", "razonsocial", "firmante", "nombre", "razon") or "").strip()
            cuit = (pick("cuit", "cuil") or "").strip()

            lim3_raw = pick("lim3", "lim 3", "limite 3", "limite_3", "limite 3ros", "limite credito 3ros", "credito 3ros", "lim_3ros")
            limp_raw = pick("lim prop", "limprop", "limite prop", "limite propio", "limite credito propio", "credito propio", "lim_propio")
            limf_raw = pick("lim fce", "limf", "limite fce", "limite credito fce", "credito fce", "lim_fce")

            exp3_raw = pick("venc 3", "venc3", "vto 3", "vencimiento 3", "vencimiento 3ros", "vto 3ros", "fecha vencimiento 3ros")
            expp_raw = pick("venc prop", "vencprop", "vto prop", "vencimiento prop", "vencimiento propio", "vto propio", "fecha vencimiento propio")

            group_name = (pick("grupo", "grupo economico", "grupo_economico", "grupoeconomico") or "").strip()
            primera_linea = parse_bool_si(pick("primera linea", "primera_linea", "primera linea?", "es primera linea"))

            cd = cuit_digits(cuit)

            if not rs:
                rows_in.append({"_row": i, "_err": "falta Razón Social"})
                continue
            if not cd:
                rows_in.append({"_row": i, "_err": "CUIT inválido/vacío"})
                continue

            def parse_partial_number(v):
                if v is None:
                    return None
                s = str(v).strip()
                if s == "":
                    return None
                n = parse_es_number(s)
# si no parsea, lo dejamos como None => no pisa
                return None if n is None else float(n)

            v3 = parse_partial_number(lim3_raw)
            vp = parse_partial_number(limp_raw)
            vf = parse_partial_number(limf_raw)

            exp3 = None if exp3_raw is None or str(exp3_raw).strip() == "" else str(exp3_raw).strip()
            expp = None if expp_raw is None or str(expp_raw).strip() == "" else str(expp_raw).strip()

            gid = _KEEP  # por defecto: no actualiza si grupo vacío
            if group_name:
                g = get_group_by_name(group_name)
                if g:
                    gid = int(g["id"])
                else:
                    unknown.add(group_name)
                    gid = {"__unknown__": group_name}

            rows_in.append({
                "_row": i,
                "_err": None,
                "razon_social": rs,
                "cuit": cuit,
                "lim3": v3,
                "limp": vp,
                "limf": vf,
                "exp3": exp3,
                "expp": expp,
                "group": gid,
                "limf": vf,
                "primera_linea": primera_linea,
            })

        unknown_list = sorted(list(unknown), key=lambda x: x.lower())

# si hay grupos desconocidos, staging y confirmación
        if unknown_list:
            batch_id = uuid.uuid4().hex
            payload = {"rows": rows_in, "unknown_groups": unknown_list}
            conn = get_db()
            conn.execute(
                f"INSERT INTO {TABLE_IMPORT_BATCHES}(batch_id, username, payload_json, created_at) VALUES (?, ?, ?, ?)",
                (batch_id, current_user(), json.dumps(payload, ensure_ascii=False), now_str())
            )
            conn.commit()
            conn.close()
            return redirect(url_for("fp_upload_confirm", batch_id=batch_id))

# aplicar directo
        ok_n = 0
        fail_n = 0
        reasons = []

        for r in rows_in:
            if r.get("_err"):
                fail_n += 1
                reasons.append(f"Fila {r.get('_row')}: {r.get('_err')}")
                continue

            ok, msg = upsert_firmante(
                r["razon_social"], r["cuit"],
                r["lim3"], r["limp"], r["limf"],
                r["exp3"], r["expp"],
                current_user(),
                allow_update=True,
                reactivate=True,
                group_id=r["group"] if r["group"] is not None else _KEEP,
                primera_linea=bool(r.get("primera_linea")),
            )
            if ok:
                ok_n += 1
            else:
                fail_n += 1
                reasons.append(f"Fila {r.get('_row')}: {msg}")

        extra = ""
        if reasons:
            extra = " | Ejemplos: " + " ; ".join(reasons[:10])

        set_flash(f"Importación finalizada. OK: {ok_n} | Fallidos: {fail_n}{extra}", "ok" if ok_n > 0 else "err")
        return redirect(url_for("fp"))

    except Exception as e:
        set_flash(f"Error importando CSV: {e}", "err")
        return redirect(url_for("fp"))

@app.route("/fp/upload/confirm", methods=["GET"], endpoint="fp_upload_confirm")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_upload_confirm():
    batch_id = (request.args.get("batch_id") or "").strip()
    if not batch_id:
        return redirect(url_for("fp"))
    conn = get_db()
    row = conn.execute(f"SELECT payload_json FROM {TABLE_IMPORT_BATCHES} WHERE batch_id=? AND username=?", (batch_id, current_user())).fetchone()
    conn.close()
    if not row:
        flash_err("Batch no encontrado o expirado.")
        return redirect(url_for("fp"))
    payload = json.loads(row["payload_json"])
    unknown_groups = payload.get("unknown_groups") or []
    return render_template_string(
        FP_UPLOAD_CONFIRM_HTML,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
        batch_id=batch_id,
        unknown_groups=unknown_groups,
        flash=pull_flash(),
    )

@app.route("/fp/upload/confirm/apply", methods=["POST"], endpoint="fp_upload_confirm_apply")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_upload_confirm_apply():
    batch_id = (request.form.get("batch_id") or "").strip()
    if not batch_id:
        return redirect(url_for("fp"))

    conn = get_db()
    row = conn.execute(f"SELECT payload_json FROM {TABLE_IMPORT_BATCHES} WHERE batch_id=? AND username=?", (batch_id, current_user())).fetchone()
    if not row:
        conn.close()
        flash_err("Batch no encontrado o expirado.")
        return redirect(url_for("fp"))
    payload = json.loads(row["payload_json"])

    unknown_groups = payload.get("unknown_groups") or []
    rows_in = payload.get("rows") or []

# decisiones del usuario
    decisions = {}  # name -> {"create":bool, "limit":float}
    for idx, name in enumerate(unknown_groups):
        act = (request.form.get(f"action__{idx}") or "yes").strip().lower()
        if act == "no":
            decisions[name] = {"create": False, "limit": 0.0}
            continue
        lim_raw = request.form.get(f"limit__{idx}")
        lim = parse_es_number(lim_raw)
        if lim is None:
            conn.close()
            flash_err(f"Límite inválido para grupo: {name}")
            return redirect(url_for("fp_upload_confirm", batch_id=batch_id))
        decisions[name] = {"create": True, "limit": float(lim)}

# crear grupos seleccionados y mapear a ids
    name_to_id = {}
    for name, d in decisions.items():
        if not d["create"]:
            continue
        ok, msg, gid = upsert_group(name, d["limit"], actor=current_user(), allow_update=False)
        if not ok:
            conn.close()
            flash_err(f"No se pudo crear grupo '{name}': {msg}")
            return redirect(url_for("fp_upload_confirm", batch_id=batch_id))
        name_to_id[name] = int(gid)

# aplicar firmantes
    ok_n = 0
    fail_n = 0
    reasons = []

    for r in rows_in:
        if r.get("_err"):
            fail_n += 1
            reasons.append(f"Fila {r.get('_row')}: {r.get('_err')}")
            continue

        gid = r.get("group", _KEEP)
# resolver unknown
        if isinstance(gid, dict) and gid.get("__unknown__"):
            name = gid.get("__unknown__")
            if decisions.get(name, {}).get("create"):
                gid = name_to_id.get(name, _KEEP)
            else:
                gid = None  # si elige NO => campo vacío

        ok, msg = upsert_firmante(
            r["razon_social"], r["cuit"],
            r.get("lim3"), r.get("limp"), r.get("limf"),
            r.get("exp3"), r.get("expp"),
            current_user(),
            allow_update=True,
            reactivate=True,
            group_id=gid,
            primera_linea=bool(r.get("primera_linea")),
        )
        if ok:
            ok_n += 1
        else:
            fail_n += 1
            reasons.append(f"Fila {r.get('_row')}: {msg}")

# borrar batch
    conn.execute(f"DELETE FROM {TABLE_IMPORT_BATCHES} WHERE batch_id=? AND username=?", (batch_id, current_user()))
    conn.commit()
    conn.close()

    extra = ""
    if reasons:
        extra = " | Ejemplos: " + " ; ".join(reasons[:10])
    set_flash(f"Importación finalizada. OK: {ok_n} | Fallidos: {fail_n}{extra}", "ok" if ok_n > 0 else "err")
    return redirect(url_for("fp"))

@app.route("/fp/<string:cd>/update_limits", methods=["POST"], endpoint="fp_update_limits")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_update_limits(cd: str):
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    rs = (f["razon_social"] or "").strip()
    cuit = (f["cuit"] or "").strip()

    lim3 = parse_es_number(request.form.get("lim3"))
    limp = parse_es_number(request.form.get("limp"))
    limf = parse_es_number(request.form.get("limf"))
    exp3 = request.form.get("exp3") or ""
    expp = request.form.get("expp") or ""
    observacion = (request.form.get("observacion") or "").strip()

    ok, msg = upsert_firmante(
        rs, cuit,
        lim3, limp, limf,
        exp3, expp,
        current_user(),
        allow_update=True,
        observacion=observacion
    )
    set_flash(msg, "ok" if ok else "err")
    return redirect(url_for("fp"))

@app.route("/fp/<string:cd>/set_first_line", methods=["POST"], endpoint="fp_set_first_line")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_set_first_line(cd: str):
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    primera_linea = request.form.get("primera_linea") == "1"
    ok, msg = upsert_firmante(
        f["razon_social"], f["cuit"],
        lim3=None, limp=None, limf=None,
        exp3=None, expp=None,
        actor=current_user(),
        allow_update=True, reactivate=False,
        group_id=_KEEP,
        primera_linea=primera_linea,
    )
    flash_ok("Clasificación actualizada.") if ok else flash_err(msg)
    return redirect(url_for("fp", q=request.args.get("q", ""), show_inactive=request.args.get("show_inactive", "0"), group_id=request.args.get("group_id", ""), first_line=request.args.get("first_line", "all"), view_scope=request.args.get("view_scope", "3ros")))

@app.route("/fp/<string:cd>/set_group", methods=["POST"], endpoint="fp_set_group")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_set_group(cd: str):
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    if request.form.get("clear_group") == "1":
# limpiar
        ok, msg = upsert_firmante(
            f["razon_social"], f["cuit"],
            lim3=None, limp=None, limf=None,
            exp3=None, expp=None,
            actor=current_user(),
            allow_update=True, reactivate=False,
            group_id=None
        )
        flash_ok("Grupo económico quitado.") if ok else flash_err(msg)
        return redirect(url_for("fp", q=request.args.get("q", ""), show_inactive=request.args.get("show_inactive", "0"), group_id=request.args.get("group_id", "")))

    group_name = (request.form.get("group_name") or "").strip()
    if not group_name:
# no pisa por vacío en UI? acá interpretamos vacío como borrar (más natural)
        ok, msg = upsert_firmante(
            f["razon_social"], f["cuit"],
            lim3=None, limp=None, limf=None, exp3=None, expp=None,
            actor=current_user(),
            allow_update=True, reactivate=False,
            group_id=None
        )
        flash_ok("Grupo económico quitado.") if ok else flash_err(msg)
        return redirect(url_for("fp"))

    g = get_group_by_name(group_name)
    if not g:
# confirmar creación
        return redirect(url_for("group_confirm", cd=cd, group_name=group_name))

    ok, msg = upsert_firmante(
        f["razon_social"], f["cuit"],
        lim3=None, limp=None, limf=None, exp3=None, expp=None,
        actor=current_user(),
        allow_update=True, reactivate=False,
        group_id=int(g["id"])
    )
    flash_ok("Grupo económico actualizado.") if ok else flash_err(msg)
    return redirect(url_for("fp"))

@app.route("/groups/confirm", methods=["GET"], endpoint="group_confirm")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def group_confirm():
    cd = cuit_digits(request.args.get("cd") or "")
    group_name = (request.args.get("group_name") or "").strip()
    if not cd or not group_name:
        return redirect(url_for("fp"))
    return render_template_string(
        GROUP_CONFIRM_HTML,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
        cd=cd,
        group_name=group_name,
    )

@app.route("/groups/confirm/no", methods=["POST"], endpoint="group_confirm_no")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def group_confirm_no():
    cd = cuit_digits(request.form.get("cd") or "")
    if not cd:
        return redirect(url_for("fp"))
    f = get_firmante(cd)
    if not f:
        return redirect(url_for("fp"))

    ok, msg = upsert_firmante(
        f["razon_social"], f["cuit"],
        lim3=None, limp=None, limf=None,
        exp3=None, expp=None,
        actor=current_user(),
        allow_update=True, reactivate=False,
        group_id=None
    )
    flash_ok("Grupo económico no asignado (campo vaciado).") if ok else flash_err(msg)
    return redirect(url_for("fp"))

@app.route("/groups/confirm/yes", methods=["POST"], endpoint="group_confirm_yes")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def group_confirm_yes():
    cd = cuit_digits(request.form.get("cd") or "")
    group_name = (request.form.get("group_name") or "").strip()
    lim = parse_es_number(request.form.get("limit"))
    if not cd or not group_name:
        return redirect(url_for("fp"))
    if lim is None:
        flash_err("Límite grupal inválido.")
        return redirect(url_for("group_confirm", cd=cd, group_name=group_name))

    ok, msg, gid = upsert_group(group_name, lim, actor=current_user(), allow_update=False)
    if not ok:
        flash_err(msg)
        return redirect(url_for("fp"))

    f = get_firmante(cd)
    if not f:
        return redirect(url_for("fp"))

    ok, msg = upsert_firmante(
        f["razon_social"], f["cuit"],
        lim3=None, limp=None, limf=None,
        exp3=None, expp=None,
        actor=current_user(),
        allow_update=True, reactivate=False,
        group_id=None
    )
    flash_ok("Grupo creado y asignado.") if ok2 else flash_err(msg2)
    return redirect(url_for("fp"))

@app.route("/fp/<string:cd>/deactivate", methods=["POST"], endpoint="fp_deactivate")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_deactivate(cd: str):
    cd = cuit_digits(cd)
    observacion = (request.form.get("observacion") or "").strip()
    ok, msg = set_firmante_active(cd, False, current_user(), observacion=observacion)
    set_flash("Firmante desactivado." if ok else msg, "ok" if ok else "err")
    return redirect(url_for("fp"))

@app.route("/fp/<string:cd>/reactivate", methods=["POST"], endpoint="fp_reactivate")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def fp_reactivate(cd: str):
    cd = cuit_digits(cd)
    observacion = (request.form.get("observacion") or "").strip()
    ok, msg = set_firmante_active(cd, True, current_user(), observacion=observacion)
    set_flash("Firmante reactivado." if ok else msg, "ok" if ok else "err")
    return redirect(url_for("fp"))


@app.route("/fp/<string:cd>/delete", methods=["POST"], endpoint="fp_delete")
@roles_required(ROLE_ADMIN)
def fp_delete(cd: str):
    cd = cuit_digits(cd)
    observacion = (request.form.get("observacion") or "").strip()
    ok, msg = delete_firmante_hard(cd, current_user(), observacion=observacion)
    set_flash(msg, "ok" if ok else "err")
    return redirect(request.referrer or url_for("fp"))

# --------------------------- | Bloqueos (desde FP y también desde cartera) | ---------------------------

@app.route("/firmantes/<string:cd>/blocks/add", methods=["POST"], endpoint="block_add")
@login_required
def block_add(cd: str):
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    scope = (request.form.get("scope") or "").strip().lower()
    if scope not in ("propio", "3ros", "fce"):
        set_flash("Elegí si el bloqueo es Propio, 3ros o FCE.", "err")
        return redirect(url_for("fp"))

    amt = parse_es_number(request.form.get("amount"))
    if amt is None or amt <= 0:
        set_flash("Monto de bloqueo inválido (debe ser > 0).", "err")
        return redirect(url_for("fp"))

    today = today_local_date()
    conn = get_db()
    conn.execute("""
        INSERT INTO limit_blocks (firmante_cuit_digits, scope, username, amount, block_date, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (cd, scope, current_user(), float(amt), today, now_str()))
    conn.commit()
    conn.close()

    set_flash("Bloqueo aplicado.", "ok")
    return redirect(url_for("fp"))

@app.route("/blocks/<int:block_id>/delete", methods=["POST"], endpoint="block_delete")
@login_required
def block_delete(block_id: int):
    conn = get_db()
    row = conn.execute("""
        SELECT id, firmante_cuit_digits, scope, username, amount
        FROM limit_blocks
        WHERE id=?
    """, (block_id,)).fetchone()

    if not row:
        conn.close()
        abort(404)

    role = current_role()
    if role == ROLE_SALES and row["username"] != current_user():
        conn.close()
        abort(403)

    conn.execute("DELETE FROM limit_blocks WHERE id=?", (block_id,))
    conn.commit()
    conn.close()

    set_flash("Bloqueo eliminado.", "ok")
    return redirect(url_for("fp"))

# --------------------------- | Cartera por firmante (detalle) | ---------------------------

@app.route("/firmantes/<string:cd>/cartera", endpoint="firmante_cartera")
@login_required
def firmante_cartera(cd: str):
    flash = pop_flash()
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    headers, all_rows, agg, cartera_status = load_cartera()
    fact_headers, fact_all_rows, fact_agg, facturas_status = load_facturas()
    blocks_agg, _ = blocks_agg_today_and_list()

    used3 = float((agg or {}).get(cd, {}).get("3ros", 0.0))
    usedp = float((agg or {}).get(cd, {}).get("propio", 0.0))
    usedf = float((fact_agg or {}).get(cd, 0.0))

    lim3 = float(f["credit_limit_3ros"] or 0.0)
    limp = float(f["credit_limit_propio"] or 0.0)
    limf = float(f["credit_limit_fce"] or 0.0)

    exp3 = (f["limit_expiry_3ros"] or "").strip()
    expp = (f["limit_expiry_propio"] or "").strip()
    line3_active = effective_line_active(f["line_active_3ros"], exp3)
    linep_active = effective_line_active(f["line_active_propio"], expp)

    blocked3 = float((blocks_agg or {}).get(cd, {}).get("3ros", 0.0))
    blockedp = float((blocks_agg or {}).get(cd, {}).get("propio", 0.0))
    blockedf = float((blocks_agg or {}).get(cd, {}).get("fce", 0.0))

    avail3 = (lim3 - used3 - blocked3) if line3_active else 0.0
    availp = (limp - usedp - blockedp) if linep_active else 0.0
    availf = limf - usedf - blockedf

    try:
        gid = int(f["grupo_id"]) if f["grupo_id"] is not None and str(f["grupo_id"]).strip() != "" else None
    except Exception:
        gid = None
    if gid is not None:
        group_avail_map = compute_group_available_map(agg=agg, facturas_agg=fact_agg, blocks_agg=blocks_agg)
        group_info = group_avail_map.get(gid) or {}
        if "avail" in group_info:
            group_avail = float(group_info.get("avail", 0.0))
            avail3 = min(avail3, group_avail)
            availp = min(availp, group_avail)
            availf = min(availf, group_avail)

# Cartera cheques filtrada por CUIT
    rows = []
    if headers and all_rows:
        headers_norm = [normalize_header(h) for h in headers]
        idx_cuit = None
        for i, h in enumerate(headers_norm):
            if h == "cuit":
                idx_cuit = i
                break
        if idx_cuit is not None:
            for r in all_rows:
                c1 = cuit_digits(r[idx_cuit])
                if c1 == cd:
                    rows.append(r)

# Cartera facturas filtrada por Firmante
    fact_rows = []
    if fact_headers and fact_all_rows:
        fact_headers_norm = [normalize_header(h) for h in fact_headers]
        idx_firmante = None
        for i, h in enumerate(fact_headers_norm):
            if h == "firmante":
                idx_firmante = i
                break
        if idx_firmante is not None:
            for r in fact_all_rows:
                c1 = cuit_digits(r[idx_firmante])
                if c1 == cd:
                    fact_rows.append(r)

    return render_template_string(
        FIRMANTE_CARTERA_HTML,
        f=f,
        headers=headers,
        rows=rows,
        fact_headers=fact_headers,
        fact_rows=fact_rows,
        lim3=lim3, limp=limp, limf=limf,
        exp3=exp3, expp=expp,
        used3=used3, usedp=usedp, usedf=usedf,
        blocked3=blocked3, blockedp=blockedp, blockedf=blockedf,
        avail3=avail3, availp=availp, availf=availf,
        line3_active=line3_active, linep_active=linep_active,
        line3_status=line_status_label(f["line_active_3ros"], exp3),
        linep_status=line_status_label(f["line_active_propio"], expp),
        cartera_status=cartera_status,
        facturas_status=facturas_status,
        fmt_ars=fmt_ars,
        flash=flash,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

@app.route("/firmantes/<string:cd>/cartera/update_limits", methods=["POST"], endpoint="firmante_update_limits_from_cartera")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def firmante_update_limits_from_cartera(cd: str):
    cd = cuit_digits(cd)
    f = get_firmante(cd)
    if not f:
        abort(404)

    rs = (f["razon_social"] or "").strip()
    cuit = (f["cuit"] or "").strip()

    lim3 = parse_es_number(request.form.get("lim3"))
    limp = parse_es_number(request.form.get("limp"))
    limf = parse_es_number(request.form.get("limf"))

    exp3 = request.form.get("exp3") or ""
    expp = request.form.get("expp") or ""
    observacion = (request.form.get("observacion") or "").strip()

    ok, msg = upsert_firmante(
        rs, cuit,
        lim3, limp, limf,
        exp3, expp,
        current_user(),
        allow_update=True,
        observacion=observacion
    )
    set_flash(msg, "ok" if ok else "err")
    return redirect(url_for("firmante_cartera", cd=cd))

# --------------------------- | Auditoría ABM | ---------------------------

@app.route("/audit", endpoint="audit")
@login_required
def audit():
    flash = pop_flash()
    q = (request.args.get("q") or "").strip()
    qd = cuit_digits(q)

    conn = get_db()

    base_sql = f"""
        SELECT
            a.id,
            a.firmante_cuit_digits,
            a.action,
            a.username,
            a.details,
            a.created_at,
            f.razon_social AS razon_social,
            f.cuit AS cuit
        FROM audit_log a
        LEFT JOIN {TABLE_FIRMANTES} f
          ON f.cuit_digits = a.firmante_cuit_digits
    """

    params = []
    if q:
        base_sql += """
        WHERE
            a.firmante_cuit_digits LIKE ?
            OR f.razon_social LIKE ? COLLATE NOCASE
            OR f.cuit LIKE ?
        """
        params = [
            f"%{qd}%" if qd else f"%{q}%",
            f"%{q}%",
            f"%{q}%"
        ]

    base_sql += " ORDER BY a.id DESC LIMIT 1500"

    rows = conn.execute(base_sql, params).fetchall()
    conn.close()

    out = [dict(r) for r in rows]

    return render_template_string(
        AUDIT_HTML,
        rows=out,
        q=q,
        flash=flash,
        is_admin=is_admin(),
        live_search_js=LIVE_SEARCH_JS,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

@app.route("/audit/<int:audit_id>/delete", methods=["POST"], endpoint="audit_delete")
@roles_required(ROLE_ADMIN)
def audit_delete(audit_id: int):
    conn = get_db()
    conn.execute("DELETE FROM audit_log WHERE id=?", (audit_id,))
    conn.commit()
    conn.close()

    set_flash("Registro ABM eliminado.", "ok")
    return redirect(request.referrer or url_for("audit"))


@app.route("/audit/delete_all", methods=["POST"], endpoint="audit_delete_all")
@roles_required(ROLE_ADMIN)
def audit_delete_all():
    conn = get_db()
    conn.execute("DELETE FROM audit_log")
    conn.commit()
    conn.close()

    set_flash("Se borraron todos los registros ABM.", "ok")
    return redirect(url_for("audit"))

# --------------------------- | Cartera (general) | ---------------------------

@app.route("/cartera", endpoint="cartera")
@login_required
def cartera():
    flash = pop_flash()
    view_kind = (request.args.get("kind") or "cheques").strip().lower()
    if view_kind not in ("cheques", "facturas"):
        view_kind = "cheques"

    headers, rows, _, status = load_cartera()
    fact_headers_raw, fact_rows_raw, _, fact_status = load_facturas()
    fact_headers, fact_rows = reorder_facturas_for_ui(fact_headers_raw, fact_rows_raw)
    status_out = status if view_kind == "cheques" else fact_status

    return render_template_string(
        CARTERA_HTML,
        headers=headers,
        rows=rows,
        fact_headers=fact_headers,
        fact_rows=fact_rows,
        view_kind=view_kind,
        status=status_out,
        flash=flash,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

# --------------------------- | Users (Admin) | ---------------------------

# --------------------------- | Routes: Grupos Económicos | ---------------------------

@app.route("/groups", methods=["GET"], endpoint="groups")
@login_required
def groups():
    apply_expired_limits_if_needed()

    headers, rows, agg, status = load_cartera()
    _, _, facturas_agg, fact_status = load_facturas()
    blocks_agg, _ = blocks_agg_today_and_list()

    conn = get_db()
    groups_rows = conn.execute(f"SELECT id, nombre, credit_limit_grupal FROM {TABLE_GROUPS} WHERE is_active=1").fetchall()
    conn.close()

    group_avail_map = compute_group_available_map(agg=agg, facturas_agg=facturas_agg, blocks_agg=blocks_agg)

    computed = []
    for g in groups_rows:
        gid = int(g["id"])
        ginfo = group_avail_map.get(gid) or {}
        d = {
            "id": gid,
            "nombre": (g["nombre"] or "").strip(),
            "_lim": float(ginfo.get("limit", g["credit_limit_grupal"] or 0.0)),
            "_used": float(ginfo.get("used", 0.0)),
            "_blocked_total": float(ginfo.get("blocked", 0.0)),
            "_avail": float(ginfo.get("avail", float(g["credit_limit_grupal"] or 0.0))),
        }
        computed.append(d)

# Orden: negativos arriba, luego disponible asc
    computed.sort(key=lambda x: (0 if x["_avail"] < 0 else 1, x["_avail"]))

    return render_template_string(
        GROUPS_HTML,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
        username=current_user(),
        role_label=ROLE_LABELS.get(current_role() or ROLE_SALES, current_role() or ROLE_SALES),
        can_edit=is_admin() or is_risk(),
        rows=computed,
        fmt_ars=fmt_ars,
        flash=pull_flash(),
    )

@app.route("/groups/create", methods=["POST"], endpoint="groups_create")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def groups_create():
    nombre = (request.form.get("nombre") or "").strip()
    lim = parse_es_number(request.form.get("limite"))
# vacío => None, no pisa si existiera; pero acá es alta manual, lo tratamos como 0 si es nuevo
    ok, msg, _ = upsert_group(nombre, lim, actor=current_user(), allow_update=False)
    flash_ok(msg) if ok else flash_err(msg)
    return redirect(url_for("groups"))

@app.route("/groups/<int:group_id>/update_limit", methods=["POST"], endpoint="groups_update_limit")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def groups_update_limit(group_id: int):
    g = get_group(group_id)
    if not g:
        abort(404)
    lim = parse_es_number(request.form.get("limite"))
    if lim is None:
        flash_err("Límite inválido.")
        return redirect(url_for("groups"))
    ok, msg, _ = upsert_group(g["nombre"], lim, actor=current_user(), allow_update=True)
    flash_ok("Límite grupal actualizado.") if ok else flash_err(msg)
    return redirect(url_for("groups"))

@app.route("/groups/upload", methods=["POST"], endpoint="groups_upload")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def groups_upload():
    f = request.files.get("file")
    if not f:
        flash_err("No se seleccionó archivo.")
        return redirect(url_for("groups"))

    try:
        raw = f.read()
        try:
            txt = raw.decode("utf-8")
        except Exception:
            txt = raw.decode("latin-1")

        reader = csv.DictReader(io.StringIO(txt))
        n_ok = 0
        n_err = 0
        for row in reader:
            nombre = (row.get("grupo") or row.get("nombre") or "").strip()
            lim_raw = row.get("limite")
            lim = None if lim_raw is None or str(lim_raw).strip() == "" else parse_es_number(lim_raw)
            if not nombre:
                n_err += 1
                continue
# si existe y lim vacío => no pisa
            ok, msg, _ = upsert_group(nombre, lim, actor=current_user(), allow_update=True)
            if ok:
                n_ok += 1
            else:
                n_err += 1

        flash_ok(f"Importación finalizada. OK={n_ok} · Errores={n_err}.")
        return redirect(url_for("groups"))
    except Exception as e:
        flash_err(f"Error importando CSV: {e}")
        return redirect(url_for("groups"))

@app.route("/groups/<int:group_id>/cartera", methods=["GET"], endpoint="group_cartera")
@login_required
def group_cartera(group_id: int):
    apply_expired_limits_if_needed()
    g = get_group(group_id)
    if not g:
        abort(404)

    headers, rows, agg, status = load_cartera()
    if status != "OK":
        pass

    conn = get_db()
    firm_rows = conn.execute(f"SELECT cuit_digits FROM {TABLE_FIRMANTES} WHERE grupo_id=?", (group_id,)).fetchall()
    conn.close()
    cuits = set([r["cuit_digits"] for r in firm_rows if r["cuit_digits"]])

# filtrar filas donde CUIT ∈ grupo (sin importar CUIT2)
    idx_cuit = None
    for i, h in enumerate(headers):
        if (h or "").strip().lower() in ("cuit", "cuit_emisor", "cuitemisor"):
            idx_cuit = i
            break

    filtered = []
    if idx_cuit is not None:
        for r in rows:
            cd = cuit_digits(r[idx_cuit]) if idx_cuit < len(r) else ""
            if cd and cd in cuits:
                filtered.append(r)

    return render_template_string(
        GROUP_CARTERA_HTML,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
        headers=headers,
        rows=filtered,
        status=status,
        group=g,
        firmantes_count=len(cuits),
        flash=pull_flash(),
    )

@app.route("/users", endpoint="users")
@roles_required(ROLE_ADMIN)
def users():
    flash = pop_flash()
    conn = get_db()
    rows = conn.execute("SELECT username, role, is_active FROM users ORDER BY username").fetchall()
    conn.close()
    return render_template_string(
        USERS_HTML,
        users=rows,
        labels=ROLE_LABELS,
        current_username=current_user(),
        flash=flash,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

@app.route("/users/create", methods=["POST"], endpoint="user_create")
@roles_required(ROLE_ADMIN)
def user_create():
    u = (request.form.get("username") or "").strip()
    p = (request.form.get("password") or "").strip()
    role = (request.form.get("role") or "").strip().lower()

    if not u or not p:
        set_flash("Usuario y password son obligatorios.", "err")
        return redirect(url_for("users"))
    if role not in (ROLE_ADMIN, ROLE_RISK, ROLE_SALES):
        set_flash("Rol inválido.", "err")
        return redirect(url_for("users"))

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO users (username, pass_hash, role, is_active, created_at)
            VALUES (?, ?, ?, 1, ?)
        """, (u, generate_password_hash(p), role, now_str()))
        conn.commit()
        set_flash("Usuario creado.", "ok")
    except sqlite3.IntegrityError:
        set_flash("Ese usuario ya existe.", "err")
    finally:
        conn.close()

    return redirect(url_for("users"))

@app.route("/users/<string:username>/deactivate", methods=["POST"], endpoint="user_deactivate")
@roles_required(ROLE_ADMIN)
def user_deactivate(username: str):
    if username == current_user():
        set_flash("No podés desactivarte a vos mismo.", "err")
        return redirect(url_for("users"))
    conn = get_db()
    conn.execute("UPDATE users SET is_active=0 WHERE username=?", (username,))
    conn.commit()
    conn.close()
    set_flash("Usuario desactivado.", "ok")
    return redirect(url_for("users"))

@app.route("/users/<string:username>/reactivate", methods=["POST"], endpoint="user_reactivate")
@roles_required(ROLE_ADMIN)
def user_reactivate(username: str):
    conn = get_db()
    conn.execute("UPDATE users SET is_active=1 WHERE username=?", (username,))
    conn.commit()
    conn.close()
    set_flash("Usuario reactivado.", "ok")
    return redirect(url_for("users"))

# --------------------------- | Boot | ---------------------------

# --------------------------- | Routes: auth | ---------------------------

# --------------------------- | Alertas / boot restaurados | ---------------------------

ALERTS_HTML = '\n<!doctype html><html><head><meta charset="utf-8" />\n<title>Alertas diarias</title>\n<style>\n  *,*::before,*::after{ box-sizing:border-box; }\n  {{ toolbar_css|safe }}\n  \n.flash-ok{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(34,197,94,0.35);\n  background: rgba(34,197,94,0.10);\n  color:#dcfce7;\n}\n.flash-err{\n  margin-top:12px; padding:12px 14px; border-radius:12px;\n  border:1px solid rgba(244,63,94,0.35);\n  background: rgba(244,63,94,0.10);\n  color:#ffe4e6;\n}\n\n  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }\n  .page{ padding:92px 24px 24px 24px; }\n  .wrap{ max-width: 1280px; margin:0 auto; }\n  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }\n  .h{ font-size:18px; font-weight:800; margin-bottom:10px; }\n  .row{ display:flex; gap:10px; flex-wrap:wrap; margin:10px 0 14px; }\n  .stat{ flex:1; min-width:220px; background: rgba(0,0,0,0.18); border:1px solid rgba(255,255,255,0.10); border-radius:14px; padding:12px; }\n  .stat .k{ color:#cbd5e1; font-size:12px; }\n  .stat .v{ font-size:18px; font-weight:800; margin-top:4px; }\n  .row2{ display:flex; gap:10px; align-items:end; flex-wrap:wrap; }\n  label{ color:#cbd5e1; font-size:12px; display:block; margin-bottom:6px; }\n  select,input{\n    padding:10px; border-radius:12px;\n    border:1px solid rgba(255,255,255,0.15);\n    background: rgba(0,0,0,0.2); color:#fff;\n    min-width:220px;\n  }\n  .btn{\n    padding:10px 12px; border-radius:12px; cursor:pointer;\n    border:1px solid rgba(255,255,255,0.18);\n    background: rgba(255,255,255,0.08);\n    color:#fff; font-weight:700;\n  }\n  .btn.primary{ background:#2563eb; border-color:#2563eb; }\n  .btn.warn{ background:#7c3aed; border-color:#7c3aed; }\n  .btn:disabled{ opacity:0.6; cursor:not-allowed; }\n  .btn:hover{ filter:brightness(1.08); }\n  .muted{ color:#cbd5e1; font-size:12px; margin-top:10px; }\n  .sep{ height:1px; background: rgba(255,255,255,0.12); margin:14px 0; border:0; }\n  .grid{ display:grid; grid-template-columns: 1.2fr 1fr; gap:14px; }\n  .status-box{ padding:12px; border-radius:14px; border:1px solid rgba(255,255,255,0.12); background: rgba(0,0,0,0.18); }\n  .badge{ display:inline-block; padding:4px 8px; border-radius:999px; font-size:12px; font-weight:700; }\n  .b-ok{ background:#14532d; color:#bbf7d0; }\n  .b-run{ background:#1d4ed8; color:#dbeafe; }\n  .b-err{ background:#7f1d1d; color:#fecaca; }\n  table{ width:100%; border-collapse:collapse; margin-top:10px; font-size:13px; }\n  th,td{ padding:8px 10px; border-bottom:1px solid rgba(255,255,255,0.08); text-align:left; vertical-align:top; }\n  th{ color:#cbd5e1; font-weight:700; }\n  .mono{ font-family: Consolas, monospace; }\n  .small{ font-size:12px; color:#cbd5e1; }\n  @media (max-width: 980px){ .grid{ grid-template-columns:1fr; } }\n</style>\n</head><body>\n{{ toolbar_html|safe }}\n<div class="page"><div class="wrap">\n\n  <div class="card">\n    <div class="h">🚨 Cheques rechazados SIN FONDOS (Módulo en desarrollo)</div>\n    {% if flash %}<div class="flash {{ flash.kind }}">{{ flash.msg }}</div>{% endif %}\n\n    <div class="row">\n      <div class="stat"><div class="k">Firmantes monitoreables</div><div class="v">{{ stats.monitoreables }}</div></div>\n      <div class="stat"><div class="k">Eventos detectados</div><div class="v">{{ stats.total_events }}</div></div>\n      <div class="stat"><div class="k">Última corrida</div><div class="v">{{ stats.last_run or \'—\' }}</div></div>\n      <div class="stat"><div class="k">Estado</div><div class="v" id="statusBadgeWrap">{% if status.running %}<span class="badge b-run">Corriendo</span>{% elif status.last_rc not in (none, 0) %}<span class="badge b-err">Error</span>{% else %}<span class="badge b-ok">En espera</span>{% endif %}</div></div>\n    </div>\n\n    <div class="status-box" id="statusBox">\n      <div style="display:flex; justify-content:space-between; gap:10px; flex-wrap:wrap; align-items:center;">\n        <div><b id="statusTitle">{{ status.title }}</b></div>\n        <div class="small" id="statusTime">{{ status.time_text }}</div>\n      </div>\n      <div class="small" id="statusText" style="margin-top:8px; white-space:pre-wrap;">{{ status.text }}</div>\n    </div>\n\n    <hr class="sep" />\n\n    <div class="row2">\n      <div>\n        <label>Alcance</label>\n        <select id="scope">\n          <option value="all">Todos los firmantes activos</option>\n          <option value="one">Solo un CUIT</option>\n        </select>\n      </div>\n      <div>\n        <label>CUIT</label>\n        <input id="cuit" placeholder="Completar solo si elegís \'Solo un CUIT\'" />\n      </div>\n      <div>\n        <button class="btn primary" id="btnRun">▶ Ejecutar monitoreo con alertas</button>\n      </div>\n      <div>\n        <button class="btn warn" id="btnSync">🔄 Sincronizar sin alertas</button>\n      </div>\n    </div>\n\n    <div class="muted">\n      Script monitor: <span class="mono">{{ monitor_script }}</span>\n    </div>\n  </div>\n\n  <div style="height:14px"></div>\n\n  <div class="grid">\n    <div class="card">\n      <div class="h">🧾 Últimos eventos detectados</div>\n      <table>\n        <thead>\n          <tr><th>Detectado</th><th>Firmante</th><th>Fecha rechazo</th><th>Cheque</th><th>Monto</th><th>Pagado</th></tr>\n        </thead>\n        <tbody>\n        {% for e in recent_events %}\n          <tr>\n            <td class="small">{{ e.detected_at or \'—\' }}</td>\n            <td><div>{{ e.razon_social or \'—\' }}</div><div class="small mono">{{ e.cuit_digits }}</div></td>\n            <td>{{ e.fecha_rechazo or \'—\' }}</td>\n            <td class="mono">{{ e.nro_cheque or \'—\' }}</td>\n            <td>{{ e.monto_fmt }}</td>\n            <td>{{ \'Sí\' if e.pagado else \'No\' }}</td>\n          </tr>\n        {% else %}\n          <tr><td colspan="6" class="small">Todavía no hay eventos detectados.</td></tr>\n        {% endfor %}\n        </tbody>\n      </table>\n    </div>\n\n    <div class="card">\n      <div class="h">⏱ Últimas corridas</div>\n      <table>\n        <thead>\n          <tr><th>Inicio</th><th>Fin</th><th>Consultas</th><th>Nuevos</th><th>Sin Rechazos</th></tr>\n        </thead>\n        <tbody>\n        {% for r in recent_runs %}\n          <tr>\n            <td class="small">{{ r.started_at or \'—\' }}</td>\n            <td class="small">{{ r.ended_at or \'—\' }}</td>\n            <td>{{ r.total_cuits or 0 }}</td>\n            <td>{{ r.new_events or 0 }}</td>\n            <td>{{ r.errors or 0 }}</td>\n          </tr>\n        {% else %}\n          <tr><td colspan="5" class="small">Aún no hay corridas registradas.</td></tr>\n        {% endfor %}\n        </tbody>\n      </table>\n    </div>\n  </div>\n\n</div></div>\n\n<script>\n(function(){\n  const btnRun = document.getElementById(\'btnRun\');\n  const btnSync = document.getElementById(\'btnSync\');\n  const scope = document.getElementById(\'scope\');\n  const cuit = document.getElementById(\'cuit\');\n  let timer = null;\n\n  function setButtons(disabled){\n    if(btnRun) btnRun.disabled = disabled;\n    if(btnSync) btnSync.disabled = disabled;\n  }\n\n  async function startJob(withAlerts){\n    const url = withAlerts ? "{{ url_for(\'alerts_run_async\') }}" : "{{ url_for(\'alerts_sync_async\') }}";\n    const fd = new FormData();\n    fd.append(\'scope\', scope.value || \'all\');\n    fd.append(\'cuit\', cuit.value || \'\');\n    const res = await fetch(url, { method:\'POST\', body: fd });\n    const j = await res.json();\n    if(!res.ok || !j.ok){\n      alert((j && (j.error || j.message)) || \'No se pudo iniciar la corrida\');\n      return;\n    }\n    poll();\n  }\n\n  async function poll(){\n    const res = await fetch("{{ url_for(\'alerts_progress\') }}", { cache:\'no-store\' });\n    const j = await res.json();\n    const statusTitle = document.getElementById(\'statusTitle\');\n    const statusText = document.getElementById(\'statusText\');\n    const statusTime = document.getElementById(\'statusTime\');\n    const statusBadgeWrap = document.getElementById(\'statusBadgeWrap\');\n    if(statusTitle) statusTitle.textContent = j.title || \'Estado\';\n    if(statusText) statusText.textContent = j.text || \'\';\n    if(statusTime) statusTime.textContent = j.time_text || \'\';\n    if(statusBadgeWrap){\n      if(j.running){ statusBadgeWrap.innerHTML = \'<span class="badge b-run">Corriendo</span>\'; }\n      else if((j.last_rc || 0) !== 0){ statusBadgeWrap.innerHTML = \'<span class="badge b-err">Error</span>\'; }\n      else { statusBadgeWrap.innerHTML = \'<span class="badge b-ok">En espera</span>\'; }\n    }\n    setButtons(!!j.running);\n    if(j.running){\n      if(!timer){ timer = setInterval(poll, 2000); }\n    } else {\n      if(timer){ clearInterval(timer); timer = null; }\n      if(j.refresh_page){ window.location.reload(); }\n    }\n  }\n\n  if(btnRun) btnRun.addEventListener(\'click\', function(){ startJob(true).catch(err => alert(err)); });\n  if(btnSync) btnSync.addEventListener(\'click\', function(){ startJob(false).catch(err => alert(err)); });\n  poll();\n})();\n</script>\n</body></html>\n'

CHEQ_MONITOR_SCRIPT = (
    os.environ.get("CRM_CHEQ_MONITOR_SCRIPT")
    or (os.path.join(APP_DIR, "cheques_monitor_snapshot_slackfmt_patched.py")
        if os.path.exists(os.path.join(APP_DIR, "cheques_monitor_snapshot_slackfmt_patched.py"))
        else os.path.join(APP_DIR, "cheques_monitor_snapshot_slackfmt.py"))
)
_CHEQ_BG_LOCK = threading.Lock()
_CHEQ_BG_STATE = {
    "running": False,
    "scope": "",
    "scope_cuit": "",
    "alerts": False,
    "started_at": "",
    "ended_at": "",
    "last_rc": 0,
    "last_stdout": "",
    "last_stderr": "",
    "just_finished": False,
}

def _ensure_snapshot_tables():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cheq_monitor_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id INTEGER,
            cuit_digits TEXT NOT NULL,
            razon_social TEXT,
            event_id TEXT NOT NULL,
            causal TEXT,
            entidad TEXT,
            nro_cheque TEXT,
            fecha_rechazo TEXT,
            monto REAL,
            pagado INTEGER NOT NULL DEFAULT 0,
            payload_json TEXT,
            detected_at TEXT NOT NULL,
            notified INTEGER NOT NULL DEFAULT 0,
            UNIQUE(cuit_digits, event_id)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS cheq_monitor_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_type TEXT,
            scope TEXT,
            scope_cuit TEXT,
            started_at TEXT,
            ended_at TEXT,
            total_cuits INTEGER,
            new_events INTEGER,
            alerts_sent INTEGER,
            errors INTEGER,
            notes TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_cheq_monitor_events_detected ON cheq_monitor_events(detected_at DESC)")
    conn.commit()
    conn.close()

def _alerts_recent_events(limit: int = 80):
    _ensure_snapshot_tables()
    conn = get_db()
    rows = conn.execute(
        """
        SELECT detected_at, cuit_digits, razon_social, fecha_rechazo, nro_cheque, monto, pagado
        FROM cheq_monitor_events
        ORDER BY id DESC
        LIMIT ?
        """,
        (int(limit),),
    ).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        d["monto_fmt"] = fmt_ars(d.get("monto") or 0)
        out.append(d)
    return out

def _alerts_recent_runs(limit: int = 20):
    _ensure_snapshot_tables()
    conn = get_db()
    rows = conn.execute(
        """
        SELECT id, run_type, scope, scope_cuit, started_at, ended_at, total_cuits, new_events, alerts_sent, errors, notes
        FROM cheq_monitor_runs
        ORDER BY id DESC
        LIMIT ?
        """,
        (int(limit),),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def _alerts_stats():
    _ensure_snapshot_tables()
    s = {"monitoreables": 0, "total_events": 0, "last_run": "—"}
    try:
        s["monitoreables"] = len(_monitoreable_cuits())
    except Exception:
        s["monitoreables"] = 0
    conn = get_db()
    try:
        r = conn.execute("SELECT COUNT(1) FROM cheq_monitor_events").fetchone()
        s["total_events"] = int(r[0] or 0) if r else 0
        rr = conn.execute("SELECT started_at, ended_at FROM cheq_monitor_runs ORDER BY id DESC LIMIT 1").fetchone()
        if rr:
            s["last_run"] = (rr[1] or rr[0] or "—")
    finally:
        conn.close()
    return s

def _alerts_progress_payload():
    with _CHEQ_BG_LOCK:
        st = dict(_CHEQ_BG_STATE)
    if st.get("running"):
        return {
            "running": True,
            "last_rc": st.get("last_rc", 0),
            "title": "Corrida en ejecución",
            "text": f"Monitoreo {'con alertas' if st.get('alerts') else 'sin alertas'} en curso. Alcance: {st.get('scope') or 'all'} {st.get('scope_cuit') or ''}".strip(),
            "time_text": st.get("started_at") or "",
            "refresh_page": False,
        }
    rc = st.get("last_rc", 0)
    text = "Esperando una nueva corrida."
    if st.get("ended_at"):
        if rc == 0:
            text = (st.get("last_stdout") or "Corrida finalizada correctamente.")[-800:]
        else:
            text = ((st.get("last_stderr") or st.get("last_stdout") or "Corrida finalizada con error."))[-800:]
    payload = {
        "running": False,
        "last_rc": rc,
        "title": "Última corrida finalizada" if st.get("ended_at") else "Sin corridas en ejecución",
        "text": text,
        "time_text": st.get("ended_at") or st.get("started_at") or "",
        "refresh_page": bool(st.get("just_finished")),
    }
    with _CHEQ_BG_LOCK:
        _CHEQ_BG_STATE["just_finished"] = False
    return payload

def _launch_snapshot_job(*, scope: str, scope_cuit: str, alerts: bool):
    _ensure_snapshot_tables()
    with _CHEQ_BG_LOCK:
        if _CHEQ_BG_STATE.get("running"):
            return False, "Ya hay una corrida en ejecución."
        _CHEQ_BG_STATE.update({
            "running": True,
            "scope": scope,
            "scope_cuit": scope_cuit,
            "alerts": bool(alerts),
            "started_at": now_str(),
            "ended_at": "",
            "last_rc": 0,
            "last_stdout": "",
            "last_stderr": "",
            "just_finished": False,
        })

    def _worker():
        cmd = [sys.executable, CHEQ_MONITOR_SCRIPT, "--db", DB_PATH, "--base-app", os.path.abspath(__file__), "--scope", "one" if scope == "one" and scope_cuit else "all", "--alerts", "1" if alerts else "0"]
        if scope == "one" and scope_cuit:
            cmd.extend(["--cuit", cuit_digits(scope_cuit)])
        try:
            proc = subprocess.run(cmd, cwd=APP_DIR, capture_output=True, text=True)
            rc = int(proc.returncode or 0)
            out = (proc.stdout or "").strip()
            err = (proc.stderr or "").strip()
        except Exception as e:
            rc = 1
            out = ""
            err = str(e)
        with _CHEQ_BG_LOCK:
            _CHEQ_BG_STATE.update({
                "running": False,
                "ended_at": now_str(),
                "last_rc": rc,
                "last_stdout": out,
                "last_stderr": err,
                "just_finished": True,
            })

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    return True, "Corrida lanzada."

@app.route("/alerts", methods=["GET"], endpoint="alerts")
@roles_required(ROLE_ADMIN, ROLE_RISK)
def alerts():
    return render_template_string(
        ALERTS_HTML,
        toolbar_html=render_toolbar_html(),
        toolbar_css=TOOLBAR_CSS,
        flash=pull_flash(),
        stats=_alerts_stats(),
        status=_alerts_progress_payload(),
        recent_events=_alerts_recent_events(80),
        recent_runs=_alerts_recent_runs(20),
        monitor_script=CHEQ_MONITOR_SCRIPT,
    )

@app.route("/alerts/run", methods=["POST"])
@roles_required(ROLE_ADMIN, ROLE_RISK)
def alerts_run():
    scope = (request.form.get("scope") or "all").strip().lower()
    scope_cuit = cuit_digits(request.form.get("cuit") or "")
    if scope == "one" and not scope_cuit:
        set_flash("Debés informar un CUIT cuando el alcance es 'Solo un CUIT'.", "err")
        return redirect(url_for("alerts"))
    ok, msg = _launch_snapshot_job(scope=scope, scope_cuit=scope_cuit, alerts=True)
    set_flash(msg, "ok" if ok else "warn")
    return redirect(url_for("alerts"))

@app.route("/alerts/sync", methods=["POST"])
@roles_required(ROLE_ADMIN, ROLE_RISK)
def alerts_sync():
    scope = (request.form.get("scope") or "all").strip().lower()
    scope_cuit = cuit_digits(request.form.get("cuit") or "")
    if scope == "one" and not scope_cuit:
        set_flash("Debés informar un CUIT cuando el alcance es 'Solo un CUIT'.", "err")
        return redirect(url_for("alerts"))
    ok, msg = _launch_snapshot_job(scope=scope, scope_cuit=scope_cuit, alerts=False)
    set_flash(msg, "ok" if ok else "warn")
    return redirect(url_for("alerts"))

@app.route("/alerts/progress", methods=["GET"])
@roles_required(ROLE_ADMIN, ROLE_RISK)
def alerts_progress():
    return json.dumps(_alerts_progress_payload(), ensure_ascii=False), 200, {"Content-Type": "application/json; charset=utf-8"}

@app.route("/alerts/run_async", methods=["POST"])
@roles_required(ROLE_ADMIN, ROLE_RISK)
def alerts_run_async():
    scope = (request.form.get("scope") or "all").strip().lower()
    scope_cuit = cuit_digits(request.form.get("cuit") or "")
    if scope == "one" and not scope_cuit:
        return json.dumps({"ok": False, "error": "Informá un CUIT válido."}, ensure_ascii=False), 400, {"Content-Type": "application/json; charset=utf-8"}
    ok, msg = _launch_snapshot_job(scope=scope, scope_cuit=scope_cuit, alerts=True)
    code = 202 if ok else 409
    return json.dumps({"ok": ok, "message": msg}, ensure_ascii=False), code, {"Content-Type": "application/json; charset=utf-8"}

@app.route("/alerts/sync_async", methods=["POST"])
@roles_required(ROLE_ADMIN, ROLE_RISK)
def alerts_sync_async():
    scope = (request.form.get("scope") or "all").strip().lower()
    scope_cuit = cuit_digits(request.form.get("cuit") or "")
    if scope == "one" and not scope_cuit:
        return json.dumps({"ok": False, "error": "Informá un CUIT válido."}, ensure_ascii=False), 400, {"Content-Type": "application/json; charset=utf-8"}
    ok, msg = _launch_snapshot_job(scope=scope, scope_cuit=scope_cuit, alerts=False)
    code = 202 if ok else 409
    return json.dumps({"ok": ok, "message": msg}, ensure_ascii=False), code, {"Content-Type": "application/json; charset=utf-8"}

@app.route("/run_alerts")
def run_alerts():
    ok, msg = _launch_snapshot_job(scope="all", scope_cuit="", alerts=True)
    payload = {"ok": ok, "message": msg}
    return json.dumps(payload, ensure_ascii=False), (200 if ok else 409), {"Content-Type": "application/json; charset=utf-8"}


# --------------------------- | Extensión: Política crediticia | ---------------------------
TABLE_CREDIT_POLICY_CONFIG = "credit_policy_config"
TABLE_EXTRAORDINARY_LIMITS = "extraordinary_limits"

_orig_get_db = get_db

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("PRAGMA busy_timeout=30000;")
        conn.execute("PRAGMA foreign_keys=ON;")
    except Exception:
        pass
    return conn

_orig_init_db = init_db

def init_db():
    _orig_init_db()
    conn = get_db()
    cur = conn.cursor()
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_CREDIT_POLICY_CONFIG} (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            nyps REAL NOT NULL DEFAULT 0,
            pymes REAL NOT NULL DEFAULT 0,
            t1_t2_ar REAL NOT NULL DEFAULT 0,
            t2_directo REAL NOT NULL DEFAULT 0,
            garantizado REAL NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL DEFAULT '',
            updated_by TEXT NOT NULL DEFAULT ''
        )
    """)
    cur.execute(
        f"""
        INSERT OR IGNORE INTO {TABLE_CREDIT_POLICY_CONFIG}
        (id, nyps, pymes, t1_t2_ar, t2_directo, garantizado, updated_at, updated_by)
        VALUES (1, 0, 0, 0, 0, 0, ?, 'system')
        """,
        (now_str(),),
    )
    cur.execute(f"""
        CREATE TABLE IF NOT EXISTS {TABLE_EXTRAORDINARY_LIMITS} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cuit TEXT NOT NULL,
            cuit_digits TEXT NOT NULL,
            razon_social TEXT NOT NULL,
            segmento TEXT NOT NULL,
            limite REAL NOT NULL DEFAULT 0,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            updated_by TEXT NOT NULL
        )
    """)
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_extra_limits_cuit ON {TABLE_EXTRAORDINARY_LIMITS}(cuit_digits)")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_extra_limits_razon ON {TABLE_EXTRAORDINARY_LIMITS}(razon_social)")
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_extra_limits_segmento ON {TABLE_EXTRAORDINARY_LIMITS}(segmento)")
    conn.commit()
    conn.close()

_orig_render_toolbar_html = render_toolbar_html

def render_toolbar_html():
    links = [
        (url_for("home"), "🏠 Home"),
        (url_for("fp"), "👥 Firmantes Precalificados"),
        (url_for("groups"), "🏢 Grupos Económicos"),
        (safe_url_for("credit_policy"), "📘 Política crediticia"),
        (url_for("cartera"), "📁 Cartera"),
        (url_for("audit"), "🧾 Auditoría ABM"),
    ]
    if not is_sales():
        links.append((safe_url_for("alerts"), "🚨 Alertas diarias"))
    if is_admin():
        links.append((url_for("users"), "👤 Usuarios"))
    links.append((url_for("logout"), "Salir"))
    right = "".join([f'<a class="tb-btn" href="{href}">{label}</a>' for href, label in links if href])
    return f"""
<div class="toolbar">
  <div class="tb-left">
    <button class="tb-btn" type="button" onclick="history.back()">◀ Anterior</button>
    <button class="tb-btn" type="button" onclick="history.forward()">Siguiente ▶</button>
  </div>
  <div class="tb-right">{right}</div>
</div>
"""

def fmt_decimal_input(value):
    try:
        n = float(value or 0)
    except Exception:
        n = 0.0
    s = f"{n:,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")

def get_credit_policy_config():
    conn = get_db()
    row = conn.execute(f"SELECT * FROM {TABLE_CREDIT_POLICY_CONFIG} WHERE id=1").fetchone()
    conn.close()
    return row

def update_credit_policy_config(nyps, pymes, t1_t2_ar, t2_directo, garantizado, actor):
    conn = get_db()
    conn.execute(
        f"""
        UPDATE {TABLE_CREDIT_POLICY_CONFIG}
           SET nyps=?, pymes=?, t1_t2_ar=?, t2_directo=?, garantizado=?, updated_at=?, updated_by=?
         WHERE id=1
        """,
        (float(nyps or 0), float(pymes or 0), float(t1_t2_ar or 0), float(t2_directo or 0), float(garantizado or 0), now_str(), actor),
    )
    conn.commit()
    conn.close()

def list_extraordinary_limits(active_only=True):
    conn = get_db()
    sql = f"SELECT * FROM {TABLE_EXTRAORDINARY_LIMITS}"
    if active_only:
        sql += " WHERE is_active=1"
    sql += " ORDER BY razon_social COLLATE NOCASE, id DESC"
    rows = conn.execute(sql).fetchall()
    conn.close()
    return rows

CREDIT_POLICY_HTML = r'''
<!doctype html><html><head><meta charset="utf-8" />
<title>Política crediticia</title>
<style>
  *,*::before,*::after{ box-sizing:border-box; }
  {{ toolbar_css|safe }}
  body{ margin:0; font-family: Arial; background:#0b1220; color:#fff; }
  .page{ padding:92px 28px 28px 28px; }
  .wrap{ max-width: 1500px; margin:0 auto; }
  .card{ background: rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:18px; }
  .muted{ color:#cbd5e1; }
  .grid{ display:grid; grid-template-columns: repeat(5, minmax(180px, 1fr)); gap:12px; margin-top:16px; }
  .mini{ background: rgba(255,255,255,0.05); border:1px solid rgba(255,255,255,0.12); border-radius:16px; padding:14px; }
  .mini .k{ color:#93c5fd; font-size:12px; text-transform:uppercase; letter-spacing:.04em; }
  .mini .v{ margin-top:8px; font-size:24px; font-weight:800; }
  .row{ display:flex; gap:16px; flex-wrap:wrap; margin-top:18px; }
  .col{ flex:1 1 480px; }
  label{ display:block; margin-top:10px; color:#cbd5e1; font-size:13px; }
  input, select{ width:100%; padding:12px; margin-top:6px; border-radius:12px; border:1px solid rgba(255,255,255,0.15); background: rgba(0,0,0,0.2); color:#fff; }
  button{ margin-top:14px; padding:12px 14px; border-radius:12px; cursor:pointer; border:none; background:#fff; color:#0b1220; font-weight:800; }
  .btn-small{ padding:8px 10px; border-radius:12px; cursor:pointer; font-size:13px; border:1px solid rgba(255,255,255,0.14); background: rgba(255,255,255,0.06); color:#fff; }
  .btn-danger{ border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.12); }
  table{ width:100%; border-collapse: collapse; margin-top: 14px; }
  th,td{ text-align:left; padding:10px; border-bottom:1px solid rgba(255,255,255,0.10); vertical-align: top; }
  th{ color:#dbeafe; font-size:13px; }
  thead input{ margin-top:8px; padding:8px 10px; border-radius:10px; }
  .flash-ok{ margin-top:12px; padding:12px 14px; border-radius:12px; border:1px solid rgba(34,197,94,0.35); background: rgba(34,197,94,0.10); color:#dcfce7; }
  .flash-err,.flash-warn{ margin-top:12px; padding:12px 14px; border-radius:12px; border:1px solid rgba(244,63,94,0.35); background: rgba(244,63,94,0.10); color:#ffe4e6; }
  @media (max-width:1200px){ .grid{ grid-template-columns: repeat(2, minmax(220px, 1fr)); } }
  @media (max-width:700px){ .grid{ grid-template-columns: 1fr; } }
</style></head><body>
{{ toolbar_html|safe }}
<div class="page"><div class="wrap">
  <div class="card">
    <h2 style="margin:0;">Política crediticia</h2>
    <div class="muted" style="margin-top:6px;">Límites de concentración individuales y límites extraordinarios vigentes.</div>
    {% if flash %}<div class="{{ 'flash-ok' if flash.kind=='ok' else 'flash-err' }}">{{ flash.msg }}</div>{% endif %}

    <div class="grid">
      <div class="mini"><div class="k">NYPS</div><div class="v">{{ fmt_ars(policy.nyps if policy else 0) }}</div></div>
      <div class="mini"><div class="k">PYMES</div><div class="v">{{ fmt_ars(policy.pymes if policy else 0) }}</div></div>
      <div class="mini"><div class="k">T1, T2 y AR</div><div class="v">{{ fmt_ars(policy.t1_t2_ar if policy else 0) }}</div></div>
      <div class="mini"><div class="k">T2 Directo</div><div class="v">{{ fmt_ars(policy.t2_directo if policy else 0) }}</div></div>
      <div class="mini"><div class="k">Garantizado</div><div class="v">{{ fmt_ars(policy.garantizado if policy else 0) }}</div></div>
    </div>

    {% if is_admin %}
    <div class="row">
      <div class="col card" style="margin-top:18px;">
        <h3 style="margin:0;">Editar límites generales</h3>
        <form method="post" action="{{ url_for('credit_policy_update_config') }}">
          <label>NYPS</label><input name="nyps" value="{{ fmt_decimal_input(policy.nyps if policy else 0) }}" />
          <label>PYMES</label><input name="pymes" value="{{ fmt_decimal_input(policy.pymes if policy else 0) }}" />
          <label>T1, T2 y AR</label><input name="t1_t2_ar" value="{{ fmt_decimal_input(policy.t1_t2_ar if policy else 0) }}" />
          <label>T2 Directo</label><input name="t2_directo" value="{{ fmt_decimal_input(policy.t2_directo if policy else 0) }}" />
          <label>Garantizado</label><input name="garantizado" value="{{ fmt_decimal_input(policy.garantizado if policy else 0) }}" />
          <button type="submit">Guardar límites</button>
        </form>
      </div>
      <div class="col card" style="margin-top:18px;">
        <h3 style="margin:0;">Nuevo límite extraordinario</h3>
        <form method="post" action="{{ url_for('credit_policy_create_extra') }}">
          <label>CUIT</label><input name="cuit" required />
          <label>Razón</label><input name="razon_social" required />
          <label>Segmento</label><input name="segmento" required />
          <label>Límite</label><input name="limite" required />
          <button type="submit">Agregar límite extraordinario</button>
        </form>
      </div>
    </div>
    {% endif %}

    <div style="display:flex; justify-content:space-between; align-items:end; gap:12px; margin-top:22px; flex-wrap:wrap;">
      <div>
        <h3 style="margin:0;">Límites extraordinarios</h3>
        <div class="muted" style="margin-top:6px;">Registros visibles para todos los usuarios. Mostrando <b id="rowCountExtra">0</b> filas.</div>
      </div>
    </div>

    <table id="extraTable">
      <thead>
        <tr>
          <th>CUIT<div><input class="colf" data-col="0" placeholder="Filtrar CUIT" /></div></th>
          <th>Razón<div><input class="colf" data-col="1" placeholder="Filtrar razón" /></div></th>
          <th>Segmento<div><input class="colf" data-col="2" placeholder="Filtrar segmento" /></div></th>
          <th>Límite</th>
          {% if is_admin %}<th>Acciones</th>{% endif %}
        </tr>
      </thead>
      <tbody>
      {% for r in rows %}
        <tr>
          <td>{{ r.cuit }}</td>
          <td>{{ r.razon_social }}</td>
          <td>{{ r.segmento }}</td>
          <td>{{ fmt_ars(r.limite or 0) }}</td>
          {% if is_admin %}
          <td>
            <form method="post" action="{{ url_for('credit_policy_update_extra', row_id=r.id) }}" style="display:grid; grid-template-columns: 1.1fr 1.5fr 1fr 1fr auto; gap:8px; align-items:center;">
              <input name="cuit" value="{{ r.cuit }}" required />
              <input name="razon_social" value="{{ r.razon_social }}" required />
              <input name="segmento" value="{{ r.segmento }}" required />
              <input name="limite" value="{{ fmt_decimal_input(r.limite or 0) }}" required />
              <button class="btn-small" type="submit">Guardar</button>
            </form>
            <form method="post" action="{{ url_for('credit_policy_delete_extra', row_id=r.id) }}" style="display:inline;" onsubmit="return confirm('¿Desactivar este límite extraordinario?');">
              <button class="btn-small btn-danger" type="submit">Desactivar</button>
            </form>
          </td>
          {% endif %}
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
</div></div>
<script>
(function(){
  const table = document.getElementById('extraTable');
  if(!table) return;
  const filters = table.querySelectorAll('thead input.colf');
  const rows = Array.from(table.querySelectorAll('tbody tr'));
  const count = document.getElementById('rowCountExtra');
  function apply(){
    let visible = 0;
    rows.forEach(tr => {
      const tds = Array.from(tr.querySelectorAll('td'));
      if(!tds.length) return;
      let ok = true;
      filters.forEach(inp => {
        const idx = parseInt(inp.getAttribute('data-col') || '0', 10);
        const q = (inp.value || '').trim().toLowerCase();
        if(!q) return;
        const val = (tds[idx]?.innerText || '').trim().toLowerCase();
        if(!val.includes(q)) ok = false;
      });
      tr.style.display = ok ? '' : 'none';
      if(ok) visible += 1;
    });
    if(count) count.textContent = String(visible);
  }
  filters.forEach(inp => inp.addEventListener('input', apply));
  apply();
})();
</script>
</body></html>
'''

@app.route("/politica-crediticia", methods=["GET"], endpoint="credit_policy")
@login_required
def credit_policy():
    flash = pop_flash()
    return render_template_string(
        CREDIT_POLICY_HTML,
        policy=get_credit_policy_config(),
        rows=list_extraordinary_limits(active_only=True),
        fmt_ars=fmt_ars,
        fmt_decimal_input=fmt_decimal_input,
        is_admin=is_admin(),
        flash=flash,
        toolbar_css=TOOLBAR_CSS,
        toolbar_html=render_toolbar_html(),
    )

@app.route("/politica-crediticia/config", methods=["POST"], endpoint="credit_policy_update_config")
@roles_required(ROLE_ADMIN)
def credit_policy_update_config_route():
    vals = {
        "nyps": float(parse_es_number(request.form.get("nyps")) or 0),
        "pymes": float(parse_es_number(request.form.get("pymes")) or 0),
        "t1_t2_ar": float(parse_es_number(request.form.get("t1_t2_ar")) or 0),
        "t2_directo": float(parse_es_number(request.form.get("t2_directo")) or 0),
        "garantizado": float(parse_es_number(request.form.get("garantizado")) or 0),
    }
    update_credit_policy_config(vals["nyps"], vals["pymes"], vals["t1_t2_ar"], vals["t2_directo"], vals["garantizado"], current_user())
    set_flash("Política crediticia actualizada.", "ok")
    return redirect(url_for("credit_policy"))

@app.route("/politica-crediticia/extra/create", methods=["POST"], endpoint="credit_policy_create_extra")
@roles_required(ROLE_ADMIN)
def credit_policy_create_extra():
    cuit = (request.form.get("cuit") or "").strip()
    razon = (request.form.get("razon_social") or "").strip()
    segmento = (request.form.get("segmento") or "").strip()
    limite = float(parse_es_number(request.form.get("limite")) or 0)
    cd = cuit_digits(cuit)
    if not cd or not razon or not segmento:
        set_flash("CUIT, Razón y Segmento son obligatorios.", "err")
        return redirect(url_for("credit_policy"))
    conn = get_db()
    conn.execute(
        f"""
        INSERT INTO {TABLE_EXTRAORDINARY_LIMITS}
        (cuit, cuit_digits, razon_social, segmento, limite, is_active, created_at, updated_at, updated_by)
        VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
        """,
        (cuit, cd, razon, segmento, limite, now_str(), now_str(), current_user()),
    )
    conn.commit()
    conn.close()
    set_flash("Límite extraordinario creado.", "ok")
    return redirect(url_for("credit_policy"))

@app.route("/politica-crediticia/extra/<int:row_id>/update", methods=["POST"], endpoint="credit_policy_update_extra")
@roles_required(ROLE_ADMIN)
def credit_policy_update_extra(row_id: int):
    cuit = (request.form.get("cuit") or "").strip()
    razon = (request.form.get("razon_social") or "").strip()
    segmento = (request.form.get("segmento") or "").strip()
    limite = float(parse_es_number(request.form.get("limite")) or 0)
    cd = cuit_digits(cuit)
    if not cd or not razon or not segmento:
        set_flash("CUIT, Razón y Segmento son obligatorios.", "err")
        return redirect(url_for("credit_policy"))
    conn = get_db()
    conn.execute(
        f"""
        UPDATE {TABLE_EXTRAORDINARY_LIMITS}
           SET cuit=?, cuit_digits=?, razon_social=?, segmento=?, limite=?, updated_at=?, updated_by=?
         WHERE id=?
        """,
        (cuit, cd, razon, segmento, limite, now_str(), current_user(), int(row_id)),
    )
    conn.commit()
    conn.close()
    set_flash("Límite extraordinario actualizado.", "ok")
    return redirect(url_for("credit_policy"))

@app.route("/politica-crediticia/extra/<int:row_id>/delete", methods=["POST"], endpoint="credit_policy_delete_extra")
@roles_required(ROLE_ADMIN)
def credit_policy_delete_extra(row_id: int):
    conn = get_db()
    conn.execute(
        f"UPDATE {TABLE_EXTRAORDINARY_LIMITS} SET is_active=0, updated_at=?, updated_by=? WHERE id=?",
        (now_str(), current_user(), int(row_id)),
    )
    conn.commit()
    conn.close()
    set_flash("Límite extraordinario desactivado.", "ok")
    return redirect(url_for("credit_policy"))

from apscheduler.schedulers.background import BackgroundScheduler

scheduler = BackgroundScheduler()

scheduler.add_job(
    func=run_alerts,
    trigger="interval",
    minutes=15
)

scheduler.start()

if __name__ == "__main__":
    init_db()
# CLI jobs de alertas deshabilitados en el base app: usar wrapper/worker externo.
    app.run(host="0.0.0.0", port=5000, debug=False)



from werkzeug.security import generate_password_hash

def create_hardcoded_admin():
    conn = get_db_connection()
    try:
        user = conn.execute(
            "SELECT id FROM users WHERE username = ?",
            ("admin",)
        ).fetchone()

        if not user:
            conn.execute(
                """
                INSERT INTO users (username, pass_hash, role, is_active, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                ("admin", generate_password_hash("admin123"), "ADMIN", 1, now_str())
            )
            conn.commit()
            print("Admin creado: admin")
    finally:
        conn.close()

init_db()
create_hardcoded_admin()
