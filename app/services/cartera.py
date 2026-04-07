"""
Meridiano CRM — Cartera service
Parse cartera.xlsx and Facturas.xlsx with mtime-based caching.
"""

import os
from app.config import CARTERA_XLSX, FACTURAS_XLSX
from app.utils import cuit_digits, cuit_digits_excel, normalize_header, parse_es_number

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ── Cache structures ───────────────────────────────────────

_cartera_cache = {"mtime": None, "headers": [], "rows": [], "agg": {}, "status": "OK"}
_facturas_cache = {"mtime": None, "headers": [], "rows": [], "agg": {}, "status": "OK"}


# ── Cartera (cheques) ──────────────────────────────────────

def load_cartera() -> tuple[list, list, dict, str]:
    """Returns (headers, rows, agg, status).
    agg = {cuit_digits: {'3ros': sum, 'propio': sum}}
    """
    global _cartera_cache

    if not os.path.exists(CARTERA_XLSX):
        return [], [], {}, "No se encontró cartera.xlsx."
    if openpyxl is None:
        return [], [], {}, "Falta openpyxl."

    mtime = os.path.getmtime(CARTERA_XLSX)
    if _cartera_cache["mtime"] == mtime:
        return _cartera_cache["headers"], _cartera_cache["rows"], _cartera_cache["agg"], _cartera_cache["status"]

    try:
        wb = openpyxl.load_workbook(CARTERA_XLSX, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return [], [], {}, "cartera.xlsx está vacío."

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        headers_norm = [normalize_header(h) for h in headers]

        idx_cuit = next((i for i, h in enumerate(headers_norm) if h.startswith("cuit") and "2" not in h), None)
        idx_cuit2 = next((i for i, h in enumerate(headers_norm) if "cuit" in h and "2" in h), None)
        idx_importe = next((i for i, h in enumerate(headers_norm) if "importe" in h or "monto" in h), None)

        rows, agg = [], {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None or str(v).strip() == "" for v in r):
                continue
            row = list(r)
            if len(row) < len(headers):
                row += [None] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            rows.append(row)

            if idx_cuit is None or idx_importe is None:
                continue
            c1 = cuit_digits(row[idx_cuit])
            if not c1:
                continue
            c2 = cuit_digits(row[idx_cuit2]) if idx_cuit2 is not None else ""
            imp = float(parse_es_number(row[idx_importe]) or 0)

            agg.setdefault(c1, {"3ros": 0.0, "propio": 0.0})
            if c2 and c2 != c1:
                agg[c1]["3ros"] += imp
            else:
                agg[c1]["propio"] += imp

        wb.close()
        status = "OK" if (idx_cuit is not None and idx_importe is not None) else "Faltan columnas (CUIT/Importe)."
        _cartera_cache = {"mtime": mtime, "headers": headers, "rows": rows, "agg": agg, "status": status}
        return headers, rows, agg, status
    except Exception as e:
        return [], [], {}, f"Error leyendo cartera.xlsx: {e}"


# ── Facturas ───────────────────────────────────────────────

def load_facturas() -> tuple[list, list, dict, str]:
    """Returns (headers, rows, agg, status).
    agg = {cuit_digits: sum_importe}
    """
    global _facturas_cache

    if not os.path.exists(FACTURAS_XLSX):
        return [], [], {}, "No se encontró Facturas.xlsx."
    if openpyxl is None:
        return [], [], {}, "Falta openpyxl."

    mtime = os.path.getmtime(FACTURAS_XLSX)
    if _facturas_cache["mtime"] == mtime:
        return _facturas_cache["headers"], _facturas_cache["rows"], _facturas_cache["agg"], _facturas_cache["status"]

    try:
        wb = openpyxl.load_workbook(FACTURAS_XLSX, data_only=True, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            return [], [], {}, "Facturas.xlsx está vacío."

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        headers_norm = [normalize_header(h) for h in headers]

        idx_firmante = next((i for i, h in enumerate(headers_norm) if h in ("firmante", "cuit", "cuit firmante")), None)
        idx_importe = next((i for i, h in enumerate(headers_norm) if h in ("importe", "monto", "importe real")), None)

        rows, agg = [], {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None or str(v).strip() == "" for v in r):
                continue
            row = list(r)
            if len(row) < len(headers):
                row += [None] * (len(headers) - len(row))
            elif len(row) > len(headers):
                row = row[:len(headers)]
            rows.append(row)

            if idx_firmante is None or idx_importe is None:
                continue
            cd = cuit_digits_excel(row[idx_firmante])
            if not cd:
                continue
            imp = float(parse_es_number(row[idx_importe]) or 0)
            agg[cd] = float(agg.get(cd, 0)) + imp

        wb.close()
        status = "OK" if (idx_firmante is not None and idx_importe is not None) else "Faltan columnas (Firmante/Importe)."
        _facturas_cache.update({"mtime": mtime, "headers": headers, "rows": rows, "agg": agg, "status": status})
        return headers, rows, agg, status
    except Exception as e:
        return [], [], {}, f"Error leyendo Facturas.xlsx: {e}"


# ── Helpers for monitor ────────────────────────────────────

def cartera_al_dia_map() -> dict[str, float]:
    """Sum of positive importes per CUIT from cartera.xlsx (main CUIT only)."""
    headers, rows, _, _ = load_cartera()
    if not headers or not rows:
        return {}
    headers_norm = [normalize_header(h) for h in headers]
    idx_cuit = next((i for i, h in enumerate(headers_norm) if h == "cuit"), None)
    idx_imp = next((i for i, h in enumerate(headers_norm) if "importe" in h or "monto" in h), None)
    if idx_cuit is None or idx_imp is None:
        return {}

    out: dict[str, float] = {}
    for r in rows:
        cd = cuit_digits(r[idx_cuit])
        if not cd:
            continue
        imp = float(parse_es_number(r[idx_imp]) or 0)
        if imp > 0:
            out[cd] = out.get(cd, 0) + imp
    return out


def cartera_clientes_map() -> dict[str, set[str]]:
    """Mapea firmante principal -> conjunto de CUIT2 asociados en cartera.
    Si una fila no tiene CUIT2 válido o es el mismo CUIT, se representa con cadena vacía
    para indicar cartera propia / sin cliente identificable.
    """
    headers, rows, _, _ = load_cartera()
    if not headers or not rows:
        return {}
    headers_norm = [normalize_header(h) for h in headers]
    idx_cuit = next((i for i, h in enumerate(headers_norm) if h == "cuit"), None)
    idx_cuit2 = next((i for i, h in enumerate(headers_norm) if "cuit" in h and "2" in h), None)
    idx_imp = next((i for i, h in enumerate(headers_norm) if "importe" in h or "monto" in h), None)
    if idx_cuit is None or idx_imp is None:
        return {}

    out: dict[str, set[str]] = {}
    for r in rows:
        cd = cuit_digits(r[idx_cuit])
        if not cd:
            continue
        imp = float(parse_es_number(r[idx_imp]) or 0)
        if imp <= 0:
            continue
        c2 = cuit_digits(r[idx_cuit2]) if idx_cuit2 is not None else ""
        if c2 == cd:
            c2 = ""
        out.setdefault(cd, set()).add(c2 or "")
    return out
